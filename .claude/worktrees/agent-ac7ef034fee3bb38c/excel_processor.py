import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

CATEGORY_RULES = [
    ('透明商标类', ['透明', '商标', '白色PET']),
    ('铝箔类', ['银', '铝箔', '电化铝箔']),
    ('接线类', ['接线']),
    ('能源能效类', ['能源', '能效']),
    ('标贴类', ['型号标贴', '机型标贴', '纸箱标贴', '不可移铜版纸', '不可移光粉纸', '指示标贴']),
    ('说明书类', ['说明书', '合格证', '保修卡', '清单', '附页', '手册', '用户', '书写纸', '参数页']),
    ('特光类', ['特光']),
]

class ExcelProcessor:
    # 字段名别名映射：标准名 -> 可能的别名列表
    FIELD_ALIASES = {
        '物料号': ['物料编码'],
        '销售订单': ['销售订单号'],
        '总缺料': ['总缺料（差异数部分标红）', '上版缺料（差异数部分标红）'],
    }

    def __init__(self):
        self.category_rules = CATEGORY_RULES

    def normalize_row(self, row):
        """将字段别名规范化为标准字段名"""
        normalized = dict(row)
        for standard_name, aliases in self.FIELD_ALIASES.items():
            if standard_name not in normalized:
                for alias in aliases:
                    if alias in normalized:
                        normalized[standard_name] = normalized[alias]
                        break
        return normalized

    def read_excel_with_color(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        data = []
        color_map = {}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_values = [cell.value for cell in row]
            if any(v is not None and str(v).strip() for v in row_values):
                row_dict = dict(zip(headers, row_values))
                data.append(self.normalize_row(row_dict))

                material = row_dict.get('物料号', '')
                if material:
                    cell = row[0]
                    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != '00000000':
                        color_map[material] = cell.fill.fgColor.rgb

        return headers, data, color_map

    def read_excel(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None and str(v).strip() for v in row):
                row_dict = dict(zip(headers, row))
                data.append(self.normalize_row(row_dict))
        return headers, data

    def make_match_key(self, row):
        material = row.get('物料号', '') or row.get('物料编码', '') or ''
        sales_order = str(row.get('销售订单', '') or row.get('销售订单号', '') or '').strip()
        sales_line = str(row.get('销售订单行号', '') or '').strip()
        internal_order = str(row.get('内需单号', '') or '').strip()
        return f"{material}|{sales_order}|{sales_line}|{internal_order}"

    def compare_and_get_diff(self, yesterday_data, today_data, shipment_dict=None):
        yesterday_dict = {}
        for row in yesterday_data:
            if row.get('物料号'):
                key = self.make_match_key(row)
                yesterday_dict[key] = row

        today_dict = {}
        for row in today_data:
            if row.get('物料号'):
                key = self.make_match_key(row)
                today_dict[key] = row

        diff_result = []

        for key, today_row in today_dict.items():
            today_total = today_row.get('总缺料', 0) or 0
            if key in yesterday_dict:
                yesterday_total = yesterday_dict[key].get('总缺料', 0) or 0
                item = dict(today_row)
                item['昨天总缺料'] = yesterday_total
                item['今天总缺料'] = today_total
                # 计算本次送货数量
                shipment_qty = (shipment_dict.get(key, 0) if shipment_dict else 0) or 0
                item['本次送货数量'] = shipment_qty
                # 剩余出货量：数值一致时扣减，不一致时保留原始值
                if yesterday_total == shipment_qty:
                    item['剩余出货量'] = 0
                else:
                    item['剩余出货量'] = yesterday_total
                # 变化量 = 今天总缺料 - 剩余出货量
                item['变化量'] = today_total - item['剩余出货量']
                diff_result.append(item)
            else:
                item = dict(today_row)
                item['昨天总缺料'] = 0
                item['今天总缺料'] = today_total
                item['本次送货数量'] = 0
                item['剩余出货量'] = 0
                item['变化量'] = today_total
                diff_result.append(item)

        for key, yesterday_row in yesterday_dict.items():
            if key not in today_dict:
                yesterday_total = yesterday_row.get('总缺料', 0) or 0
                item = dict(yesterday_row)
                item['昨天总缺料'] = yesterday_total
                item['今天总缺料'] = 0
                item['本次送货数量'] = 0
                item['剩余出货量'] = yesterday_total
                item['变化量'] = -yesterday_total
                diff_result.append(item)

        return diff_result

    def sort_diff_data(self, diff_data):
        def sort_key(item):
            change_value = item['变化量']
            is_red_change = 1 if change_value != item['今天总缺料'] else 0
            return (is_red_change, -change_value)

        return sorted(diff_data, key=sort_key)

    def classify_material(self, material_description):
        if not material_description:
            return '未分类'

        for category_name, keywords in self.category_rules:
            for keyword in keywords:
                if keyword in material_description:
                    return category_name
        return '其他'

    def export_diff_to_excel(self, diff_data, output_path, color_map=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "缺料差异"

        if not diff_data:
            wb.save(output_path)
            return output_path

        priority_headers = ['序号', '物料号', '供应商在途量', '昨天总缺料', '今天总缺料', '变化量', '剩余出货量', '本次送货数量', '物料描述', '供方',
                           '异常', '备注', '销售订单', '销售订单行号', '内需单号']

        all_fields = []
        for item in diff_data:
            for key in item.keys():
                if key not in all_fields:
                    all_fields.append(key)

        headers = []
        for h in priority_headers:
            if h in all_fields and h not in headers:
                headers.append(h)
        for field in all_fields:
            if field not in headers:
                headers.append(field)

        if '序号' not in headers:
            headers.insert(0, '序号')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        highlight_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        highlight_font = Font(color="FFFFFF", bold=True)
        red_font = Font(color="FF0000", bold=True)

        for row_idx, item in enumerate(diff_data, 1):
            col_idx = 1
            if '序号' in headers:
                ws.cell(row=row_idx + 1, column=col_idx, value=row_idx)
                col_idx += 1

            for header in headers[1:]:
                value = item.get(header, '')
                cell = ws.cell(row=row_idx + 1, column=col_idx, value=value)

                if header == '物料号':
                    if color_map and item['物料号'] in color_map:
                        color_code = color_map[item['物料号']]
                        if color_code and len(color_code) >= 6:
                            try:
                                cell.fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
                            except:
                                pass

                elif header == '昨天总缺料':
                    if item.get('昨天总缺料', 0) == 0:
                        cell.font = red_font

                elif header == '今天总缺料':
                    if item.get('今天总缺料', 0) == 0:
                        cell.fill = highlight_fill
                        cell.font = highlight_font

                elif header == '变化量':
                    change_value = item.get('变化量', 0)
                    if change_value != item.get('今天总缺料', 0):
                        cell.font = red_font

                elif header == '剩余出货量':
                    shipment_qty = item.get('本次送货数量', 0) or 0
                    remaining_qty = (item.get('今天总缺料', 0) or 0) - shipment_qty
                    cell.value = remaining_qty
                    if item.get('昨天总缺料', 0) == 0:
                        cell.font = red_font

                elif header == '供应商在途量':
                    transit_qty = value
                    if transit_qty != '' and transit_qty != 0 and transit_qty is not None:
                        cell.font = red_font

                elif header == '物料描述':
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                col_idx += 1

            for col in range(1, len(headers) + 1):
                if headers[col-1] != '物料描述':
                    ws.cell(row=row_idx + 1, column=col).alignment = Alignment(horizontal="center", vertical="center")

        column_widths = {
            '序号': 8,
            '物料号': 18,
            '昨天总缺料': 14,
            '今天总缺料': 14,
            '变化量': 10,
            '剩余出货量': 12,
            '本次送货数量': 14,
            '供应商在途量': 15,
            '物料描述': 60,
            '供方': 15,
            '异常': 8,
            '备注': 15,
            '销售订单': 15,
            '销售订单行号': 15,
            '内需单号': 15,
        }

        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            width = column_widths.get(header, 15)
            ws.column_dimensions[col_letter].width = width

        wb.save(output_path)
        return output_path

    def export_category_files(self, data, output_dir):
        category_dict = {}
        for row in data:
            category = self.classify_material(row.get('物料描述', ''))
            if category not in category_dict:
                category_dict[category] = []
            category_dict[category].append(row)

        output_files = []
        for category, items in category_dict.items():
            if items:
                safe_category = category.replace('/', '_').replace('\\', '_')
                output_file = f"{output_dir}/{safe_category}.xlsx"
                self.export_to_excel(items, output_file, category)
                output_files.append((category, output_file))

        return output_files

    def export_to_excel(self, data, output_path, sheet_name="Sheet1"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not data:
            wb.save(output_path)
            return output_path

        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row_data in enumerate(data, 1):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx + 1, column=col_idx, value=row_data.get(header, ''))

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        # 冻结第一行表头
        ws.freeze_panes = 'A2'
        # 添加筛选功能
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        wb.save(output_path)
        return output_path

    def export_with_sequence(self, data, output_path, columns_order):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "数据"

        for col, header in enumerate(columns_order, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row_data in enumerate(data, 1):
            for col_idx, col_name in enumerate(columns_order, 1):
                value = row_data.get(col_name, '')
                cell = ws.cell(row=row_idx + 1, column=col_idx, value=value)
                if col_name in ['物料号', '数量']:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for col in range(1, len(columns_order) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        wb.save(output_path)
        return output_path

    def import_excel_with_highlight(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        highlighted_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                highlighted_data.append(row)

        return highlighted_data

    def save_shipment_highlighted(self, data, output_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "出货记录"

        headers = ['物料号', '数量', '备注']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        for row_idx, row_data in enumerate(data, 1):
            ws.cell(row=row_idx + 1, column=1, value=row_data.get('物料号', ''))
            ws.cell(row=row_idx + 1, column=2, value=row_data.get('数量', ''))
            ws.cell(row=row_idx + 1, column=3, value=row_data.get('备注', ''))

            for col in [1, 2]:
                cell = ws.cell(row=row_idx + 1, column=col)
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        wb.save(output_path)
        return output_path