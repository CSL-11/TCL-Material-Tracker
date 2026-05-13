import os
import sys
from datetime import datetime
from PyQt5.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QListWidget, QStackedWidget, QLabel, QPushButton,
                             QLineEdit, QTextEdit, QTableWidget, QTableWidgetItem,
                             QFileDialog, QMessageBox, QHeaderView, QFrame,
                             QGridLayout, QListWidgetItem, QAbstractItemView,
                             QScrollArea, QGroupBox, QDialog, QMenuBar, QAction,
                             QApplication, QCheckBox, QSystemTrayIcon, QMenu)
from PyQt5.QtCore import Qt, QSize, QThread, pyqtSignal, QTimer
from PyQt5.QtGui import QIcon, QFont, QColor, QPalette

from database import DatabaseManager
from excel_processor import ExcelProcessor
from network_manager import network_manager
from version import __version__


def get_app_dir():
    """获取应用所在目录，兼容PyInstaller打包后的EXE"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


class ServerThread(QThread):
    """服务器后台运行线程"""
    started = pyqtSignal()
    stopped = pyqtSignal()
    error = pyqtSignal(str)

    def __init__(self, host='0.0.0.0', port=5000):
        super().__init__()
        self.host = host
        self.port = port
        self._is_running = False
        self._server = None

    def run(self):
        try:
            from flask import Flask, jsonify
            from flask_cors import CORS
            from server_db import ServerDatabase
            import sys
            import os
            import traceback

            # 第一步：确定程序目录
            if getattr(sys, 'frozen', False):
                app_dir = os.path.dirname(sys.executable)
            else:
                app_dir = get_app_dir()

            # 第二步：切换工作目录
            os.chdir(app_dir)

            # 修复中文Windows主机名GBK编码导致socket.getfqdn() UTF-8解码失败
            import socket
            _orig_getfqdn = socket.getfqdn
            def _safe_getfqdn(*args, **kwargs):
                try:
                    return _orig_getfqdn(*args, **kwargs)
                except UnicodeDecodeError:
                    return 'localhost'
            socket.getfqdn = _safe_getfqdn

            # 第三步：创建Flask应用（不依赖__name__避免PyInstaller路径解析问题）
            app = Flask('tcl_server',
                        template_folder=os.path.join(app_dir, 'templates'),
                        static_folder=os.path.join(app_dir, 'static'))
            CORS(app)

            # 第四步：初始化数据库（使用绝对路径）
            db_path = os.path.join(app_dir, 'tcl_server_data.db')
            db = ServerDatabase(db_path)

            # 复制server.py中的路由定义
            @app.route('/api/health', methods=['GET'])
            def health_check():
                server_name = db.load_config('server_name', 'TCL表格比对系统服务器')
                return jsonify({
                    'status': 'ok',
                    'server': 'TCL表格比对系统服务器',
                    'name': server_name,
                    'version': __version__,
                    'timestamp': datetime.now().isoformat()
                })

            @app.route('/api/stats', methods=['GET'])
            def get_statistics():
                stats = db.get_statistics()
                stats['online_clients'] = len(connected_clients)
                stats['clients'] = [{'id': k, 'last_active': v.isoformat()} for k, v in connected_clients.items()]
                return jsonify(stats)

            @app.route('/api/db/data', methods=['GET', 'POST', 'DELETE'])
            def handle_db_data():
                from flask import request
                if request.method == 'GET':
                    headers, data = db.load_db_data()
                    return jsonify({'success': True, 'headers': headers, 'data': data, 'count': len(data)})
                elif request.method == 'POST':
                    try:
                        data = request.json
                        headers = data.get('headers', [])
                        rows = data.get('data', [])
                        db.save_db_data(headers, rows)
                        return jsonify({'success': True, 'message': f'成功保存 {len(rows)} 条记录'})
                    except Exception as e:
                        return jsonify({'success': False, 'error': str(e)}), 500
                elif request.method == 'DELETE':
                    try:
                        data = request.json or {}
                        item_ids = data.get('ids')
                        db.delete_db_data(item_ids)
                        return jsonify({'success': True, 'message': '删除成功'})
                    except Exception as e:
                        return jsonify({'success': False, 'error': str(e)}), 500

            @app.route('/api/batch/data', methods=['GET', 'POST', 'DELETE'])
            def handle_batch_data():
                from flask import request
                if request.method == 'GET':
                    headers, data = db.load_batch_import_data()
                    return jsonify({'success': True, 'headers': headers, 'data': data, 'count': len(data)})
                elif request.method == 'POST':
                    try:
                        data = request.json
                        headers = data.get('headers', [])
                        rows = data.get('data', [])
                        db.save_batch_import_data(headers, rows)
                        return jsonify({'success': True, 'message': f'成功保存 {len(rows)} 条记录'})
                    except Exception as e:
                        return jsonify({'success': False, 'error': str(e)}), 500
                elif request.method == 'DELETE':
                    try:
                        data = request.json or {}
                        indices = data.get('indices')
                        db.delete_batch_import_items(indices)
                        return jsonify({'success': True, 'message': '删除成功'})
                    except Exception as e:
                        return jsonify({'success': False, 'error': str(e)}), 500

            @app.route('/api/config/<key>', methods=['GET', 'POST'])
            def handle_config(key):
                from flask import request
                if request.method == 'GET':
                    value = db.load_config(key)
                    return jsonify({'success': True, 'key': key, 'value': value})
                elif request.method == 'POST':
                    try:
                        value = request.json.get('value')
                        db.save_config(key, value)
                        return jsonify({'success': True, 'message': f'配置 [{key}] 已更新'})
                    except Exception as e:
                        return jsonify({'success': False, 'error': str(e)}), 500

            @app.route('/api/config/all', methods=['GET'])
            def get_all_config():
                config = db.load_all_config()
                return jsonify({'success': True, 'config': config})

            @app.route('/api/logs', methods=['GET'])
            def get_logs():
                from flask import request
                limit = request.args.get('limit', 100, type=int)
                logs = db.get_operation_logs(limit)
                return jsonify({'success': True, 'logs': logs, 'count': len(logs)})

            @app.route('/api/backup', methods=['POST'])
            def backup_data():
                backup_path = db.backup_database()
                return jsonify({'success': True, 'path': backup_path, 'message': '备份成功'})

            @app.route('/api/restore', methods=['POST'])
            def restore_data():
                from flask import request
                backup_path = request.json.get('path')
                db.restore_database(backup_path)
                return jsonify({'success': True, 'message': '恢复成功'})

            @app.route('/api/backups', methods=['GET'])
            def list_backups():
                backups = db.list_backups()
                return jsonify({'success': True, 'backups': backups})

            # 密码认证中间件
            connected_clients = {}

            @app.before_request
            def check_auth():
                from flask import request
                # 追踪在线客户端
                client_id = request.headers.get('X-Client-ID', 'unknown')
                connected_clients[client_id] = datetime.now()
                # 清理超过60秒的客户端
                now = datetime.now()
                expired = [k for k, v in connected_clients.items() if (now - v).seconds > 60]
                for k in expired:
                    del connected_clients[k]
                # /api/health 无需认证
                if request.path == '/api/health':
                    return
                # 检查密码
                server_password = db.load_config('server_password', '')
                if server_password:
                    auth_token = request.headers.get('X-Auth-Token', '')
                    if auth_token != server_password:
                        return jsonify({'success': False, 'error': '认证失败，请输入正确的连接密码'}), 401

            self._is_running = True
            self.started.emit()
            app.run(host=self.host, port=self.port, debug=False, use_reloader=False)

        except Exception as e:
            # 将完整错误堆栈写入日志文件，方便定位问题
            try:
                log_path = os.path.join(
                    os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__)),
                    'server_error.log'
                )
                with open(log_path, 'w', encoding='utf-8') as f:
                    f.write(f"[{datetime.now().isoformat()}] 服务器启动错误:\n")
                    f.write(traceback.format_exc())
                    f.write(f"\nsys.frozen: {getattr(sys, 'frozen', False)}\n")
                    f.write(f"sys.executable: {sys.executable}\n")
                    f.write(f"os.getcwd(): {os.getcwd()}\n")
            except Exception:
                pass
            self.error.emit(str(e))
        finally:
            self._is_running = False
            self.stopped.emit()

    def stop(self):
        self._is_running = False


class ServerSettingsDialog(QDialog):
    """服务器设置对话框"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.server_thread = None
        self.is_server_running = False
        self.search_thread = None
        self.found_servers = []

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("服务器设置")
        self.setMinimumSize(500, 550)
        self.setModal(True)

        layout = QVBoxLayout()
        layout.setSpacing(10)

        # ========== 网络模式设置 ==========
        mode_group = QGroupBox("网络模式")
        mode_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 1px solid #D1D5DB;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
        """)
        mode_layout = QVBoxLayout()

        # 模式选择
        mode_btn_layout = QHBoxLayout()
        self.local_mode_btn = QPushButton("本地模式")
        self.local_mode_btn.setCheckable(True)
        self.local_mode_btn.setStyleSheet("""
            QPushButton {
                background-color: #10B981;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 20px;
                font-weight: bold;
            }
            QPushButton:checked {
                background-color: #059669;
                font-weight: bold;
            }
        """)
        self.server_mode_btn = QPushButton("服务器模式")
        self.server_mode_btn.setCheckable(True)
        self.server_mode_btn.setStyleSheet("""
            QPushButton {
                background-color: #9CA3AF;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 20px;
                font-weight: bold;
            }
            QPushButton:checked {
                background-color: #3B82F6;
                font-weight: bold;
            }
        """)
        self.local_mode_btn.clicked.connect(lambda: self.set_network_mode(False))
        self.server_mode_btn.clicked.connect(lambda: self.set_network_mode(True))

        mode_btn_layout.addWidget(self.local_mode_btn)
        mode_btn_layout.addWidget(self.server_mode_btn)
        mode_btn_layout.addStretch()
        mode_layout.addLayout(mode_btn_layout)

        # 服务器地址输入
        url_layout = QHBoxLayout()
        url_label = QLabel("服务器地址:")
        url_label.setFixedWidth(70)
        self.server_url_edit = QLineEdit()
        self.server_url_edit.setPlaceholderText("http://服务器IP:5000")
        self.server_url_edit.setStyleSheet("""
            QLineEdit {
                padding: 5px;
                border: 1px solid #D1D5DB;
                border-radius: 3px;
            }
        """)
        self.test_connection_btn = QPushButton("测试连接")
        self.test_connection_btn.setFixedWidth(80)
        self.test_connection_btn.setStyleSheet("""
            QPushButton {
                background-color: #F59E0B;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px 10px;
            }
            QPushButton:hover {
                background-color: #D97706;
            }
        """)
        self.test_connection_btn.clicked.connect(self.test_connection)

        url_layout.addWidget(url_label)
        url_layout.addWidget(self.server_url_edit)
        url_layout.addWidget(self.test_connection_btn)
        mode_layout.addLayout(url_layout)

        # 连接状态
        self.connection_status = QLabel("● 本地模式")
        self.connection_status.setStyleSheet("color: #10B981; font-size: 12px; font-weight: bold;")
        mode_layout.addWidget(self.connection_status)

        # 连接密码（客户端模式）
        password_layout = QHBoxLayout()
        password_label = QLabel("连接密码:")
        password_label.setFixedWidth(70)
        self.client_password_edit = QLineEdit()
        self.client_password_edit.setPlaceholderText("无密码请留空")
        self.client_password_edit.setEchoMode(QLineEdit.Password)
        self.client_password_edit.setStyleSheet("""
            QLineEdit {
                padding: 5px;
                border: 1px solid #D1D5DB;
                border-radius: 3px;
            }
        """)
        password_layout.addWidget(password_label)
        password_layout.addWidget(self.client_password_edit)
        mode_layout.addLayout(password_layout)

        # 开机自动连接服务器（客户端模式）
        client_autostart_layout = QHBoxLayout()
        self.client_autostart_checkbox = QCheckBox("开机自动连接到此服务器")
        self.client_autostart_checkbox.setChecked(False)
        self.client_autostart_checkbox.stateChanged.connect(self._toggle_client_autostart)
        self.client_autostart_checkbox.setEnabled(False)  # 本地模式下禁用
        client_autostart_layout.addWidget(self.client_autostart_checkbox)
        client_autostart_layout.addStretch()
        mode_layout.addLayout(client_autostart_layout)

        mode_group.setLayout(mode_layout)
        layout.addWidget(mode_group)

        # ========== 本机服务器设置 ==========
        server_group = QGroupBox("本机服务器")
        server_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 1px solid #D1D5DB;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
        """)
        server_layout = QVBoxLayout()

        # 服务器名称
        name_layout = QHBoxLayout()
        name_label = QLabel("服务器名称:")
        name_label.setFixedWidth(70)
        self.server_name_edit = QLineEdit("TCL表格比对系统服务器")
        self.server_name_edit.setStyleSheet("""
            QLineEdit {
                padding: 5px;
                border: 1px solid #D1D5DB;
                border-radius: 3px;
            }
        """)
        name_layout.addWidget(name_label)
        name_layout.addWidget(self.server_name_edit)
        server_layout.addLayout(name_layout)

        # 服务器密码
        srv_password_layout = QHBoxLayout()
        srv_password_label = QLabel("连接密码:")
        srv_password_label.setFixedWidth(70)
        self.server_password_edit = QLineEdit()
        self.server_password_edit.setPlaceholderText("无密码请留空")
        self.server_password_edit.setEchoMode(QLineEdit.Password)
        self.server_password_edit.setStyleSheet("""
            QLineEdit {
                padding: 5px;
                border: 1px solid #D1D5DB;
                border-radius: 3px;
            }
        """)
        srv_password_layout.addWidget(srv_password_label)
        srv_password_layout.addWidget(self.server_password_edit)
        server_layout.addLayout(srv_password_layout)

        # 端口设置
        port_layout = QHBoxLayout()
        port_label = QLabel("监听端口:")
        port_label.setFixedWidth(70)
        self.port_edit = QLineEdit("5000")
        self.port_edit.setFixedWidth(80)
        self.port_edit.setStyleSheet("""
            QLineEdit {
                padding: 5px;
                border: 1px solid #D1D5DB;
                border-radius: 3px;
            }
        """)
        port_layout.addWidget(port_label)
        port_layout.addWidget(self.port_edit)
        port_layout.addStretch()

        # 启动/停止按钮
        self.start_btn = QPushButton("启动服务器")
        self.start_btn.setFixedWidth(100)
        self.start_btn.setStyleSheet("""
            QPushButton {
                background-color: #10B981;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #059669;
            }
        """)
        self.start_btn.clicked.connect(self.toggle_server)
        port_layout.addWidget(self.start_btn)

        server_layout.addLayout(port_layout)

        # 服务器状态
        self.status_label = QLabel("● 服务器未启动")
        self.status_label.setStyleSheet("color: #6B7280; font-size: 12px;")
        server_layout.addWidget(self.status_label)

        # 服务器信息
        self.info_label = QLabel("")
        self.info_label.setStyleSheet("color: #374151; font-size: 11px;")
        self.info_label.setWordWrap(True)
        server_layout.addWidget(self.info_label)

        # 数据管理按钮
        data_btn_layout = QHBoxLayout()
        self.backup_btn = QPushButton("备份数据")
        self.backup_btn.setFixedWidth(80)
        self.backup_btn.setStyleSheet("""
            QPushButton {
                background-color: #8B5CF6;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 6px 12px;
            }
            QPushButton:hover {
                background-color: #7C3AED;
            }
        """)
        self.backup_btn.clicked.connect(self.backup_data)

        self.restore_btn = QPushButton("恢复数据")
        self.restore_btn.setFixedWidth(80)
        self.restore_btn.setStyleSheet("""
            QPushButton {
                background-color: #F59E0B;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 6px 12px;
            }
            QPushButton:hover {
                background-color: #D97706;
            }
        """)
        self.restore_btn.clicked.connect(self.restore_data)

        data_btn_layout.addWidget(self.backup_btn)
        data_btn_layout.addWidget(self.restore_btn)
        data_btn_layout.addStretch()
        server_layout.addLayout(data_btn_layout)

        # 日志和在线客户端按钮
        info_btn_layout = QHBoxLayout()
        self.view_logs_btn = QPushButton("查看日志")
        self.view_logs_btn.setFixedWidth(80)
        self.view_logs_btn.setStyleSheet("""
            QPushButton {
                background-color: #3B82F6;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 6px 12px;
            }
            QPushButton:hover {
                background-color: #2563EB;
            }
        """)
        self.view_logs_btn.clicked.connect(self.view_logs)

        self.online_clients_label = QLabel("在线客户端: 0")
        self.online_clients_label.setStyleSheet("color: #374151; font-size: 11px;")

        info_btn_layout.addWidget(self.view_logs_btn)
        info_btn_layout.addWidget(self.online_clients_label)
        info_btn_layout.addStretch()
        server_layout.addLayout(info_btn_layout)

        # 开机自动启动
        autostart_layout = QHBoxLayout()
        self.autostart_checkbox = QCheckBox("开机自动启动服务器")
        self.autostart_checkbox.setChecked(self._check_autostart_task())
        self.autostart_checkbox.stateChanged.connect(self._toggle_autostart)
        autostart_layout.addWidget(self.autostart_checkbox)
        autostart_layout.addStretch()
        server_layout.addLayout(autostart_layout)

        server_group.setLayout(server_layout)
        layout.addWidget(server_group)

        # ========== 服务器搜索 ==========
        search_group = QGroupBox("搜索局域网服务器")
        search_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 1px solid #D1D5DB;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
        """)
        search_layout = QVBoxLayout()

        # 搜索按钮
        search_btn_layout = QHBoxLayout()
        self.search_btn = QPushButton("搜索服务器")
        self.search_btn.setFixedWidth(100)
        self.search_btn.setStyleSheet("""
            QPushButton {
                background-color: #3B82F6;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2563EB;
            }
        """)
        self.search_btn.clicked.connect(self.search_servers)
        search_btn_layout.addWidget(self.search_btn)
        search_btn_layout.addStretch()

        self.search_status = QLabel("")
        self.search_status.setStyleSheet("color: #6B7280; font-size: 11px;")
        search_btn_layout.addWidget(self.search_status)
        search_layout.addLayout(search_btn_layout)

        # 搜索结果表格
        self.server_table = QTableWidget()
        self.server_table.setColumnCount(4)
        self.server_table.setHorizontalHeaderLabels(["服务器名称", "服务器地址", "端口", "状态"])
        self.server_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.server_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.server_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.server_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.server_table.setSelectionMode(QTableWidget.SingleSelection)
        self.server_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.server_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #E5E7EB;
                border-radius: 3px;
                gridline-color: #F3F4F6;
            }
            QTableWidget::item:selected {
                background-color: #DBEAFE;
                color: #1E40AF;
            }
            QHeaderView::section {
                background-color: #F9FAFB;
                border: none;
                border-bottom: 1px solid #E5E7EB;
                padding: 5px;
                font-weight: bold;
            }
        """)
        search_layout.addWidget(self.server_table)

        # 使用选中服务器按钮
        self.use_server_btn = QPushButton("使用选中的服务器")
        self.use_server_btn.setFixedWidth(130)
        self.use_server_btn.setEnabled(False)
        self.use_server_btn.setStyleSheet("""
            QPushButton {
                background-color: #8B5CF6;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #7C3AED;
            }
            QPushButton:disabled {
                background-color: #D1D5DB;
                color: #9CA3AF;
            }
        """)
        self.use_server_btn.clicked.connect(self.use_selected_server)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_layout.addWidget(self.use_server_btn)
        search_layout.addLayout(btn_layout)

        search_group.setLayout(search_layout)
        layout.addWidget(search_group)

        # ========== 关闭按钮 ==========
        close_btn = QPushButton("关闭")
        close_btn.setFixedWidth(80)
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #6B7280;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 15px;
            }
            QPushButton:hover {
                background-color: #4B5563;
            }
        """)
        close_btn.clicked.connect(self.close)

        bottom_layout = QHBoxLayout()
        bottom_layout.addStretch()
        bottom_layout.addWidget(close_btn)
        layout.addLayout(bottom_layout)

        self.setLayout(layout)

        # 连接表格选择事件
        self.server_table.itemSelectionChanged.connect(self.on_server_selected)

        # 初始化状态
        self._init_mode_state()

    def _init_mode_state(self):
        """初始化网络模式状态"""
        import subprocess
        parent = self.parent()
        if parent and hasattr(parent, 'network'):
            network = parent.network
            if network.is_server_mode:
                self.server_mode_btn.setChecked(True)
                self.local_mode_btn.setChecked(False)
                self.server_url_edit.setText(network.server_url)
                self.connection_status.setText("● 服务器模式")
                self.connection_status.setStyleSheet("color: #3B82F6; font-size: 12px; font-weight: bold;")
            else:
                self.local_mode_btn.setChecked(True)
                self.server_mode_btn.setChecked(False)
                self.connection_status.setText("● 本地模式")
                self.connection_status.setStyleSheet("color: #10B981; font-size: 12px; font-weight: bold;")

            # 初始化客户端自启动状态
            task_exists = self._check_client_autostart_task()
            config_enabled = getattr(network, 'auto_connect', False)
            self.client_autostart_checkbox.setChecked(task_exists and config_enabled)
            self.client_autostart_checkbox.setEnabled(network.is_server_mode)
            # 自愈：任务存在但配置为False，清理孤立任务
            if task_exists and not config_enabled:
                subprocess.run('schtasks /delete /tn "TCL表格比对客户端" /f', shell=True, capture_output=True, text=True, timeout=10)

    def set_network_mode(self, is_server_mode):
        """设置网络模式"""
        if is_server_mode:
            server_url = self.server_url_edit.text().strip()
            if not server_url:
                QMessageBox.warning(self, "提示", "请输入服务器地址\n格式: http://服务器IP:5000")
                self.local_mode_btn.setChecked(True)
                self.server_mode_btn.setChecked(False)
                return

            # 测试连接
            self.test_connection_btn.setEnabled(False)
            self.connection_status.setText("● 连接中...")
            self.connection_status.setStyleSheet("color: #F59E0B; font-size: 12px;")
            QApplication.processEvents()

            parent = self.parent()
            if parent and hasattr(parent, 'network'):
                network = parent.network
                original_url = network.server_url
                original_mode = network.is_server_mode
                original_password = network.server_password

                password = self.client_password_edit.text().strip()
                network.set_server_mode(server_url, True, password)
                success, msg = network.test_connection()

                if success:
                    self.server_mode_btn.setChecked(True)
                    self.local_mode_btn.setChecked(False)
                    self.connection_status.setText("● 服务器模式")
                    self.connection_status.setStyleSheet("color: #3B82F6; font-size: 12px; font-weight: bold;")
                    self.client_autostart_checkbox.setEnabled(True)
                    # 更新主界面
                    if hasattr(parent, 'update_network_status'):
                        parent.update_network_status()
                else:
                    # 恢复原始状态
                    network.set_server_mode(original_url, original_mode, original_password)
                    self.local_mode_btn.setChecked(True)
                    self.server_mode_btn.setChecked(False)
                    self.connection_status.setText("● 连接失败")
                    self.connection_status.setStyleSheet("color: #EF4444; font-size: 12px;")
                    self.client_autostart_checkbox.setEnabled(False)
                    QMessageBox.warning(self, "连接失败", f"无法连接到服务器:\n{msg}")
            self.test_connection_btn.setEnabled(True)
        else:
            self.local_mode_btn.setChecked(True)
            self.server_mode_btn.setChecked(False)
            self.connection_status.setText("● 本地模式")
            self.connection_status.setStyleSheet("color: #10B981; font-size: 12px; font-weight: bold;")
            # 禁用并取消客户端自启动
            if self.client_autostart_checkbox.isChecked():
                self.client_autostart_checkbox.setChecked(False)
            self.client_autostart_checkbox.setEnabled(False)

            parent = self.parent()
            if parent and hasattr(parent, 'network'):
                parent.network.set_server_mode('', False)
                if hasattr(parent, 'update_network_status'):
                    parent.update_network_status()

    def test_connection(self):
        """测试服务器连接"""
        server_url = self.server_url_edit.text().strip()
        if not server_url:
            QMessageBox.warning(self, "提示", "请先输入服务器地址")
            return

        self.test_connection_btn.setEnabled(False)
        self.connection_status.setText("● 测试中...")
        self.connection_status.setStyleSheet("color: #F59E0B; font-size: 12px;")
        QApplication.processEvents()

        parent = self.parent()
        if parent and hasattr(parent, 'network'):
            network = parent.network
            original_url = network.server_url
            original_mode = network.is_server_mode

            network.set_server_mode(server_url, True)
            success, msg = network.test_connection()
            network.set_server_mode(original_url, original_mode)

            if success:
                self.connection_status.setText("● 连接成功")
                self.connection_status.setStyleSheet("color: #10B981; font-size: 12px; font-weight: bold;")
                QMessageBox.information(self, "连接成功", f"✓ {msg}\n\n可以切换到服务器模式使用")
            else:
                self.connection_status.setText("● 连接失败")
                self.connection_status.setStyleSheet("color: #EF4444; font-size: 12px;")
                QMessageBox.warning(self, "连接失败", msg)

        self.test_connection_btn.setEnabled(True)

    def toggle_server(self):
        """启动或停止服务器"""
        if self.is_server_running:
            self.stop_server()
        else:
            self.start_server()

    def start_server(self):
        """启动服务器"""
        try:
            port = int(self.port_edit.text().strip())
            if port < 1 or port > 65535:
                QMessageBox.warning(self, "提示", "端口号必须在1-65535之间")
                return
        except ValueError:
            QMessageBox.warning(self, "提示", "请输入有效的端口号")
            return

        # 检查flask是否可用
        try:
            import flask
            import flask_cors
        except ImportError:
            QMessageBox.warning(self, "缺少依赖", "服务器功能需要安装Flask和flask_cors\n\n请运行: pip install flask flask_cors")
            return

        # 检查server_db是否可用
        try:
            from server_db import ServerDatabase
        except ImportError:
            QMessageBox.warning(self, "缺少模块", "找不到server_db模块，请确保文件存在")
            return

        self.start_btn.setEnabled(False)
        self.start_btn.setText("启动中...")
        self.status_label.setText("● 启动中...")
        self.status_label.setStyleSheet("color: #F59E0B; font-size: 12px;")
        QApplication.processEvents()

        # 保存服务器名称和密码到数据库配置
        try:
            from server_db import ServerDatabase
            import sys
            if getattr(sys, 'frozen', False):
                app_dir = os.path.dirname(sys.executable)
            else:
                app_dir = os.path.dirname(os.path.abspath(__file__))
            db_path = os.path.join(app_dir, 'tcl_server_data.db')
            db = ServerDatabase(db_path)
            db.save_config('server_name', self.server_name_edit.text().strip() or 'TCL表格比对系统服务器')
            db.save_config('server_password', self.server_password_edit.text().strip())
            db.save_config('server_port', port)
        except Exception as e:
            print(f"[WARN] 保存服务器配置失败: {e}")

        # 创建并启动服务器线程
        self.server_thread = ServerThread(host='0.0.0.0', port=port)
        self.server_thread.started.connect(self._on_server_started)
        self.server_thread.stopped.connect(self._on_server_stopped)
        self.server_thread.error.connect(self._on_server_error)
        self.server_thread.start()

    def stop_server(self):
        """停止服务器"""
        if self.server_thread and self.server_thread.isRunning():
            self.start_btn.setEnabled(False)
            self.start_btn.setText("停止中...")
            self.status_label.setText("● 停止中...")
            self.status_label.setStyleSheet("color: #F59E0B; font-size: 12px;")
            QApplication.processEvents()

            self.server_thread.stop()
            self.server_thread.wait(3000)

            if self.server_thread.isRunning():
                self.server_thread.terminate()

            self._on_server_stopped()

    def _on_server_started(self):
        """服务器启动成功回调"""
        self.is_server_running = True
        port = self.port_edit.text().strip()
        self.start_btn.setText("停止服务器")
        self.start_btn.setStyleSheet("""
            QPushButton {
                background-color: #EF4444;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #DC2626;
            }
        """)
        self.start_btn.setEnabled(True)
        self.status_label.setText(f"● 服务器运行中")
        self.status_label.setStyleSheet("color: #10B981; font-size: 12px; font-weight: bold;")

        # 获取本机IP
        import socket
        try:
            hostname = socket.gethostname()
            local_ip = socket.gethostbyname(hostname)
        except:
            local_ip = "127.0.0.1"

        self.info_label.setText(f"地址: http://{local_ip}:{port}\n其他客户端可通过此地址连接")

        # 自动添加Windows防火墙规则，允许其他电脑连接
        self._add_firewall_rule(port)

        # 启动定时器更新在线客户端数量
        self._update_clients_timer = QTimer()
        self._update_clients_timer.timeout.connect(self._update_online_clients)
        self._update_clients_timer.start(5000)  # 每5秒更新
        self._update_online_clients()

    def _on_server_stopped(self):
        """服务器停止回调"""
        self.is_server_running = False
        # 停止在线客户端更新定时器
        if hasattr(self, '_update_clients_timer'):
            self._update_clients_timer.stop()
        self.online_clients_label.setText("在线客户端: 0")
        self.start_btn.setText("启动服务器")
        self.start_btn.setStyleSheet("""
            QPushButton {
                background-color: #10B981;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #059669;
            }
        """)
        self.start_btn.setEnabled(True)
        self.status_label.setText("● 服务器未启动")
        self.status_label.setStyleSheet("color: #6B7280; font-size: 12px;")
        self.info_label.setText("")

    def _on_server_error(self, error_msg):
        """服务器启动失败回调"""
        self.is_server_running = False
        self.start_btn.setText("启动服务器")
        self.start_btn.setStyleSheet("""
            QPushButton {
                background-color: #10B981;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #059669;
            }
        """)
        self.start_btn.setEnabled(True)
        self.status_label.setText("● 启动失败")
        self.status_label.setStyleSheet("color: #EF4444; font-size: 12px;")
        QMessageBox.warning(self, "服务器启动失败", f"错误信息:\n{error_msg}")

    def _add_firewall_rule(self, port):
        """尝试添加Windows防火墙规则，允许其他电脑连接"""
        import subprocess
        rule_name = f"TCL表格比对服务器端口{port}"
        try:
            # 先检查是否已存在同名规则
            result = subprocess.run(
                ['netsh', 'advfirewall', 'firewall', 'show', 'rule', f'name={rule_name}'],
                capture_output=True, text=True, timeout=5
            )
            if rule_name in (result.stdout or ''):
                return  # 规则已存在

            # 添加端口入站规则（比程序规则更可靠）
            result = subprocess.run(
                ['netsh', 'advfirewall', 'firewall', 'add', 'rule',
                 f'name={rule_name}', 'dir=in', 'action=allow', 'protocol=TCP',
                 f'localport={port}',
                 'enable=yes', 'description=TCL表格比对系统服务器端口'],
                capture_output=True, text=True, timeout=5
            )
            if result.returncode == 0:
                print(f"[OK] 已自动添加防火墙规则: {rule_name}")
            else:
                # 权限不足，提示手动添加
                self._show_firewall_guide(port)
        except Exception:
            self._show_firewall_guide(port)

    def _show_firewall_guide(self, port):
        """显示防火墙手动配置指南"""
        QMessageBox.information(self, "防火墙设置",
            f"如其他电脑无法连接，请手动添加防火墙规则：\n\n"
            f"方法一（推荐）：\n"
            f"  以管理员身份运行此程序，启动服务器时会自动添加规则\n\n"
            f"方法二（手动添加）：\n"
            f"  1. 按 Win+R，输入 wf.msc 回车\n"
            f"  2. 左侧点击「入站规则」\n"
            f"  3. 右侧点击「新建规则」\n"
            f"  4. 选择「端口」→ 下一步\n"
            f"  5. 选择 TCP，特定端口填: {port} → 下一步\n"
            f"  6. 选择「允许连接」→ 下一步 → 完成")

    def backup_data(self):
        """备份服务器数据"""
        try:
            import sys
            if getattr(sys, 'frozen', False):
                app_dir = os.path.dirname(sys.executable)
            else:
                app_dir = os.path.dirname(os.path.abspath(__file__))
            db_path = os.path.join(app_dir, 'tcl_server_data.db')
            from server_db import ServerDatabase
            db = ServerDatabase(db_path)
            backup_path = db.backup_database()
            QMessageBox.information(self, "备份成功", f"数据已备份到:\n{backup_path}")
        except Exception as e:
            QMessageBox.warning(self, "备份失败", f"备份数据时出错:\n{str(e)}")

    def restore_data(self):
        """恢复服务器数据"""
        import sys
        if getattr(sys, 'frozen', False):
            app_dir = os.path.dirname(sys.executable)
        else:
            app_dir = os.path.dirname(os.path.abspath(__file__))

        backup_dir = os.path.join(app_dir, 'backups')
        if not os.path.exists(backup_dir):
            QMessageBox.warning(self, "提示", "没有找到备份文件")
            return

        from PyQt5.QtWidgets import QFileDialog
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择备份文件", backup_dir, "数据库文件 (*.db)")
        if not file_path:
            return

        reply = QMessageBox.question(self, "确认恢复",
            f"恢复数据将覆盖当前所有数据，确定要继续吗？\n\n备份文件: {os.path.basename(file_path)}",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        try:
            db_path = os.path.join(app_dir, 'tcl_server_data.db')
            from server_db import ServerDatabase
            db = ServerDatabase(db_path)
            db.restore_database(file_path)
            QMessageBox.information(self, "恢复成功", "数据已恢复，请重新启动服务器以加载新数据。")
        except Exception as e:
            QMessageBox.warning(self, "恢复失败", f"恢复数据时出错:\n{str(e)}")

    def view_logs(self):
        """查看服务器操作日志"""
        dialog = QDialog(self)
        dialog.setWindowTitle("服务器操作日志")
        dialog.setMinimumSize(600, 400)
        dialog.setModal(True)

        layout = QVBoxLayout()

        # 日志表格
        log_table = QTableWidget()
        log_table.setColumnCount(4)
        log_table.setHorizontalHeaderLabels(["时间", "客户端ID", "操作", "详情"])
        log_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        log_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        log_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        log_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        log_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(log_table)

        # 按钮
        btn_layout = QHBoxLayout()
        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(lambda: self._load_logs(log_table))
        clear_btn = QPushButton("清空日志")
        clear_btn.clicked.connect(lambda: self._clear_logs(log_table))
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(dialog.close)
        btn_layout.addWidget(refresh_btn)
        btn_layout.addWidget(clear_btn)
        btn_layout.addStretch()
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)

        dialog.setLayout(layout)
        self._load_logs(log_table)
        dialog.exec_()

    def _load_logs(self, table):
        """加载日志数据"""
        try:
            import sys
            if getattr(sys, 'frozen', False):
                app_dir = os.path.dirname(sys.executable)
            else:
                app_dir = os.path.dirname(os.path.abspath(__file__))
            db_path = os.path.join(app_dir, 'tcl_server_data.db')
            from server_db import ServerDatabase
            db = ServerDatabase(db_path)
            logs = db.get_operation_logs(100)

            table.setRowCount(len(logs))
            for i, log in enumerate(logs):
                table.setItem(i, 0, QTableWidgetItem(str(log.get('timestamp', ''))))
                table.setItem(i, 1, QTableWidgetItem(str(log.get('client_id', ''))))
                table.setItem(i, 2, QTableWidgetItem(str(log.get('operation', ''))))
                table.setItem(i, 3, QTableWidgetItem(str(log.get('details', ''))))
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载日志失败:\n{str(e)}")

    def _clear_logs(self, table):
        """清空日志"""
        reply = QMessageBox.question(self, "确认", "确定要清空所有日志吗？",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        try:
            import sys
            if getattr(sys, 'frozen', False):
                app_dir = os.path.dirname(sys.executable)
            else:
                app_dir = os.path.dirname(os.path.abspath(__file__))
            db_path = os.path.join(app_dir, 'tcl_server_data.db')
            from server_db import ServerDatabase
            db = ServerDatabase(db_path)
            conn = db.get_connection()
            cursor = conn.cursor()
            cursor.execute('DELETE FROM operation_log')
            conn.commit()
            conn.close()
            table.setRowCount(0)
            QMessageBox.information(self, "完成", "日志已清空")
        except Exception as e:
            QMessageBox.warning(self, "错误", f"清空日志失败:\n{str(e)}")

    def _update_online_clients(self):
        """更新在线客户端数量"""
        try:
            import urllib.request
            port = self.port_edit.text().strip()
            url = f"http://127.0.0.1:{port}/api/stats"
            req = urllib.request.Request(url)
            with urllib.request.urlopen(req, timeout=2) as response:
                import json
                data = json.loads(response.read().decode('utf-8'))
                count = data.get('online_clients', 0)
                self.online_clients_label.setText(f"在线客户端: {count}")
        except Exception:
            pass

    def _check_autostart_task(self):
        """检查是否已设置开机自动启动"""
        import subprocess
        try:
            result = subprocess.run(
                ['schtasks', '/query', '/tn', 'TCL表格比对服务器'],
                capture_output=True, text=True, timeout=5
            )
            return result.returncode == 0
        except Exception:
            return False

    def _toggle_autostart(self, state):
        """切换开机自动启动"""
        import subprocess
        task_name = 'TCL表格比对服务器'
        try:
            if state == 2:  # Qt.Checked
                # 获取程序路径
                import sys
                if getattr(sys, 'frozen', False):
                    exe_path = sys.executable
                else:
                    exe_path = os.path.abspath(sys.argv[0])
                # 创建计划任务
                cmd = f'schtasks /create /tn "{task_name}" /tr "\\"{exe_path}\\" --server" /sc onlogon /f'
                result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=10)
                if result.returncode == 0:
                    # 互斥：删除客户端自启动任务
                    subprocess.run('schtasks /delete /tn "TCL表格比对客户端" /f', shell=True, capture_output=True, text=True, timeout=10)
                    if hasattr(self, 'client_autostart_checkbox'):
                        self.client_autostart_checkbox.setChecked(False)
                    QMessageBox.information(self, "设置成功", "已设置开机自动启动服务器")
                else:
                    QMessageBox.warning(self, "设置失败", f"设置开机自启失败:\n{result.stderr}")
                    self.autostart_checkbox.setChecked(False)
            else:
                # 删除计划任务
                cmd = f'schtasks /delete /tn "{task_name}" /f'
                result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=10)
                if result.returncode == 0:
                    QMessageBox.information(self, "取消成功", "已取消开机自动启动")
                else:
                    QMessageBox.warning(self, "取消失败", f"取消开机自启失败:\n{result.stderr}")
                    self.autostart_checkbox.setChecked(True)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"操作失败:\n{str(e)}")

    def _check_client_autostart_task(self):
        """检查是否已设置客户端开机自动连接"""
        import subprocess
        try:
            result = subprocess.run(
                ['schtasks', '/query', '/tn', 'TCL表格比对客户端'],
                capture_output=True, text=True, timeout=5
            )
            return result.returncode == 0
        except Exception:
            return False

    def _toggle_client_autostart(self, state):
        """切换客户端开机自动连接"""
        import subprocess
        task_name = 'TCL表格比对客户端'
        try:
            if state == 2:  # Qt.Checked
                server_url = self.server_url_edit.text().strip()
                if not server_url:
                    QMessageBox.warning(self, "提示", "请先输入服务器地址并连接")
                    self.client_autostart_checkbox.setChecked(False)
                    return

                if getattr(sys, 'frozen', False):
                    exe_path = sys.executable
                else:
                    exe_path = os.path.abspath(sys.argv[0])
                # 创建计划任务
                cmd = f'schtasks /create /tn "{task_name}" /tr "\\"{exe_path}\\" --client" /sc onlogon /f'
                result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=10)
                if result.returncode == 0:
                    # 互斥：删除服务器自启动任务
                    subprocess.run('schtasks /delete /tn "TCL表格比对服务器" /f', shell=True, capture_output=True, text=True, timeout=10)
                    if hasattr(self, 'autostart_checkbox'):
                        self.autostart_checkbox.setChecked(False)
                    # 保存 auto_connect 标记
                    parent = self.parent()
                    if parent and hasattr(parent, 'network'):
                        parent.network.auto_connect = True
                        parent.network.save_config()
                    QMessageBox.information(self, "设置成功", "已设置开机自动连接服务器")
                else:
                    QMessageBox.warning(self, "设置失败", f"设置开机自启失败:\n{result.stderr}")
                    self.client_autostart_checkbox.setChecked(False)
            else:
                # 删除计划任务
                cmd = f'schtasks /delete /tn "{task_name}" /f'
                result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=10)
                if result.returncode == 0:
                    parent = self.parent()
                    if parent and hasattr(parent, 'network'):
                        parent.network.auto_connect = False
                        parent.network.save_config()
                    QMessageBox.information(self, "取消成功", "已取消开机自动连接")
                else:
                    QMessageBox.warning(self, "取消失败", f"取消开机自启失败:\n{result.stderr}")
                    self.client_autostart_checkbox.setChecked(True)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"操作失败:\n{str(e)}")

    def search_servers(self):
        """搜索局域网内的服务器"""
        self.search_btn.setEnabled(False)
        self.search_btn.setText("搜索中...")
        self.search_status.setText("正在搜索...")
        self.server_table.setRowCount(0)
        QApplication.processEvents()

        # 获取本机IP段（使用UDP方式避免DNS解析编码问题）
        import socket
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            local_ip = s.getsockname()[0]
            s.close()
            ip_parts = local_ip.split('.')
            network_prefix = f"{ip_parts[0]}.{ip_parts[1]}.{ip_parts[2]}"
        except:
            network_prefix = "192.168.1"

        # 在后台线程中搜索
        self.search_thread = ServerSearchThread(network_prefix)
        self.search_thread.server_found.connect(self.on_server_found)
        self.search_thread.search_finished.connect(self.on_search_finished)
        self.search_thread.start()

    def on_server_found(self, ip, port, is_online, server_name=""):
        """发现服务器回调"""
        row = self.server_table.rowCount()
        self.server_table.insertRow(row)

        self.server_table.setItem(row, 0, QTableWidgetItem(server_name))
        self.server_table.setItem(row, 1, QTableWidgetItem(ip))
        self.server_table.setItem(row, 2, QTableWidgetItem(str(port)))

        status_item = QTableWidgetItem("在线" if is_online else "离线")
        status_item.setForeground(QColor("#10B981" if is_online else "#EF4444"))
        self.server_table.setItem(row, 3, status_item)

    def on_search_finished(self, count):
        """搜索完成回调"""
        self.search_btn.setEnabled(True)
        self.search_btn.setText("搜索服务器")
        self.search_status.setText(f"找到 {count} 个服务器" if count > 0 else "未找到服务器")

    def on_server_selected(self):
        """服务器选中事件"""
        selected = self.server_table.selectedItems()
        self.use_server_btn.setEnabled(len(selected) > 0)

    def use_selected_server(self):
        """使用选中的服务器"""
        selected = self.server_table.selectedItems()
        if not selected:
            return

        ip = selected[1].text()
        port = selected[2].text()
        server_url = f"http://{ip}:{port}"

        # 设置地址并切换到服务器模式
        self.server_url_edit.setText(server_url)
        self.set_network_mode(True)

        if not self.connection_status.text().startswith("● 服务器模式"):
            # 连接失败，不关闭窗口
            return

        self.accept()

    def closeEvent(self, event):
        """关闭事件"""
        # 停止搜索线程
        if self.search_thread and self.search_thread.isRunning():
            self.search_thread.terminate()
            self.search_thread.wait(1000)
        event.accept()


class ServerSearchThread(QThread):
    """服务器搜索线程"""
    server_found = pyqtSignal(str, int, bool, str)  # ip, port, is_online, server_name
    search_finished = pyqtSignal(int)  # found_count

    def __init__(self, network_prefix):
        super().__init__()
        self.network_prefix = network_prefix
        self.ports_to_scan = [5000, 8080, 8000, 3000]

    def run(self):
        import socket
        import concurrent.futures
        import json

        found_count = 0

        def check_server(ip, port):
            nonlocal found_count
            try:
                # 第一步：快速检测端口是否开放
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.settimeout(0.5)
                result = sock.connect_ex((ip, port))
                sock.close()

                if result != 0:
                    return  # 端口未开放，跳过

                # 第二步：端口开放，尝试HTTP请求验证是否是TCL服务器
                try:
                    import urllib.request
                    import urllib.error

                    url = f"http://{ip}:{port}/api/health"
                    req = urllib.request.Request(url, method='GET')
                    req.add_header('User-Agent', 'TCL-Client/1.0')

                    with urllib.request.urlopen(req, timeout=2) as response:
                        if response.status == 200:
                            data = json.loads(response.read().decode('utf-8'))
                            # 验证是否是TCL服务器（检查返回的JSON结构）
                            if 'status' in data and data.get('status') == 'ok':
                                server_name = data.get('name', 'TCL表格比对系统服务器')
                                self.server_found.emit(ip, port, True, server_name)
                                found_count += 1
                except (urllib.error.URLError, urllib.error.HTTPError, json.JSONDecodeError, KeyError):
                    pass  # 不是TCL服务器，忽略
                except Exception:
                    pass

            except Exception:
                pass

        # 并发搜索
        with concurrent.futures.ThreadPoolExecutor(max_workers=50) as executor:
            futures = []
            for i in range(1, 255):
                ip = f"{self.network_prefix}.{i}"
                for port in self.ports_to_scan:
                    futures.append(executor.submit(check_server, ip, port))

            concurrent.futures.wait(futures)

        self.search_finished.emit(found_count)


class TCLApplication(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db = DatabaseManager()
        self.excel_processor = ExcelProcessor()

        # 初始化网络管理器（局域网模式）
        self.network = network_manager
        self._server_thread = None
        self._is_server_running = False

        self.yesterday_file = ""
        self.today_file = ""
        self.shipment_file = ""
        self.output_dir = get_app_dir()

        self.shipment_compare_file = ""
        self.shipment_target_file = ""

        self.sequence_file = ""
        self.query_file = ""

        self.category_output_dir = get_app_dir()
        self.pending_supply_file = ""
        self.pending_delivery_file = ""

        self.current_shipment_data = []
        self.pending_data = []
        self.current_db_file = ""
        self.current_db_headers = []

        # 数据持久化文件路径（配置文件在 data/ 目录）
        self.db_data_file = os.path.join(get_app_dir(), "data", "db_cache.json")

        # 批量导入数据持久化文件路径
        self.batch_import_data_file = os.path.join(get_app_dir(), "data", "batch_import_cache.json")

        # 输出目录配置文件路径（保存上一次使用的输出目录）
        self.output_dir_config_file = os.path.join(get_app_dir(), "data", "output_dir_config.json")

        self.init_ui()
        # 延迟加载保存的数据，确保UI完全初始化
        import time
        time.sleep(0.1)
        QApplication.processEvents()
        self.load_db_data()
        # 加载批量导入数据
        self.load_batch_import_data()
        # 加载上一次使用的输出目录配置
        self.load_output_dir_config()

        # 检查是否需要自动启动服务器（开机自启模式）
        if '--server' in sys.argv:
            QTimer.singleShot(1000, self.auto_start_server)
        elif '--client' in sys.argv:
            QTimer.singleShot(1000, self.auto_connect_client)

    def is_widget_valid(self, widget):
        """安全检查Qt widget对象是否仍然有效"""
        if widget is None:
            return False
        try:
            import sip
            return not sip.isdeleted(widget)
        except:
            try:
                # 如果无法使用sip，尝试调用一个简单的方法来测试有效性
                widget.isVisible()
                return True
            except:
                return False

    def showEvent(self, event):
        """窗口显示后刷新表格数据"""
        super().showEvent(event)
        # 确保表格在窗口显示后正确渲染
        if hasattr(self, 'db_table') and self.db_table.rowCount() > 0:
            self.db_table.viewport().update()

    def _read_json_file(self, file_path):
        """读取JSON文件，自动处理UTF-8和GBK编码"""
        import json
        # 先尝试UTF-8
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except UnicodeDecodeError:
            pass
        # 回退到GBK（旧版本Windows可能用GBK保存）
        try:
            with open(file_path, 'r', encoding='gbk') as f:
                data = json.load(f)
            # 用UTF-8重新保存，后续就不再有编码问题
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return data
        except Exception:
            return None

    def init_ui(self):
        self.setWindowTitle(f"TCL表格比对系统 v{__version__}")
        self.setGeometry(100, 100, 1200, 800)

        # 设置窗口图标（图标在 resources/ 目录）
        icon_path = os.path.join(get_app_dir(), "resources", "icon.png")
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        else:
            # 尝试从PyInstaller资源目录加载
            if getattr(sys, 'frozen', False):
                base_path = sys._MEIPASS
                icon_path = os.path.join(base_path, "icon.png")
                if os.path.exists(icon_path):
                    self.setWindowIcon(QIcon(icon_path))

        self.setup_styles()

        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        main_layout = QHBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        central_widget.setLayout(main_layout)

        self.setup_sidebar(main_layout)

        self.stacked_widget = QStackedWidget()
        main_layout.addWidget(self.stacked_widget)

        self.setup_page1()
        self.setup_page5()  # 数据查询（移到第2位）
        self.setup_page2()  # 数据库管理
        self.setup_page3()
        self.setup_page4()
        self.setup_page6()
        self.setup_page7()

        # 初始化系统托盘图标
        self.init_system_tray()

    def init_system_tray(self):
        """初始化系统托盘图标"""
        # 创建系统托盘图标
        self.tray_icon = QSystemTrayIcon(self)

        # 设置托盘图标（图标在 resources/ 目录）
        icon_path = os.path.join(get_app_dir(), "resources", "icon.png")
        if not os.path.exists(icon_path) and getattr(sys, 'frozen', False):
            # 尝试从PyInstaller资源目录加载
            base_path = sys._MEIPASS
            icon_path = os.path.join(base_path, "icon.png")

        if os.path.exists(icon_path):
            self.tray_icon.setIcon(QIcon(icon_path))
        else:
            # 使用默认图标
            self.tray_icon.setIcon(self.style().standardIcon(self.style().SP_ComputerIcon))

        # 创建托盘右键菜单
        tray_menu = QMenu()

        # 显示主窗口选项
        show_action = QAction("显示主窗口", self)
        show_action.triggered.connect(self.show_main_window)
        tray_menu.addAction(show_action)

        tray_menu.addSeparator()

        # 退出选项
        quit_action = QAction("退出", self)
        quit_action.triggered.connect(self.quit_application)
        tray_menu.addAction(quit_action)

        # 设置托盘图标的菜单
        self.tray_icon.setContextMenu(tray_menu)

        # 点击托盘图标时显示主窗口
        self.tray_icon.activated.connect(self.tray_icon_activated)

        # 显示托盘图标
        self.tray_icon.show()

    def tray_icon_activated(self, reason):
        """托盘图标被点击时的处理"""
        if reason == QSystemTrayIcon.DoubleClick:
            self.show_main_window()

    def show_main_window(self):
        """显示主窗口"""
        self.showNormal()
        self.activateWindow()
        self.raise_()

    def quit_application(self):
        """完全退出应用程序"""
        # 停止服务器线程（如果正在运行）
        if self._server_thread and self._server_thread.isRunning():
            self._server_thread.terminate()
            self._server_thread.wait(1000)

        # 关闭数据库连接
        if self.db:
            self.db.close()

        # 隐藏托盘图标
        self.tray_icon.hide()

        # 退出应用程序
        QApplication.quit()

    def setup_styles(self):
        self.sidebar_width = 80

        self.sidebar_style = """
            QListWidget {
                background-color: #F3F4F6;
                border: none;
                outline: none;
                padding: 0;
            }
            QListWidget::item {
                height: 70px;
                padding: 10px 0;
                text-align: center;
                color: #6B7280;
                font-size: 12px;
            }
            QListWidget::item:selected {
                background-color: #FFFFFF;
                color: #F97316;
                border-left: 3px solid #F97316;
            }
            QListWidget::item:hover:!selected {
                background-color: #E5E7EB;
            }
        """

        self.card_style = """
            QWidget {
                background-color: #FFFFFF;
                border-radius: 8px;
                padding: 20px;
            }
            QGroupBox {
                border: 1px solid #E5E7EB;
                border-radius: 8px;
                margin-top: 10px;
                font-size: 14px;
                font-weight: bold;
                color: #374151;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
        """

        self.btn_primary_style = """
            QPushButton {
                background-color: #F97316;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #EA580C;
            }
            QPushButton:pressed {
                background-color: #C2410C;
            }
        """

        self.btn_secondary_style = """
            QPushButton {
                background-color: #F3F4F6;
                color: #374151;
                border: 1px solid #D1D5DB;
                border-radius: 6px;
                padding: 8px 16px;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #E5E7EB;
            }
        """

        self.btn_danger_style = """
            QPushButton {
                background-color: #EF4444;
                color: #FFFFFF;
                border: 1px solid #DC2626;
                border-radius: 6px;
                padding: 8px 16px;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #DC2626;
            }
        """

        self.line_edit_style = """
            QLineEdit {
                border: 1px solid #D1D5DB;
                border-radius: 4px;
                padding: 8px 12px;
                font-size: 13px;
                background-color: #FFFFFF;
            }
            QLineEdit:focus {
                border-color: #F97316;
            }
        """

        self.table_style = """
            QTableWidget {
                border: 1px solid #E5E7EB;
                border-radius: 8px;
                gridline-color: #E5E7EB;
                background-color: #FFFFFF;
            }
            QTableWidget::item {
                padding: 8px;
            }
            QHeaderView::section {
                background-color: #4472C4;
                color: white;
                padding: 8px;
                border: none;
                font-weight: bold;
            }
        """

    def setup_sidebar(self, main_layout):
        sidebar = QFrame()
        sidebar.setFixedWidth(self.sidebar_width + 50)  # 加宽以容纳网络配置
        sidebar.setStyleSheet("background-color: #F3F4F6; border-right: 1px solid #E5E7EB;")

        sidebar_layout = QVBoxLayout()
        sidebar_layout.setContentsMargins(0, 10, 0, 10)
        sidebar_layout.setSpacing(5)
        sidebar.setLayout(sidebar_layout)

        title_label = QLabel("TCL")
        title_label.setStyleSheet("""
            font-size: 18px;
            font-weight: bold;
            color: #F97316;
            padding: 10px;
        """)
        title_label.setAlignment(Qt.AlignCenter)
        sidebar_layout.addWidget(title_label)

        # ========== 网络状态显示 ==========
        network_group = QGroupBox("网络状态")
        network_group.setStyleSheet("""
            QGroupBox {
                font-size: 11px;
                font-weight: bold;
                color: #374151;
                border: 1px solid #D1D5DB;
                border-radius: 4px;
                margin-top: 6px;
                padding-top: 6px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 8px;
                padding: 0 2px;
            }
        """)

        network_layout = QVBoxLayout()
        network_layout.setSpacing(5)

        # 模式状态显示
        self.network_mode_label = QLabel("● 本地模式")
        self.network_mode_label.setStyleSheet("""
            QLabel {
                color: #10B981;
                font-size: 11px;
                font-weight: bold;
                padding: 3px;
            }
        """)
        self.network_mode_label.setAlignment(Qt.AlignCenter)
        network_layout.addWidget(self.network_mode_label)

        # 服务器地址显示
        self.server_address_label = QLabel("")
        self.server_address_label.setStyleSheet("""
            QLabel {
                color: #6B7280;
                font-size: 9px;
                padding: 2px;
            }
        """)
        self.server_address_label.setAlignment(Qt.AlignCenter)
        self.server_address_label.setWordWrap(True)
        network_layout.addWidget(self.server_address_label)

        # ========== 服务器设置按钮 ==========
        self.open_server_settings_btn = QPushButton("服务器设置")
        self.open_server_settings_btn.setStyleSheet("""
            QPushButton {
                background-color: #3B82F6;
                color: white;
                border: none;
                border-radius: 3px;
                padding: 6px 8px;
                font-size: 10px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2563EB;
            }
        """)
        self.open_server_settings_btn.clicked.connect(self.open_server_settings)
        network_layout.addWidget(self.open_server_settings_btn)

        network_group.setLayout(network_layout)
        sidebar_layout.addWidget(network_group)

        # 初始化网络状态显示
        self.update_network_status()

        # 分隔线
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setStyleSheet("background-color: #D1D5DB;")
        line.setFixedHeight(1)
        sidebar_layout.addWidget(line)

        self.nav_list = QListWidget()
        self.nav_list.setStyleSheet(self.sidebar_style)
        self.nav_list.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.nav_list.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.nav_list.currentRowChanged.connect(self.on_nav_changed)

        nav_items = [
            ("表格比对", "1"),
            ("数据查询", "5"),
            ("数据库", "2"),
            ("出货管理", "3"),
            ("序号导入", "4"),
            ("分类导出", "6"),
            ("还需交货", "7"),
        ]

        for text, num in nav_items:
            item = QListWidgetItem(text)
            item.setTextAlignment(Qt.AlignCenter)
            self.nav_list.addItem(item)

        sidebar_layout.addWidget(self.nav_list)
        main_layout.addWidget(sidebar)

    def on_nav_changed(self, index):
        self.stacked_widget.setCurrentIndex(index)

    def create_page_layout(self):
        page = QWidget()
        page.setStyleSheet("background-color: #F9FAFB;")
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        page.setLayout(layout)
        return page

    def create_card(self, parent, title):
        card = QFrame()
        card.setStyleSheet("""
            QFrame {
                background-color: #FFFFFF;
                border-radius: 8px;
                border: 1px solid #E5E7EB;
                padding: 15px;
            }
        """)
        layout = QVBoxLayout()
        layout.setContentsMargins(10, 10, 10, 10)

        if title:
            title_label = QLabel(title)
            title_label.setStyleSheet("""
                font-size: 15px;
                font-weight: bold;
                color: #374151;
                padding-bottom: 10px;
            """)
            layout.addWidget(title_label)

        card.setLayout(layout)
        return card, layout

    def create_file_row(self, parent_layout, label_text, current_value, browse_callback):
        row_layout = QHBoxLayout()
        row_layout.setSpacing(10)

        label = QLabel(label_text)
        label.setFixedWidth(80)
        label.setAlignment(Qt.AlignCenter)
        label.setStyleSheet("color: #374151; font-size: 13px; background-color: #F3F4F6; padding: 5px; border-radius: 4px;")

        line_edit = QLineEdit()
        line_edit.setText(current_value)
        line_edit.setStyleSheet(self.line_edit_style)
        line_edit.setMinimumHeight(36)

        browse_btn = QPushButton("浏览")
        browse_btn.setFixedWidth(60)
        browse_btn.setStyleSheet(self.btn_secondary_style)
        browse_btn.clicked.connect(lambda: browse_callback(line_edit))

        row_layout.addWidget(label)
        row_layout.addWidget(line_edit, 1)
        row_layout.addWidget(browse_btn)

        parent_layout.addLayout(row_layout)
        return line_edit

    def setup_page1(self):
        page = self.create_page_layout()
        layout = page.layout()

        card, card_layout = self.create_card(page, "表格比对设置")

        yesterday_edit = self.create_file_row(
            card_layout, "昨天表格:", self.yesterday_file,
            lambda le: self.browse_file(le, "昨天表格"))
        today_edit = self.create_file_row(
            card_layout, "今天表格:", self.today_file,
            lambda le: self.browse_file(le, "今天表格"))
        shipment_edit = self.create_file_row(
            card_layout, "出货表格:", self.shipment_file,
            lambda le: self.browse_file(le, "出货表格"))
        output_edit = self.create_file_row(
            card_layout, "输出目录:", self.output_dir,
            lambda le: self.browse_dir(le))

        self.yesterday_edit = yesterday_edit
        self.today_edit = today_edit
        self.shipment_edit = shipment_edit
        self.output_edit = output_edit

        btn_layout = QHBoxLayout()
        compare_btn = QPushButton("开始比对并导出差异表格")
        compare_btn.setStyleSheet(self.btn_primary_style)
        compare_btn.setFixedHeight(40)
        compare_btn.clicked.connect(self.compare_and_export)
        btn_layout.addWidget(compare_btn)
        btn_layout.addStretch()
        card_layout.addLayout(btn_layout)

        result_card, result_layout = self.create_card(page, "比对结果")
        self.result_text = QTextEdit()
        self.result_text.setReadOnly(True)
        self.result_text.setStyleSheet("""
            QTextEdit {
                border: 1px solid #E5E7EB;
                border-radius: 4px;
                padding: 10px;
                font-family: 'Consolas', 'Microsoft YaHei';
                font-size: 12px;
                background-color: #F9FAFB;
            }
        """)
        result_layout.addWidget(self.result_text)

        layout.addWidget(card)
        layout.addWidget(result_card)
        self.stacked_widget.addWidget(page)

    def setup_page2(self):
        page = self.create_page_layout()
        layout = page.layout()

        search_layout = QHBoxLayout()
        search_label = QLabel("搜索:")
        search_label.setStyleSheet("color: #374151; font-size: 13px;")
        self.db_search_edit = QLineEdit()
        self.db_search_edit.setStyleSheet("""
            QLineEdit {
                border: 1px solid #D1D5DB;
                border-radius: 4px;
                padding: 8px 12px;
                font-size: 13px;
            }
            QLineEdit:focus {
                border-color: #F97316;
            }
        """)
        self.db_search_edit.setPlaceholderText("输入物料号精确搜索，多个用逗号分隔\n精确匹配格式：物料号|销售订单|销售订单行号|内需单号")
        self.db_search_edit.textChanged.connect(self.filter_db_data)

        import_search_btn = QPushButton("导入搜索条件")
        import_search_btn.setStyleSheet(self.btn_secondary_style)
        import_search_btn.setFixedHeight(35)
        import_search_btn.clicked.connect(self.import_search_conditions)

        clear_search_btn = QPushButton("清空")
        clear_search_btn.setStyleSheet(self.btn_secondary_style)
        clear_search_btn.setFixedHeight(35)
        clear_search_btn.clicked.connect(self.clear_search)

        search_layout.addWidget(search_label)
        search_layout.addWidget(self.db_search_edit, 1)
        search_layout.addWidget(import_search_btn)
        self.export_search_btn = QPushButton("导出搜索结果")
        self.export_search_btn.setStyleSheet(self.btn_secondary_style)
        self.export_search_btn.setFixedHeight(35)
        self.export_search_btn.clicked.connect(self.export_search_result)
        self.export_search_btn.setEnabled(True)
        search_layout.addWidget(self.export_search_btn)
        search_layout.addWidget(clear_search_btn)
        layout.addLayout(search_layout)

        self.db_table = QTableWidget()
        self.db_table.setColumnCount(0)
        self.db_table.setHorizontalHeaderLabels([])
        self.db_table.verticalHeader().setVisible(False)
        self.db_table.horizontalHeader().setVisible(True)
        self.db_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #E5E7EB;
                border-radius: 8px;
                gridline-color: #E5E7EB;
                background-color: #FFFFFF;
                selection-background-color: #E3F2FD;
                selection-color: #000000;
            }
            QHeaderView::section {
                background-color: #FFFFFF;
                color: #000000;
                padding: 8px 12px;
                border: none;
                border-bottom: 2px solid #E5E7EB;
                font-weight: bold;
                font-size: 13px;
            }
            QTableWidget::item:selected {
                color: #000000;
                background-color: #E3F2FD;
            }
        """)
        self.db_table.setAlternatingRowColors(True)
        self.db_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.db_table.setMinimumHeight(400)
        self.db_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        # 允许多选行
        self.db_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.db_table.setSelectionMode(QAbstractItemView.MultiSelection)
        layout.addWidget(self.db_table)

        # 勾选列按钮区域
        check_layout = QHBoxLayout()
        self.select_all_btn = QPushButton("全选")
        self.select_all_btn.setStyleSheet(self.btn_secondary_style)
        self.select_all_btn.setFixedHeight(30)
        self.select_all_btn.clicked.connect(self.select_all_rows)
        self.deselect_all_btn = QPushButton("取消全选")
        self.deselect_all_btn.setStyleSheet(self.btn_secondary_style)
        self.deselect_all_btn.setFixedHeight(30)
        self.deselect_all_btn.clicked.connect(self.deselect_all_rows)
        self.db_info_label = QLabel("请选择Excel文件导入")
        self.db_info_label.setStyleSheet("color: #6B7280; font-size: 12px; padding: 0 10px;")
        self.db_info_label.setMinimumSize(200, 20)
        check_layout.addWidget(QLabel("勾选列:"))
        check_layout.addWidget(self.select_all_btn)
        check_layout.addWidget(self.deselect_all_btn)
        check_layout.addWidget(self.db_info_label)
        check_layout.addStretch()
        layout.addLayout(check_layout)

        # 出货对比结果显示区域
        self.shipment_compare_label = QLabel("出货对比结果（待删除）")
        self.shipment_compare_label.setStyleSheet("color: #374151; font-size: 13px; font-weight: bold; padding: 5px 0;")
        self.shipment_compare_label.setVisible(False)
        layout.addWidget(self.shipment_compare_label)

        self.shipment_compare_table = QTableWidget()
        self.shipment_compare_table.setColumnCount(0)
        self.shipment_compare_table.setHorizontalHeaderLabels([])
        self.shipment_compare_table.verticalHeader().setVisible(False)
        self.shipment_compare_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #E5E7EB;
                border-radius: 8px;
                gridline-color: #E5E7EB;
                background-color: #FFF5F5;
                selection-background-color: #FFCCCC;
                selection-color: #000000;
            }
            QHeaderView::section {
                background-color: #FFCCCC;
                color: #000000;
                padding: 8px 12px;
                border: none;
                border-bottom: 2px solid #E5E7EB;
                font-weight: bold;
                font-size: 13px;
            }
            QTableWidget::item:selected {
                color: #000000;
                background-color: #FFCCCC;
            }
        """)
        self.shipment_compare_table.setAlternatingRowColors(True)
        self.shipment_compare_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.shipment_compare_table.setMinimumHeight(150)
        self.shipment_compare_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.shipment_compare_table.setVisible(False)
        layout.addWidget(self.shipment_compare_table)

        bottom_layout = QHBoxLayout()
        import_shipment_btn = QPushButton("导入出货Excel")
        import_shipment_btn.setStyleSheet(self.btn_secondary_style)
        import_shipment_btn.setFixedHeight(35)
        import_shipment_btn.clicked.connect(self.import_shipment_to_compare)
        export_shipment_btn = QPushButton("导出出货Excel")
        export_shipment_btn.setStyleSheet(self.btn_secondary_style)
        export_shipment_btn.setFixedHeight(35)
        export_shipment_btn.clicked.connect(self.export_shipment_data)
        self.delete_shipment_btn = QPushButton("删除匹配项")
        self.delete_shipment_btn.setStyleSheet("""
            QPushButton {
                background-color: #DC2626;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-size: 13px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #B91C1C;
            }
            QPushButton:disabled {
                background-color: #D1D5DB;
                color: #9CA3AF;
            }
        """)
        self.delete_shipment_btn.setFixedHeight(35)
        self.delete_shipment_btn.clicked.connect(self.delete_shipment_matched)
        # 默认启用删除按钮，支持复选框删除
        import_db_btn = QPushButton("导入Excel")
        import_db_btn.setStyleSheet(self.btn_primary_style)
        import_db_btn.setFixedHeight(35)
        import_db_btn.clicked.connect(self.open_excel_to_table)
        export_db_btn = QPushButton("导出Excel")
        export_db_btn.setStyleSheet(self.btn_secondary_style)
        export_db_btn.setFixedHeight(35)
        export_db_btn.clicked.connect(self.export_db_table)
        clear_db_btn = QPushButton("清空")
        clear_db_btn.setStyleSheet(self.btn_secondary_style)
        clear_db_btn.setFixedHeight(35)
        clear_db_btn.clicked.connect(self.clear_db_table)
        self.export_search_btn = QPushButton("导出搜索结果")
        self.export_search_btn.setStyleSheet(self.btn_secondary_style)
        self.export_search_btn.setFixedHeight(35)
        self.export_search_btn.clicked.connect(self.export_search_result)
        self.export_search_btn.setEnabled(False)
        bottom_layout.addWidget(import_shipment_btn)
        bottom_layout.addWidget(export_shipment_btn)
        bottom_layout.addWidget(self.delete_shipment_btn)
        bottom_layout.addStretch()
        bottom_layout.addWidget(import_db_btn)
        bottom_layout.addWidget(export_db_btn)
        bottom_layout.addWidget(clear_db_btn)
        layout.addLayout(bottom_layout)

        self.stacked_widget.addWidget(page)

    def setup_page3(self):
        page = self.create_page_layout()
        layout = page.layout()

        card, card_layout = self.create_card(page, "导入已出货表格")

        self.shipment_file_edit = self.create_file_row(
            card_layout, "出货表格:", self.shipment_file,
            lambda le: self.browse_file(le, "出货表格"))

        btn_layout = QHBoxLayout()
        import_btn = QPushButton("导入并高亮显示")
        import_btn.setStyleSheet(self.btn_primary_style)
        import_btn.clicked.connect(self.import_shipment)
        btn_layout.addWidget(import_btn)
        btn_layout.addStretch()
        card_layout.addLayout(btn_layout)

        self.shipment_table = QTableWidget()
        self.shipment_table.setColumnCount(3)
        self.shipment_table.setHorizontalHeaderLabels(['物料号', '数量', '订单号'])
        self.shipment_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.shipment_table.setStyleSheet(self.table_style)
        self.shipment_table.setAlternatingRowColors(True)
        self.shipment_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.shipment_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        card_layout.addWidget(self.shipment_table)

        btn_layout2 = QHBoxLayout()
        delete_btn = QPushButton("检查确认后删除选中")
        delete_btn.setStyleSheet(self.btn_secondary_style)
        delete_btn.clicked.connect(self.delete_shipment)
        export_highlight_btn = QPushButton("导出高亮表格")
        export_highlight_btn.setStyleSheet(self.btn_secondary_style)
        export_highlight_btn.clicked.connect(self.export_shipment_highlighted)
        btn_layout2.addWidget(delete_btn)
        btn_layout2.addWidget(export_highlight_btn)
        btn_layout2.addStretch()
        card_layout.addLayout(btn_layout2)

        card2, card2_layout = self.create_card(page, "出货表格比对")

        self.shipment_compare_edit = self.create_file_row(
            card2_layout, "出货表格:", self.shipment_compare_file,
            lambda le: self.browse_file(le, "出货表格"))
        self.shipment_target_edit = self.create_file_row(
            card2_layout, "对比表格:", self.shipment_target_file,
            lambda le: self.browse_file(le, "对比表格"))

        compare_btn = QPushButton("比对出货与对比表格")
        compare_btn.setStyleSheet(self.btn_primary_style)
        compare_btn.clicked.connect(self.compare_shipment)
        card2_layout.addWidget(compare_btn)

        self.shipment_compare_table = QTableWidget()
        self.shipment_compare_table.setColumnCount(4)
        self.shipment_compare_table.setHorizontalHeaderLabels(['物料号', '出货数量', '对比数量', '差异'])
        self.shipment_compare_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.shipment_compare_table.setStyleSheet(self.table_style)
        self.shipment_compare_table.setAlternatingRowColors(True)
        self.shipment_compare_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        card2_layout.addWidget(self.shipment_compare_table)

        layout.addWidget(card)
        layout.addWidget(card2)
        self.stacked_widget.addWidget(page)

    def setup_page4(self):
        page = self.create_page_layout()
        layout = page.layout()

        card, card_layout = self.create_card(page, "导入带序号和订单区分的表格")

        self.sequence_file_edit = self.create_file_row(
            card_layout, "序号表格:", self.sequence_file,
            lambda le: self.browse_file(le, "序号表格"))

        col_layout = QHBoxLayout()
        col_layout.setSpacing(15)

        seq_label = QLabel("序号列名:")
        seq_label.setStyleSheet("color: #374151; font-size: 13px;")
        self.sequence_col_edit = QLineEdit()
        self.sequence_col_edit.setText("序号")
        self.sequence_col_edit.setStyleSheet(self.line_edit_style)
        self.sequence_col_edit.setFixedWidth(150)

        order_label = QLabel("订单号列名:")
        order_label.setStyleSheet("color: #374151; font-size: 13px;")
        self.order_col_edit = QLineEdit()
        self.order_col_edit.setText("订单号")
        self.order_col_edit.setStyleSheet(self.line_edit_style)
        self.order_col_edit.setFixedWidth(150)

        col_layout.addWidget(seq_label)
        col_layout.addWidget(self.sequence_col_edit)
        col_layout.addWidget(order_label)
        col_layout.addWidget(self.order_col_edit)
        col_layout.addStretch()
        card_layout.addLayout(col_layout)

        import_btn = QPushButton("导入序号信息")
        import_btn.setStyleSheet(self.btn_primary_style)
        import_btn.clicked.connect(self.import_sequence)
        card_layout.addWidget(import_btn)

        info_label = QLabel("导入说明：此功能用于将带有序号和订单区分的表格导入到数据库，\n以便在出货时可以直接查询序号和订单信息。")
        info_label.setStyleSheet("color: #6B7280; font-size: 12px; padding: 20px 0;")
        info_label.setWordWrap(True)
        card_layout.addWidget(info_label)

        layout.addWidget(card)
        self.stacked_widget.addWidget(page)

    def setup_page5(self):
        page = self.create_page_layout()
        layout = page.layout()

        # 查询卡片
        card, card_layout = self.create_card(page, "查询并导出数据库信息")

        self.query_file_edit = self.create_file_row(
            card_layout, "查询表格:", self.query_file,
            lambda le: self.browse_file(le, "查询表格"))

        # 查询控件行
        query_row_layout = QHBoxLayout()
        col_label = QLabel("物料号列名:")
        col_label.setStyleSheet("color: #374151; font-size: 13px;")
        self.query_col_edit = QLineEdit()
        self.query_col_edit.setText("物料号")
        self.query_col_edit.setStyleSheet(self.line_edit_style)
        self.query_col_edit.setFixedWidth(150)

        self.query_btn = QPushButton("查询并导出")
        self.query_btn.setStyleSheet(self.btn_primary_style)
        self.query_btn.clicked.connect(self.query_and_export)

        self.import_to_main_btn = QPushButton("导入到主表格")
        self.import_to_main_btn.setStyleSheet(self.btn_primary_style)
        self.import_to_main_btn.clicked.connect(self.load_excel_to_query_table)

        query_row_layout.addWidget(col_label)
        query_row_layout.addWidget(self.query_col_edit)
        query_row_layout.addWidget(self.query_btn)
        query_row_layout.addWidget(self.import_to_main_btn)
        query_row_layout.addStretch()
        card_layout.addLayout(query_row_layout)

        # 主查询表格
        self.query_table = QTableWidget()
        self.query_table.setColumnCount(0)
        self.query_table.setHorizontalHeaderLabels([])
        self.query_table.setRowCount(0)
        self.query_table.verticalHeader().setVisible(False)
        self.query_table.horizontalHeader().setVisible(True)
        self.query_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #E5E7EB;
                border-radius: 8px;
                gridline-color: #E5E7EB;
                background-color: #FFFFFF;
                selection-background-color: #E3F2FD;
                selection-color: #000000;
            }
            QTableWidget::item:selected {
                color: #000000;
                background-color: #E3F2FD;
            }
        """)
        self.query_table.horizontalHeader().setStyleSheet("""
            QHeaderView::section {
                background-color: #4472C4;
                color: #FFFFFF;
                padding: 8px 12px;
                border: none;
                border-bottom: 2px solid #1E3A5F;
                font-weight: bold;
                font-size: 13px;
            }
        """)
        self.query_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.query_table.horizontalHeader().setStretchLastSection(True)
        self.query_table.horizontalHeader().setMinimumSectionSize(50)
        self.query_table.setAlternatingRowColors(True)
        self.query_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.query_table.setMinimumHeight(300)
        self.query_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.query_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.query_table.setSelectionMode(QAbstractItemView.MultiSelection)
        card_layout.addWidget(self.query_table)

        self.query_stats_label = QLabel("")
        self.query_stats_label.setStyleSheet("color: #374151; font-size: 13px; padding: 4px 0;")
        card_layout.addWidget(self.query_stats_label)

        self.batch_import_data = []
        self.batch_import_headers = []  # 存储从Excel读取的原始列头（用于动态显示）

        # 批量导入区域（不使用卡片容器，让界面更紧凑）

        self.batch_import_table = QTableWidget()
        self.batch_import_table.setColumnCount(1)  # 初始只设置1列（勾选列），后续根据Excel动态调整
        self.batch_import_table.setHorizontalHeaderLabels(['勾选'])
        self.batch_import_table.verticalHeader().setVisible(False)
        self.batch_import_table.horizontalHeader().setVisible(True)

        # 完全复制db_table的成功配置（与数据库标签页保持100%一致）
        self.batch_import_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #E5E7EB;
                border-radius: 8px;
                gridline-color: #E5E7EB;
                background-color: #FFFFFF;
                selection-background-color: #E3F2FD;
                selection-color: #000000;
            }
            QHeaderView::section {
                background-color: #FFFFFF;
                color: #000000;
                padding: 8px 12px;
                border: none;
                border-bottom: 2px solid #E5E7EB;
                font-weight: bold;
                font-size: 13px;
            }
            QTableWidget::item:selected {
                color: #000000;
                background-color: #E3F2FD;
            }
        """)
        self.batch_import_table.setAlternatingRowColors(True)
        self.batch_import_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.batch_import_table.setMinimumHeight(250)
        self.batch_import_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.batch_import_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.batch_import_table.setSelectionMode(QAbstractItemView.MultiSelection)

        # 搜索功能区域 - 放在表格上方
        search_layout = QHBoxLayout()
        search_label = QLabel("搜索:")
        search_label.setStyleSheet("color: #374151; font-size: 13px;")
        self.batch_search_edit = QLineEdit()
        self.batch_search_edit.setStyleSheet("""
            QLineEdit {
                border: 1px solid #D1D5DB;
                border-radius: 4px;
                padding: 8px 12px;
                font-size: 13px;
            }
            QLineEdit:focus {
                border-color: #F97316;
            }
        """)
        self.batch_search_edit.setPlaceholderText("输入物料号搜索，多个用逗号分隔")
        self.batch_search_edit.textChanged.connect(self.filter_batch_data)

        import_batch_search_btn = QPushButton("导入搜索条件")
        import_batch_search_btn.setStyleSheet(self.btn_secondary_style)
        import_batch_search_btn.setFixedHeight(35)
        import_batch_search_btn.clicked.connect(self.import_batch_search_conditions)

        export_batch_search_btn = QPushButton("导出搜索结果")
        export_batch_search_btn.setStyleSheet(self.btn_secondary_style)
        export_batch_search_btn.setFixedHeight(35)
        export_batch_search_btn.clicked.connect(self.export_batch_search_result)

        clear_batch_search_btn = QPushButton("清空")
        clear_batch_search_btn.setStyleSheet(self.btn_secondary_style)
        clear_batch_search_btn.setFixedHeight(35)
        clear_batch_search_btn.clicked.connect(self.clear_batch_search)

        search_layout.addWidget(search_label)
        search_layout.addWidget(self.batch_search_edit, 1)
        search_layout.addWidget(import_batch_search_btn)
        search_layout.addWidget(export_batch_search_btn)
        search_layout.addWidget(clear_batch_search_btn)
        layout.addLayout(search_layout)

        # 关键修复：将表格直接添加到页面布局，而不是卡片容器中！
        layout.addWidget(self.batch_import_table)

        # 调试信息：打印父容器和样式表
        print(f"[DEBUG] batch_import_table 父容器: {self.batch_import_table.parent().__class__.__name__}")
        print(f"[DEBUG] batch_import_table 样式表已设置: {bool(self.batch_import_table.styleSheet())}")
        print(f"[DEBUG] batch_import_table 列数: {self.batch_import_table.columnCount()}")
        print(f"[DEBUG] batch_import_table 表头可见: {self.batch_import_table.horizontalHeader().isVisible()}")

        # 勾选列按钮行
        check_layout = QHBoxLayout()
        self.batch_select_all_btn = QPushButton("全选")
        self.batch_select_all_btn.setStyleSheet(self.btn_secondary_style)
        self.batch_select_all_btn.setFixedHeight(30)
        self.batch_select_all_btn.clicked.connect(self.batch_select_all_rows)
        self.batch_deselect_all_btn = QPushButton("取消全选")
        self.batch_deselect_all_btn.setStyleSheet(self.btn_secondary_style)
        self.batch_deselect_all_btn.setFixedHeight(30)
        self.batch_deselect_all_btn.clicked.connect(self.batch_deselect_all_rows)
        self.batch_info_label = QLabel("已导入 0 条")
        self.batch_info_label.setStyleSheet("color: #6B7280; font-size: 12px; padding: 0 10px;")
        self.batch_info_label.setMinimumSize(150, 20)

        check_layout.addWidget(QLabel("勾选列:"))
        check_layout.addWidget(self.batch_select_all_btn)
        check_layout.addWidget(self.batch_deselect_all_btn)
        check_layout.addWidget(self.batch_info_label)
        check_layout.addStretch()
        layout.addLayout(check_layout)

        # 操作按钮行
        action_layout = QHBoxLayout()
        # 左侧：导入出货Excel和导出Excel按钮
        self.query_import_shipment_btn = QPushButton("导入出货Excel")
        self.query_import_shipment_btn.setStyleSheet(self.btn_secondary_style)
        self.query_import_shipment_btn.setFixedHeight(30)
        self.query_import_shipment_btn.clicked.connect(self.import_shipment_to_compare_for_query)

        self.query_export_shipment_btn = QPushButton("导出出货Excel")
        self.query_export_shipment_btn.setStyleSheet(self.btn_secondary_style)
        self.query_export_shipment_btn.setFixedHeight(30)
        self.query_export_shipment_btn.clicked.connect(self.export_batch_data_for_query)

        self.query_delete_match_btn = QPushButton("删除匹配项")
        self.query_delete_match_btn.setStyleSheet("""
            QPushButton {
                background-color: #DC2626;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-size: 13px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #B91C1C;
            }
            QPushButton:disabled {
                background-color: #D1D5DB;
                color: #9CA3AF;
            }
        """)
        self.query_delete_match_btn.setFixedHeight(30)
        self.query_delete_match_btn.clicked.connect(self.delete_query_matched)

        # 右侧：导入、导出、清空按钮
        self.batch_import_btn = QPushButton("导入excel")
        self.batch_import_btn.setStyleSheet(self.btn_primary_style)
        self.batch_import_btn.clicked.connect(self.batch_import_excel)
        self.export_btn = QPushButton("导出Excel")
        self.export_btn.setStyleSheet(self.btn_secondary_style)
        self.export_btn.clicked.connect(self.batch_export_selected)
        self.clear_batch_btn = QPushButton("清空")
        self.clear_batch_btn.setStyleSheet(self.btn_secondary_style)
        self.clear_batch_btn.setFixedHeight(30)
        self.clear_batch_btn.clicked.connect(self.clear_batch_table)

        action_layout.addWidget(self.query_import_shipment_btn)    # 左下角：导入出货
        action_layout.addWidget(self.query_export_shipment_btn)     # 左下角：导出
        action_layout.addWidget(self.query_delete_match_btn)        # 左下角：删除匹配项
        action_layout.addStretch()                                  # 弹性空间
        action_layout.addWidget(self.batch_import_btn)             # 右下角：导入
        action_layout.addWidget(self.export_btn)                   # 右下角：导出
        action_layout.addWidget(self.clear_batch_btn)              # 右下角：清空
        layout.addLayout(action_layout)

        self.stacked_widget.addWidget(page)

    def setup_page6(self):
        page = self.create_page_layout()
        layout = page.layout()

        card, card_layout = self.create_card(page, "按物料描述分类导出")

        info_label = QLabel("选择数据库中的一级表格数据，按物料描述分类导出为多个Excel文件。")
        info_label.setStyleSheet("color: #374151; font-size: 13px; padding-bottom: 15px;")
        card_layout.addWidget(info_label)

        self.category_output_edit = self.create_file_row(
            card_layout, "输出目录:", self.category_output_dir,
            lambda le: self.browse_dir(le))

        export_btn = QPushButton("导出分类表格")
        export_btn.setStyleSheet(self.btn_primary_style)
        export_btn.clicked.connect(self.export_by_category)
        card_layout.addWidget(export_btn)

        info_text = QLabel("""分类规则：
1. 透明商标类：透明、商标、白色PET
2. 铝箔类：银、铝箔、电化铝箔
3. 接线类：接线
4. 能源能效类：能源、能效
5. 标贴类：型号标贴、机型标贴、纸箱标贴、不可移铜版纸、不可移光粉纸、指示标贴
6. 说明书类：说明书、合格证、保修卡、清单、附页、手册、用户、书写纸、参数页
7. 特光类：特光""")
        info_text.setStyleSheet("color: #6B7280; font-size: 12px; padding: 15px 0;")
        info_text.setWordWrap(True)
        card_layout.addWidget(info_text)

        layout.addWidget(card)
        self.stacked_widget.addWidget(page)

    def setup_page7(self):
        page = self.create_page_layout()
        layout = page.layout()

        card, card_layout = self.create_card(page, "还需交货统计（供货计划 - 已送货）")

        self.pending_supply_edit = self.create_file_row(
            card_layout, "供货计划:", "大诚供货计划26-4-28.xlsx",
            lambda le: self.browse_file(le, "供货计划"))
        self.pending_delivery_edit = self.create_file_row(
            card_layout, "送货单:", "送货单明细视图1777775654529.xlsx",
            lambda le: self.browse_file(le, "送货单"))

        btn_layout = QHBoxLayout()
        calc_btn = QPushButton("计算还需交货")
        calc_btn.setStyleSheet(self.btn_primary_style)
        calc_btn.clicked.connect(self.calculate_pending_delivery)
        export_btn = QPushButton("导出Excel")
        export_btn.setStyleSheet(self.btn_secondary_style)
        export_btn.clicked.connect(self.export_pending_delivery)
        refresh_btn = QPushButton("刷新")
        refresh_btn.setStyleSheet(self.btn_secondary_style)
        refresh_btn.clicked.connect(self.calculate_pending_delivery)
        btn_layout.addWidget(calc_btn)
        btn_layout.addWidget(export_btn)
        btn_layout.addWidget(refresh_btn)
        btn_layout.addStretch()
        card_layout.addLayout(btn_layout)

        self.pending_table = QTableWidget()
        self.pending_table.setColumnCount(7)
        self.pending_table.setHorizontalHeaderLabels(['物料号', '物料描述', '总缺料', '已送货', '还需交货', '在途', '已承诺'])
        self.pending_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.pending_table.setStyleSheet(self.table_style)
        self.pending_table.setAlternatingRowColors(True)
        self.pending_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        card_layout.addWidget(self.pending_table)

        layout.addWidget(card)
        self.stacked_widget.addWidget(page)

    def browse_file(self, line_edit, title="选择文件"):
        filename, _ = QFileDialog.getOpenFileName(
            self, title, "", "Excel files (*.xlsx *.xls);;All Files (*)"
        )
        if filename:
            line_edit.setText(filename)

    def browse_dir(self, line_edit):
        dirname = QFileDialog.getExistingDirectory(self, "选择目录")
        if dirname:
            line_edit.setText(dirname)
            # 判断是哪个输出目录，并保存配置
            if line_edit == self.output_edit:
                self.output_dir = dirname
            elif hasattr(self, 'category_output_edit') and line_edit == self.category_output_edit:
                self.category_output_dir = dirname
            # 保存输出目录配置
            self.save_output_dir_config()

    def get_material(self, row):
        return row.get('物料号', '') or row.get('物料编码', '') or ''

    def get_sales_order(self, row):
        return str(row.get('销售订单', '') or row.get('销售订单号', '') or '').strip()

    def make_match_key(self, row):
        material = self.get_material(row)
        sales_order = self.get_sales_order(row)
        sales_line = str(row.get('销售订单行号', '') or '').strip()
        internal_order = str(row.get('内需单号', '') or '').strip()
        # 出货表格可能用销售订单号而不是销售订单
        if not sales_order:
            sales_order = str(row.get('销售订单号', '') or '').strip()
        return f"{material}|{sales_order}|{sales_line}|{internal_order}"

    def compare_and_export(self):
        yesterday_path = self.yesterday_edit.text()
        today_path = self.today_edit.text()
        shipment_path = self.shipment_edit.text()
        output_dir = self.output_edit.text()

        if not yesterday_path or not today_path:
            QMessageBox.critical(self, "错误", "请选择昨天和今天的表格文件")
            return

        try:
            self.result_text.clear()
            self.result_text.append("正在读取表格...\n")

            _, yesterday_data = self.excel_processor.read_excel(yesterday_path)
            _, today_data, today_color_map = self.excel_processor.read_excel_with_color(today_path)

            self.result_text.append(f"昨天表格: {len(yesterday_data)} 条记录\n")
            self.result_text.append(f"今天表格: {len(today_data)} 条记录\n")
            self.result_text.append(f"今天表格带颜色: {len(today_color_map)} 条\n")

            shipment_data = []
            shipment_dict = {}
            if shipment_path:
                _, shipment_data = self.excel_processor.read_excel(shipment_path)
                self.result_text.append(f"出货表格: {len(shipment_data)} 条记录\n")

                shipment_dict = self.excel_processor.build_shipment_quantity_map(shipment_data)

                self.result_text.append(f"出货匹配键数量: {len(shipment_dict)} 条\n")

                # 不再扣减出货，保留今天表格原始总缺料
                # today_data = self.subtract_shipment(today_data, shipment_dict)
                # self.result_text.append(f"扣除出货后今天表格: {len(today_data)} 条记录\n")

            diff_data = self.excel_processor.compare_and_get_diff(yesterday_data, today_data)
            diff_data = self.excel_processor.attach_shipment_quantities(diff_data, shipment_dict)
            diff_data = self.excel_processor.sort_diff_data(diff_data)

            self.result_text.append(f"发现 {len(diff_data)} 条差异记录\n")

            for item in diff_data:
                self.result_text.append(
                    f"物料号: {item['物料号']}, 销售订单: {item.get('销售订单', '')}, "
                    f"本次送货: {item.get('本次送货数量', 0)}, 变化: {item['变化量']}"
                )

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = os.path.join(output_dir, f"缺料差异_{timestamp}.xlsx")

            self.excel_processor.export_diff_to_excel(diff_data, output_file, today_color_map)

            self.result_text.append(f"\n差异表格已导出到: {output_file}\n")

            # 保存输出目录配置（比对完成后）
            self.output_dir = output_dir
            self.save_output_dir_config()

            data_for_db = self.excel_processor.build_level1_records(diff_data, today_data)

            self.db.insert_or_update_level1(data_for_db)

            self.result_text.append(f"已更新数据库 {len(data_for_db)} 条记录\n")

            QMessageBox.information(self, "完成", f"比对完成！\n差异记录: {len(diff_data)} 条\n文件已保存到:\n{output_file}")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"处理失败: {str(e)}")
            self.result_text.append(f"\n错误: {str(e)}\n")

    def subtract_shipment(self, today_data, shipment_input):
        if isinstance(shipment_input, dict):
            shipment_dict = shipment_input
        else:
            shipment_dict = {}
            for row in shipment_input:
                key = self.make_match_key(row)
                quantity = row.get('本次送货数量', 0) or 0
                if key and key != '|||':
                    shipment_dict[key] = shipment_dict.get(key, 0) + quantity

        result_data = []
        for row in today_data:
            new_row = dict(row)
            key = self.make_match_key(row)
            if key and key in shipment_dict:
                original_qty = row.get('总缺料', 0) or 0
                new_row['总缺料'] = max(0, original_qty - shipment_dict.get(key, 0))
            result_data.append(new_row)

        return result_data

    def refresh_level1(self):
        self.level1_table.setRowCount(0)
        data = self.db.get_level1_all()
        for row in data[:100]:
            row_idx = self.level1_table.rowCount()
            self.level1_table.insertRow(row_idx)
            self.level1_table.setItem(row_idx, 0, QTableWidgetItem(str(row.get('ID', ''))))
            self.level1_table.setItem(row_idx, 1, QTableWidgetItem(str(row.get('物料号', ''))))
            self.level1_table.setItem(row_idx, 2, QTableWidgetItem(str(row.get('物料描述', ''))[:30]))
            self.level1_table.setItem(row_idx, 3, QTableWidgetItem(str(row.get('供方', ''))))
            self.level1_table.setItem(row_idx, 4, QTableWidgetItem(str(row.get('总缺料', ''))))
            self.level1_table.setItem(row_idx, 5, QTableWidgetItem(str(row.get('分类', ''))))

    def refresh_level2(self):
        self.level2_table.setRowCount(0)
        data = self.db.get_level2_all()
        for row in data[:100]:
            row_idx = self.level2_table.rowCount()
            self.level2_table.insertRow(row_idx)
            self.level2_table.setItem(row_idx, 0, QTableWidgetItem(str(row.get('ID', ''))))
            self.level2_table.setItem(row_idx, 1, QTableWidgetItem(str(row.get('物料号', ''))))
            self.level2_table.setItem(row_idx, 2, QTableWidgetItem(str(row.get('物料描述', ''))[:30]))
            self.level2_table.setItem(row_idx, 3, QTableWidgetItem(str(row.get('序号', ''))))
            self.level2_table.setItem(row_idx, 4, QTableWidgetItem(str(row.get('订单号', ''))))
            self.level2_table.setItem(row_idx, 5, QTableWidgetItem(str(row.get('总缺料', ''))))
            self.level2_table.setItem(row_idx, 6, QTableWidgetItem(str(row.get('分类', ''))))

    def export_level1(self):
        filename, _ = QFileDialog.getSaveFileName(
            self, "保存一级表格", "", "Excel files (*.xlsx)"
        )
        if filename:
            data = self.db.get_level1_all()
            export_data = []
            for row in data:
                export_data.append({
                    'ID': row.get('ID'),
                    '物料号': row.get('物料号'),
                    '物料描述': row.get('物料描述'),
                    '供方': row.get('供方'),
                    '总缺料': row.get('总缺料'),
                    '分类': row.get('分类'),
                    '序号': row.get('序号'),
                    '订单号': row.get('订单号')
                })
            self.excel_processor.export_to_excel(export_data, filename)
            QMessageBox.information(self, "完成", f"一级表格已导出到:\n{filename}")

    def export_level2(self):
        filename, _ = QFileDialog.getSaveFileName(
            self, "保存二级表格", "", "Excel files (*.xlsx)"
        )
        if filename:
            data = self.db.get_level2_all()
            export_data = []
            for row in data:
                export_data.append({
                    'ID': row.get('ID'),
                    '物料号': row.get('物料号'),
                    '物料描述': row.get('物料描述'),
                    '序号': row.get('序号'),
                    '订单号': row.get('订单号'),
                    '总缺料': row.get('总缺料'),
                    '分类': row.get('分类'),
                    '下单时间': row.get('下单时间')
                })
            self.excel_processor.export_to_excel(export_data, filename)
            QMessageBox.information(self, "完成", f"二级表格已导出到:\n{filename}")

    def import_level2(self):
        filename, _ = QFileDialog.getOpenFileName(
            self, "选择二级表格文件", "", "Excel files (*.xlsx *.xls)"
        )
        if not filename:
            return

        try:
            headers, data = self.excel_processor.read_excel(filename)

            data_for_db = []
            for row in data:
                material_number = row.get('物料号', '')
                if material_number:
                    data_for_db.append({
                        '物料号': material_number,
                        '物料描述': row.get('物料描述', ''),
                        '序号': row.get('序号', ''),
                        '订单号': row.get('订单号', ''),
                        '总缺料': row.get('总缺料', 0),
                        '分类': row.get('分类', ''),
                        '下单时间': row.get('下单时间', '')
                    })

            if data_for_db:
                self.db.insert_level2(data_for_db)
                self.refresh_level2()
                QMessageBox.information(self, "完成", f"已导入 {len(data_for_db)} 条记录到二级表格")
            else:
                QMessageBox.warning(self, "提示", "文件中没有找到有效的物料数据")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"导入失败: {str(e)}")

    def open_excel_to_table(self):
        filenames, _ = QFileDialog.getOpenFileNames(
            self, "选择Excel文件", "", "Excel files (*.xlsx *.xls)"
        )
        if not filenames:
            return

        # 多个文件时，导入模式选择
        if len(filenames) > 1:
            reply = QMessageBox.question(self, "批量导入",
                f"已选择 {len(filenames)} 个文件\n\n是：追加合并\n否：替换数据",
                QMessageBox.Yes | QMessageBox.No)
        else:
            reply = None

        all_data = []
        all_headers = []

        for filename in filenames:
            try:
                headers, data = self.excel_processor.read_excel(filename)
                if not all_headers:
                    all_headers = list(headers)
                all_data.extend(data)
            except Exception as e:
                print(f"读取文件失败 {filename}: {e}")
                continue

        if not all_data:
            QMessageBox.warning(self, "提示", "没有读取到有效数据")
            return

        data = all_data
        headers = all_headers

        # 如果已有数据，询问用户是追加还是替换
        if hasattr(self, 'db_all_data') and self.db_all_data and reply != QMessageBox.No:
            # 追加模式：合并数据，保持原有表头顺序
            existing_headers_list = list(self.current_db_headers)

            # 合并表头（保持顺序，不重复）
            merged_headers = list(existing_headers_list)
            for h in headers:
                if h and h not in merged_headers:
                    merged_headers.append(h)

            # 确保每行数据都有所有列（用空字符串填充缺失的列）
            merged_data = []
            for existing_row in self.db_all_data:
                new_row = {}
                for h in merged_headers:
                    new_row[h] = existing_row.get(h, '')
                merged_data.append(new_row)

            for row in data:
                # 使用物料号+销售订单+销售订单行号+内需单号作为唯一键去重
                material = str(row.get('物料号', '') or '').strip()
                sales_order = str(row.get('销售订单', '') or '').strip()
                sales_line = str(row.get('销售订单行号', '') or '').strip()
                internal_order = str(row.get('内需单号', '') or '').strip()
                key = f"{material}|{sales_order}|{sales_line}|{internal_order}"

                # 检查是否已存在
                exists = False
                for merged_row in merged_data:
                    exist_material = str(merged_row.get('物料号', '') or '').strip()
                    exist_sales_order = str(merged_row.get('销售订单', '') or '').strip()
                    exist_sales_line = str(merged_row.get('销售订单行号', '') or '').strip()
                    exist_internal_order = str(merged_row.get('内需单号', '') or '').strip()
                    exist_key = f"{exist_material}|{exist_sales_order}|{exist_sales_line}|{exist_internal_order}"
                    if key == exist_key:
                        exists = True
                        break

                if not exists:
                    # 新增行也需要有所有列
                    new_row = {}
                    for h in merged_headers:
                        new_row[h] = row.get(h, '')
                    merged_data.append(new_row)

            data = merged_data
            headers = merged_headers

        col_count = len(headers)
        row_count = len(data)

        self.db_table.setColumnCount(col_count + 1)
        self.db_table.setRowCount(row_count)

        # 第一列：复选框列
        select_header_item = QTableWidgetItem("选择")
        select_header_item.setTextAlignment(Qt.AlignCenter)
        self.db_table.setHorizontalHeaderItem(0, select_header_item)
        for col_idx, header in enumerate(headers):
            item = QTableWidgetItem(str(header) if header else '')
            self.db_table.setHorizontalHeaderItem(col_idx + 1, item)

        for row_idx in range(row_count):
            row_data = data[row_idx]
            # 第一列：复选框使用QCheckBox居中显示
            checkbox_widget = QWidget()
            checkbox_layout = QHBoxLayout(checkbox_widget)
            checkbox_layout.setContentsMargins(0, 0, 0, 0)
            checkbox_layout.setAlignment(Qt.AlignCenter)
            checkbox = QCheckBox()
            checkbox.setStyleSheet("QCheckBox { margin-left: 15px; }")
            checkbox_layout.addWidget(checkbox)
            self.db_table.setCellWidget(row_idx, 0, checkbox_widget)
            # 数据列
            for col_idx in range(col_count):
                key = headers[col_idx]
                value = row_data.get(key, '') if isinstance(row_data, dict) else ''
                if value is None:
                    value = ''
                self.db_table.setItem(row_idx, col_idx + 1, QTableWidgetItem(str(value)))

        for col_idx in range(col_count + 1):
            if col_idx == 0:
                self.db_table.setColumnWidth(col_idx, 50)
            else:
                self.db_table.resizeColumnToContents(col_idx)
                if col_idx > 0 and headers[col_idx - 1] == '物料描述':
                    self.db_table.setColumnWidth(col_idx, 300)

        self.db_all_data = data
        self.db_original_data = data
        self.current_db_file = ",".join([os.path.basename(f) for f in filenames])
        self.current_db_headers = list(headers)
        self.red_highlight_indices = set()

        file_count = len(filenames)
        if file_count > 1:
            self.db_info_label.setText(f"已批量导入 {file_count} 个文件 - 共 {len(data)} 条数据")
        else:
            self.db_info_label.setText(f"已加载: {os.path.basename(filenames[0])} - 共 {len(data)} 条数据")

        # 启用导出搜索结果按钮
        self.export_search_btn.setEnabled(True)

        # 保存数据到本地文件
        self.save_db_data()

    def import_shipment_to_compare(self):
        """导入出货Excel表格，与数据库对比并显示匹配项"""
        if not hasattr(self, 'db_all_data') or not self.db_all_data:
            QMessageBox.warning(self, "提示", "请先导入数据库Excel文件")
            return

        filename, _ = QFileDialog.getOpenFileName(
            self, "选择出货Excel文件", "", "Excel files (*.xlsx *.xls)"
        )
        if not filename:
            return

        try:
            shipment_headers, shipment_data = self.excel_processor.read_excel(filename)

            if not shipment_data:
                QMessageBox.warning(self, "提示", "出货Excel文件中没有数据")
                return

            # 查找出货表格中的关键列
            col_names = {
                'material': None,
                'sales_order': None,
                'sales_line': None,
                'internal_order': None
            }
            for h in shipment_headers:
                if h:
                    h_str = str(h)
                    if '物料号' in h_str or '物料编码' in h_str:
                        col_names['material'] = h
                    elif '销售订单' in h_str and '行号' not in h_str:
                        col_names['sales_order'] = h
                    elif '销售订单行号' in h_str:
                        col_names['sales_line'] = h
                    elif '内需单号' in h_str:
                        col_names['internal_order'] = h

            # 构建出货数据的匹配键列表
            shipment_keys = []
            for row in shipment_data:
                material_val = row.get(col_names['material'])
                material = str(material_val if material_val is not None else '').strip()
                if material:
                    sales_order_val = row.get(col_names['sales_order'])
                    sales_order = str(sales_order_val if sales_order_val is not None else '').strip()
                    sales_line_val = row.get(col_names['sales_line'])
                    sales_line = str(sales_line_val if sales_line_val is not None else '').strip()
                    internal_order_val = row.get(col_names['internal_order'])
                    internal_order = str(internal_order_val if internal_order_val is not None else '').strip()
                    # 构建完整匹配键
                    shipment_keys.append({
                        'material': material,
                        'sales_order': sales_order,
                        'sales_line': sales_line,
                        'internal_order': internal_order
                    })

            if not shipment_keys:
                QMessageBox.warning(self, "提示", "出货Excel文件中没有找到有效的物料号")
                return

            # 获取数据库的关键列名
            headers = self.current_db_headers
            db_col_names = {
                'material': None,
                'sales_order': None,
                'sales_line': None,
                'internal_order': None
            }
            for h in headers:
                if h:
                    h_str = str(h)
                    if '物料号' in h_str or '物料编码' in h_str:
                        db_col_names['material'] = h
                    elif '销售订单' in h_str and '行号' not in h_str:
                        db_col_names['sales_order'] = h
                    elif '销售订单行号' in h_str:
                        db_col_names['sales_line'] = h
                    elif '内需单号' in h_str:
                        db_col_names['internal_order'] = h

            # 在数据库中查找匹配的记录索引
            matched_indices = []
            for idx, row in enumerate(self.db_all_data):
                db_material = str(row.get(db_col_names['material'], '') or '').strip()
                db_sales_order = str(row.get(db_col_names['sales_order'], '') or '').strip()
                db_sales_line = str(row.get(db_col_names['sales_line'], '') or '').strip()
                db_internal_order = str(row.get(db_col_names['internal_order'], '') or '').strip()

                # 检查是否匹配任一出货键
                for ship_key in shipment_keys:
                    # 必须物料号匹配
                    if db_material.lower() != ship_key['material'].lower():
                        continue
                    # 检查销售订单（如果出货数据中有值）
                    if ship_key['sales_order'] and db_sales_order.lower() != ship_key['sales_order'].lower():
                        continue
                    # 检查销售订单行号（如果出货数据中有值）
                    if ship_key['sales_line'] and db_sales_line.lower() != ship_key['sales_line'].lower():
                        continue
                    # 检查内需单号（如果出货数据中有值）
                    if ship_key['internal_order'] and db_internal_order.lower() != ship_key['internal_order'].lower():
                        continue
                    # 所有条件满足
                    matched_indices.append(idx)
                    break

            if not matched_indices:
                QMessageBox.information(self, "提示", f"出货表格中的 {len(shipment_keys)} 个物料号在数据库中未找到匹配项")
                return

            # 保存匹配的索引用于后续删除
            self.shipment_matched_indices = matched_indices

            # 在主表格中高亮显示匹配结果
            self.display_shipment_matched(matched_indices)

            # 启用删除按钮
            self.delete_shipment_btn.setEnabled(True)

            # 更新标签
            self.db_info_label.setText(f"出货对比模式 - 共 {len(matched_indices)} 条匹配记录（红色高亮显示）")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"处理失败: {str(e)}")

    def display_shipment_matched(self, indices):
        """在主表格中高亮显示出货匹配的记录，匹配项排在顶部"""
        red_fill = QColor(255, 100, 100)  # 红色高亮
        headers = self.current_db_headers
        data = self.db_all_data
        matched_set = set(indices)

        col_count = len(headers)

        # 将数据分为匹配和非匹配两组，匹配项放前面
        matched_data = [row for idx, row in enumerate(data) if idx in matched_set]
        unmatched_data = [row for idx, row in enumerate(data) if idx not in matched_set]
        display_data = matched_data + unmatched_data

        row_count = len(display_data)
        self.db_table.setRowCount(row_count)

        for row_idx in range(row_count):
            row_data = display_data[row_idx]
            is_matched = row_idx < len(matched_data)  # 前len(matched_data)行是匹配的
            # 第一列：复选框使用QCheckBox居中显示
            checkbox_widget = QWidget()
            checkbox_layout = QHBoxLayout(checkbox_widget)
            checkbox_layout.setContentsMargins(0, 0, 0, 0)
            checkbox_layout.setAlignment(Qt.AlignCenter)
            checkbox = QCheckBox()
            checkbox.setChecked(is_matched)
            if is_matched:
                checkbox.setStyleSheet("QCheckBox { margin-left: 15px; background-color: rgb(255, 100, 100); }")
            else:
                checkbox.setStyleSheet("QCheckBox { margin-left: 15px; }")
            checkbox_layout.addWidget(checkbox)
            self.db_table.setCellWidget(row_idx, 0, checkbox_widget)
            # 数据列
            for col_idx in range(col_count):
                key = headers[col_idx]
                value = row_data.get(key, '') if isinstance(row_data, dict) else ''
                if value is None:
                    value = ''
                item = QTableWidgetItem(str(value))
                # 匹配的记录高亮显示
                if is_matched:
                    item.setBackground(red_fill)
                self.db_table.setItem(row_idx, col_idx + 1, item)

        # 保存红色高亮行的原始索引
        self.red_highlight_indices = set(indices)

        for col_idx in range(col_count + 1):
            if col_idx == 0:
                self.db_table.setColumnWidth(col_idx, 50)
            else:
                self.db_table.resizeColumnToContents(col_idx)
                if col_idx > 0 and headers[col_idx - 1] == '物料描述':
                    self.db_table.setColumnWidth(col_idx, 300)

    def delete_shipment_matched(self):
        """删除已显示的匹配记录或选中的行"""
        print("delete_shipment_matched 被调用")
        # 获取选中的行索引（复选框勾选的）
        checked_indices = self.get_selected_rows_indices()
        print(f"checked_indices: {checked_indices}")

        if checked_indices:
            # 复选框删除模式
            checked_count = len(checked_indices)

            reply = QMessageBox.question(self, "确认删除",
                f"确定要删除这 {checked_count} 条选中的记录吗？\n\n此操作不可撤销！",
                QMessageBox.Yes | QMessageBox.No)

            if reply == QMessageBox.Yes:
                self.db_all_data = [row for idx, row in enumerate(self.db_all_data)
                                    if idx not in checked_indices]
                self.refresh_db_table()
                self.save_db_data()
                QMessageBox.information(self, "完成", f"已删除 {checked_count} 条记录")
                return

        # 出货对比删除模式
        if not hasattr(self, 'shipment_matched_indices') or not self.shipment_matched_indices:
            QMessageBox.warning(self, "提示", "请先勾选要删除的行\n或者导入出货Excel进行匹配对比")
            return

        matched_count = len(self.shipment_matched_indices)

        reply = QMessageBox.question(self, "确认删除",
            f"确定要删除这 {matched_count} 条匹配记录吗？\n\n此操作不可撤销！",
            QMessageBox.Yes | QMessageBox.No)

        if reply == QMessageBox.Yes:
            # 使用索引删除匹配的行
            self.db_all_data = [row for idx, row in enumerate(self.db_all_data)
                                if idx not in self.shipment_matched_indices]

            # 刷新表格
            self.refresh_db_table()
            self.save_db_data()

            # 清空对比结果状态
            self.shipment_matched_indices = []
            self.red_highlight_indices = set()
            self.delete_shipment_btn.setEnabled(False)

            QMessageBox.information(self, "完成", f"已删除 {matched_count} 条匹配记录")

    def export_shipment_data(self):
        """导出主表格数据为出货Excel格式（只导出勾选行）"""
        if not hasattr(self, 'db_all_data') or not self.db_all_data:
            QMessageBox.warning(self, "提示", "没有数据可导出")
            return

        checked_indices = set(self.get_selected_rows_indices())

        if not checked_indices:
            QMessageBox.warning(self, "提示", "请勾选要导出的行")
            return

        # 生成文件名：导出出货Excel_北京时间
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"导出出货Excel_{timestamp}.xlsx"
        default_path = os.path.join(self.output_dir if hasattr(self, 'output_dir') else get_app_dir(), default_filename)

        filename, _ = QFileDialog.getSaveFileName(
            self, "导出出货数据", default_path, "Excel files (*.xlsx)"
        )
        if filename:
            headers = self.current_db_headers
            data = self.db_all_data
            export_data = [data[idx] for idx in checked_indices if idx < len(data)]
            self.excel_processor.export_to_excel(export_data, filename)
            QMessageBox.information(self, "完成", f"已导出 {len(export_data)} 条数据到:\n{filename}")

    def refresh_db_table(self):
        """刷新数据库表格显示"""
        if not hasattr(self, 'db_all_data'):
            return

        data = self.db_all_data
        headers = self.current_db_headers

        col_count = len(headers)
        row_count = len(data)

        self.db_table.setColumnCount(col_count + 1)
        self.db_table.setRowCount(row_count)

        # 第一列：复选框列
        select_header_item = QTableWidgetItem("选择")
        select_header_item.setTextAlignment(Qt.AlignCenter)
        self.db_table.setHorizontalHeaderItem(0, select_header_item)
        for col_idx, header in enumerate(headers):
            item = QTableWidgetItem(str(header) if header else '')
            self.db_table.setHorizontalHeaderItem(col_idx + 1, item)

        for row_idx in range(row_count):
            row_data = data[row_idx]
            # 第一列：复选框使用QCheckBox居中显示
            checkbox_widget = QWidget()
            checkbox_layout = QHBoxLayout(checkbox_widget)
            checkbox_layout.setContentsMargins(0, 0, 0, 0)
            checkbox_layout.setAlignment(Qt.AlignCenter)
            checkbox = QCheckBox()
            checkbox.setStyleSheet("QCheckBox { margin-left: 15px; }")
            checkbox_layout.addWidget(checkbox)
            self.db_table.setCellWidget(row_idx, 0, checkbox_widget)
            # 数据列
            for col_idx in range(col_count):
                key = headers[col_idx]
                value = row_data.get(key, '') if isinstance(row_data, dict) else ''
                if value is None:
                    value = ''
                self.db_table.setItem(row_idx, col_idx + 1, QTableWidgetItem(str(value)))

        for col_idx in range(col_count + 1):
            if col_idx == 0:
                self.db_table.setColumnWidth(col_idx, 50)
            else:
                self.db_table.resizeColumnToContents(col_idx)
                if col_idx > 0 and headers[col_idx - 1] == '物料描述':
                    self.db_table.setColumnWidth(col_idx, 300)

        self.db_info_label.setText(f"共 {len(data)} 条数据")

    def save_db_data(self):
        """保存当前数据到本地文件"""
        import json
        try:
            save_data = {
                'headers': self.current_db_headers if hasattr(self, 'current_db_headers') else [],
                'data': self.db_all_data if hasattr(self, 'db_all_data') else []
            }
            with open(self.db_data_file, 'w', encoding='utf-8') as f:
                json.dump(save_data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"保存数据失败: {e}")

    def load_db_data(self):
        """从本地文件加载数据"""
        import json
        try:
            if os.path.exists(self.db_data_file):
                save_data = self._read_json_file(self.db_data_file)
                if save_data is None:
                    return

                headers = save_data.get('headers', [])
                data = save_data.get('data', [])

                print(f"load_db_data: headers数量={len(headers)}, data数量={len(data)}")

                if data and headers:
                    self.db_all_data = data
                    self.db_original_data = data
                    self.current_db_headers = headers
                    self.current_db_file = "已加载保存的数据"
                    self.red_highlight_indices = set()

                    # 显示数据
                    col_count = len(headers)
                    row_count = len(data)
                    print(f"设置表格: {col_count + 1} 列, {row_count} 行")
                    self.db_table.setColumnCount(col_count + 1)
                    self.db_table.setRowCount(row_count)

                    # 第一列：复选框列（表头不显示文字，仅作占位）
                    self.db_table.setHorizontalHeaderItem(0, QTableWidgetItem(""))
                    for col_idx, header in enumerate(headers):
                        item = QTableWidgetItem(str(header) if header else '')
                        self.db_table.setHorizontalHeaderItem(col_idx + 1, item)

                    for row_idx in range(row_count):
                        row_data = data[row_idx]
                        # 第一列：复选框使用QCheckBox居中显示
                        checkbox_widget = QWidget()
                        checkbox_layout = QHBoxLayout(checkbox_widget)
                        checkbox_layout.setContentsMargins(0, 0, 0, 0)
                        checkbox_layout.setAlignment(Qt.AlignCenter)
                        checkbox = QCheckBox()
                        checkbox.setStyleSheet("QCheckBox { margin-left: 15px; }")
                        checkbox_layout.addWidget(checkbox)
                        self.db_table.setCellWidget(row_idx, 0, checkbox_widget)
                        # 数据列
                        for col_idx in range(col_count):
                            key = headers[col_idx]
                            value = row_data.get(key, '') if isinstance(row_data, dict) else ''
                            if value is None:
                                value = ''
                            self.db_table.setItem(row_idx, col_idx + 1, QTableWidgetItem(str(value)))

                    print(f"数据填充完成，准备更新标签")
                    for col_idx in range(col_count + 1):
                        if col_idx == 0:
                            self.db_table.setColumnWidth(col_idx, 50)
                        else:
                            self.db_table.resizeColumnToContents(col_idx)
                            if col_idx > 0 and headers[col_idx - 1] == '物料描述':
                                self.db_table.setColumnWidth(col_idx, 300)

                    print(f"设置标签文本: 已加载保存的数据 - 共 {len(data)} 条")
                    self.db_info_label.setText(f"已加载保存的数据 - 共 {len(data)} 条")
                    self.export_search_btn.setEnabled(True)
                    print(f"已加载 {len(data)} 条数据")

                    # 强制更新表格显示
                    self.db_table.viewport().update()
                    self.db_table.update()
                    # 滚动到顶部
                    self.db_table.scrollToTop()
                    self.db_table.horizontalScrollBar().setValue(0)
        except Exception as e:
            print(f"加载数据失败: {e}")

    def save_batch_import_data(self):
        """保存批量导入数据到本地文件"""
        import json
        try:
            save_data = {
                'headers': self.batch_import_headers if hasattr(self, 'batch_import_headers') else [],
                'data': self.batch_import_data if hasattr(self, 'batch_import_data') else []
            }
            with open(self.batch_import_data_file, 'w', encoding='utf-8') as f:
                json.dump(save_data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"保存批量导入数据失败: {e}")

    def load_batch_import_data(self):
        """从本地文件加载批量导入数据"""
        import json
        try:
            if os.path.exists(self.batch_import_data_file):
                save_data = self._read_json_file(self.batch_import_data_file)
                if save_data is None:
                    return

                headers = save_data.get('headers', [])
                data = save_data.get('data', [])

                if data and headers:
                    self.batch_import_headers = headers
                    self.batch_import_data = data

                    print(f"加载批量导入数据: headers数量={len(headers)}, data数量={len(data)}")

                    # 显示数据
                    self.batch_refresh_table()
                    self.batch_info_label.setText(f"已加载保存的数据 - 共 {len(data)} 条")
        except Exception as e:
            print(f"加载批量导入数据失败: {e}")

    def save_output_dir_config(self):
        """保存输出目录配置到本地文件"""
        import json
        try:
            config_data = {
                'output_dir': self.output_dir,
                'category_output_dir': self.category_output_dir,
                'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            with open(self.output_dir_config_file, 'w', encoding='utf-8') as f:
                json.dump(config_data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"保存输出目录配置失败: {e}")

    def load_output_dir_config(self):
        """从本地文件加载输出目录配置"""
        import json
        try:
            if os.path.exists(self.output_dir_config_file):
                config_data = self._read_json_file(self.output_dir_config_file)
                if config_data is None:
                    return

                output_dir = config_data.get('output_dir', '')
                category_output_dir = config_data.get('category_output_dir', '')

                if output_dir and os.path.isdir(output_dir):
                    self.output_dir = output_dir
                    print(f"加载输出目录配置: {output_dir}")
                    # 更新UI上的输出目录显示（如果控件已创建）
                    if hasattr(self, 'output_edit') and self.output_edit:
                        self.output_edit.setText(output_dir)

                if category_output_dir and os.path.isdir(category_output_dir):
                    self.category_output_dir = category_output_dir
                    print(f"加载分类输出目录配置: {category_output_dir}")
                    # 更新UI上的分类输出目录显示（如果控件已创建）
                    if hasattr(self, 'category_output_edit') and self.category_output_edit:
                        self.category_output_edit.setText(category_output_dir)

                return True
        except Exception as e:
            print(f"加载输出目录配置失败: {e}")
        return False

    def filter_db_data(self, text):
        if not hasattr(self, 'db_all_data') or not self.db_all_data:
            return

        headers = self.current_db_headers

        keywords = [k.strip() for k in text.split(',') if k.strip()]
        if not keywords:
            data = self.db_all_data
            self.db_search_keywords = []
        else:
            self.db_search_keywords = keywords

            # 查找关键列
            col_names = {
                'material': None,
                'sales_order': None,
                'sales_line': None,
                'internal_order': None
            }
            for h in headers:
                if h:
                    h_str = str(h)
                    if '物料号' in h_str or '物料编码' in h_str:
                        col_names['material'] = h
                    elif '销售订单' in h_str and '行号' not in h_str:
                        col_names['sales_order'] = h
                    elif '销售订单行号' in h_str:
                        col_names['sales_line'] = h
                    elif '内需单号' in h_str:
                        col_names['internal_order'] = h

            # 提取精确匹配条件列表
            match_conditions = []
            for kw in keywords:
                # 关键词可能是完整的匹配键（物料号|销售订单|销售订单行号|内需单号）
                if '|' in kw:
                    parts = [p.strip() for p in kw.split('|')]
                    match_conditions.append({
                        'material': parts[0] if len(parts) > 0 else '',
                        'sales_order': parts[1] if len(parts) > 1 else '',
                        'sales_line': parts[2] if len(parts) > 2 else '',
                        'internal_order': parts[3] if len(parts) > 3 else ''
                    })
                else:
                    match_conditions.append({
                        'material': kw,
                        'sales_order': '',
                        'sales_line': '',
                        'internal_order': ''
                    })

            data = []
            for row in self.db_all_data:
                material_val = str(row.get(col_names['material'], '')).strip()
                sales_order_val = str(row.get(col_names['sales_order'], '')).strip()
                sales_line_val = str(row.get(col_names['sales_line'], '')).strip()
                internal_order_val = str(row.get(col_names['internal_order'], '')).strip()

                matched = False
                for cond in match_conditions:
                    # 必须有物料号
                    if not cond['material']:
                        continue

                    # 物料号模糊匹配（包含）
                    if cond['material'].lower() not in material_val.lower():
                        continue

                    # 检查销售订单（如果条件中有值，模糊匹配）
                    if cond['sales_order'] and cond['sales_order'].lower() not in sales_order_val.lower():
                        continue

                    # 检查销售订单行号（如果条件中有值，模糊匹配）
                    if cond['sales_line'] and cond['sales_line'].lower() not in sales_line_val.lower():
                        continue

                    # 检查内需单号（如果条件中有值，模糊匹配）
                    if cond['internal_order'] and cond['internal_order'].lower() not in internal_order_val.lower():
                        continue

                    # 所有条件都满足
                    matched = True
                    break

                if matched:
                    data.append(row)

        col_count = len(headers)
        row_count = len(data)

        self.db_table.setRowCount(row_count)

        for row_idx in range(row_count):
            row_data = data[row_idx]
            # 第一列：复选框使用QCheckBox居中显示
            checkbox_widget = QWidget()
            checkbox_layout = QHBoxLayout(checkbox_widget)
            checkbox_layout.setContentsMargins(0, 0, 0, 0)
            checkbox_layout.setAlignment(Qt.AlignCenter)
            checkbox = QCheckBox()
            if keywords:
                checkbox.setStyleSheet("QCheckBox { margin-left: 15px; background-color: rgb(173, 216, 230); }")
            else:
                checkbox.setStyleSheet("QCheckBox { margin-left: 15px; }")
            checkbox_layout.addWidget(checkbox)
            self.db_table.setCellWidget(row_idx, 0, checkbox_widget)
            # 数据列
            for col_idx in range(col_count):
                key = headers[col_idx]
                value = row_data.get(key, '') if isinstance(row_data, dict) else ''
                if value is None:
                    value = ''
                item = QTableWidgetItem(str(value))
                if keywords:
                    item.setBackground(QColor(173, 216, 230, 153))
                self.db_table.setItem(row_idx, col_idx + 1, item)

        for col_idx in range(col_count + 1):
            col_idx_header = headers[col_idx - 1] if col_idx > 0 else '选择'
            col_width = 50

            if col_idx == 0:
                col_width = 50
            elif col_idx_header == '物料描述':
                col_width = 350
            elif col_idx_header == '物料号':
                col_width = 150
            elif col_idx_header in ['序号', '销售订单行号']:
                col_width = 60
            elif col_idx_header in ['总缺料', '数量']:
                col_width = 80
            elif col_idx_header in ['销售订单', '内需单号', '订单号']:
                col_width = 150
            elif col_idx_header in ['供方', '采购组', '采购组名称']:
                col_width = 120
            elif col_idx_header in ['下单时间', '送货日期']:
                col_width = 100

            self.db_table.setColumnWidth(col_idx, col_width)

        self.db_table.horizontalHeader().setStretchLastSection(True)

        # 保存当前搜索结果数据（无论有没有搜索关键词都保存）
        self.current_search_data = data
        self.current_search_headers = headers

        # 更新导出搜索结果按钮状态
        try:
            if data:
                self.export_search_btn.setEnabled(True)
            else:
                self.export_search_btn.setEnabled(False)
        except Exception as e:
            print(f"启用按钮失败: {e}")

        # 保存当前显示的数据用于导出
        self.current_search_data = data

        self.db_info_label.setText(f"显示 {row_count} / {len(self.db_all_data)} 条数据")

    def import_search_conditions(self):
        filename, _ = QFileDialog.getOpenFileName(
            self, "选择搜索条件文件", "", "Excel files (*.xlsx *.xls)"
        )
        if not filename:
            return

        try:
            search_headers, search_data = self.excel_processor.read_excel(filename)

            # 查找搜索条件文件中的关键列
            col_names = {
                'material': None,
                'sales_order': None,
                'sales_line': None,
                'internal_order': None
            }
            for h in search_headers:
                if h:
                    h_str = str(h)
                    if '物料号' in h_str or '物料编码' in h_str:
                        col_names['material'] = h
                    elif '销售订单' in h_str and '行号' not in h_str:
                        col_names['sales_order'] = h
                    elif '销售订单行号' in h_str:
                        col_names['sales_line'] = h
                    elif '内需单号' in h_str:
                        col_names['internal_order'] = h

            # 构建完整匹配键列表
            match_keys = []
            for row in search_data:
                material_val = row.get(col_names['material'])
                material = str(material_val if material_val is not None else '').strip()
                if material:
                    sales_order_val = row.get(col_names['sales_order'])
                    sales_order = str(sales_order_val if sales_order_val is not None else '').strip()
                    sales_line_val = row.get(col_names['sales_line'])
                    sales_line = str(sales_line_val if sales_line_val is not None else '').strip()
                    internal_order_val = row.get(col_names['internal_order'])
                    internal_order = str(internal_order_val if internal_order_val is not None else '').strip()
                    # 构建完整匹配键格式：物料号|销售订单|销售订单行号|内需单号
                    match_key = f"{material}|{sales_order}|{sales_line}|{internal_order}"
                    match_keys.append(match_key)

            if match_keys:
                # 保存原始数据库备份（如果有的话）
                if not hasattr(self, 'db_all_data') or not self.db_all_data:
                    QMessageBox.warning(self, "提示", "请先导入数据库Excel文件，再导入搜索条件")
                    return

                # 保存原始数据库数据
                self.db_original_data = self.db_all_data
                self.db_original_headers = self.current_db_headers

                # 使用完整匹配键格式搜索
                search_text = ','.join(match_keys[:100])
                self.db_search_edit.setText(search_text)
                self.filter_db_data(search_text)
                # 确保导出搜索结果按钮可用
                if hasattr(self, 'export_search_btn'):
                    self.export_search_btn.setEnabled(True)
                QMessageBox.information(self, "完成", f"已导入 {len(match_keys)} 个搜索条件\n匹配结果已高亮显示")
            else:
                QMessageBox.warning(self, "提示",
                    f"文件中没有找到有效的物料号\n\n"
                    f"文件: {os.path.basename(filename)}\n"
                    f"headers数量: {len(search_headers)}\n"
                    f"data数量: {len(search_data)}\n"
                    f"找到的列: {col_names}")

        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "错误", f"导入失败: {str(e)}")

    def clear_search(self):
        self.db_search_edit.setText('')
        # 如果有原始数据备份，恢复原始数据
        if hasattr(self, 'db_original_data') and self.db_original_data:
            self.db_all_data = self.db_original_data
            if hasattr(self, 'db_original_headers'):
                self.current_db_headers = self.db_original_headers
        self.filter_db_data('')

    def export_db_table(self):
        if self.db_table.rowCount() == 0:
            QMessageBox.warning(self, "提示", "没有数据可导出")
            return

        # 使用时间戳生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"导出Excel_{timestamp}.xlsx"
        save_dir = self.output_dir if hasattr(self, 'output_dir') else get_app_dir()
        default_path = os.path.join(save_dir, default_filename)

        filename, _ = QFileDialog.getSaveFileName(
            self, "导出Excel", default_path, "Excel files (*.xlsx)"
        )
        if filename:
            # 使用原始数据导出，保持列顺序
            if hasattr(self, 'db_all_data') and self.db_all_data:
                export_data = self.db_all_data
                headers = self.current_db_headers
            else:
                # 如果没有原始数据，从表格获取
                export_data = []
                headers = []
                for col in range(self.db_table.columnCount()):
                    headers.append(self.db_table.horizontalHeaderItem(col).text())
                for row in range(self.db_table.rowCount()):
                    row_data = {}
                    for col in range(self.db_table.columnCount()):
                        header = headers[col] if col < len(headers) else f"列{col}"
                        item = self.db_table.item(row, col)
                        row_data[header] = item.text() if item else ''
                    export_data.append(row_data)

            self.excel_processor.export_to_excel(export_data, filename)
            QMessageBox.information(self, "完成", f"已导出到:\n{filename}")

    def clear_db_table(self):
        reply = QMessageBox.question(self, "确认清空",
            "确定要清空所有数据吗？\n\n此操作不可撤销！",
            QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        # 清空表格数据
        self.db_table.setRowCount(0)
        self.db_table.setColumnCount(0)
        self.db_table.setHorizontalHeaderLabels([])

        # 清空数据变量
        if hasattr(self, 'db_all_data'):
            self.db_all_data = []
        if hasattr(self, 'db_original_data'):
            self.db_original_data = []
        if hasattr(self, 'current_search_data'):
            self.current_search_data = []

        # 清空出货对比状态
        if hasattr(self, 'shipment_matched_indices'):
            self.shipment_matched_indices = []

        # 隐藏出货对比结果表格
        if hasattr(self, 'shipment_compare_label'):
            self.shipment_compare_label.setVisible(False)
        if hasattr(self, 'shipment_compare_table'):
            self.shipment_compare_table.setVisible(False)
            self.shipment_compare_table.setRowCount(0)

        # 禁用按钮
        if hasattr(self, 'export_search_btn'):
            self.export_search_btn.setEnabled(False)
        if hasattr(self, 'delete_shipment_btn'):
            self.delete_shipment_btn.setEnabled(False)

        # 更新状态
        self.db_info_label.setText("请选择Excel文件导入")

    def export_search_result(self):
        # 从表格直接获取当前显示的数据
        if self.db_table.rowCount() == 0:
            QMessageBox.warning(self, "提示", "没有数据可导出")
            return

        # 使用时间戳生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"导出搜索结果_{timestamp}.xlsx"
        save_dir = self.output_dir if hasattr(self, 'output_dir') else get_app_dir()
        default_path = os.path.join(save_dir, default_filename)

        filename, _ = QFileDialog.getSaveFileName(
            self, "导出搜索结果", default_path, "Excel files (*.xlsx)"
        )
        if filename:
            # 从表格获取当前显示的数据（跳过第一列复选框）
            export_data = []
            for row in range(self.db_table.rowCount()):
                row_data = {}
                for col in range(1, self.db_table.columnCount()):
                    header = self.db_table.horizontalHeaderItem(col).text()
                    item = self.db_table.item(row, col)
                    row_data[header] = item.text() if item else ''
                export_data.append(row_data)
            headers = self.current_search_headers
            self.excel_processor.export_to_excel(export_data, filename)
            QMessageBox.information(self, "完成", f"已导出搜索结果到:\n{filename}")

    def open_excel_fullscreen(self):
        filename, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", "", "Excel files (*.xlsx *.xls)"
        )
        if not filename:
            return

        try:
            headers, data = self.excel_processor.read_excel(filename)

            dialog = QDialog(self)
            dialog.setWindowTitle(f"Excel数据 - {os.path.basename(filename)}")
            dialog.setGeometry(50, 50, 1400, 900)
            dialog.setStyleSheet("background-color: #F9FAFB;")

            layout = QVBoxLayout()
            layout.setContentsMargins(20, 20, 20, 20)

            title = QLabel(f"共 {len(data)} 条数据")
            title.setStyleSheet("font-size: 20px; font-weight: bold; color: #374151; padding-bottom: 15px;")
            layout.addWidget(title)

            table = QTableWidget()
            table.setColumnCount(len(headers))
            table.setHorizontalHeaderLabels(headers)
            table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            table.setStyleSheet("""
                QTableWidget {
                    border: 1px solid #E5E7EB;
                    border-radius: 8px;
                    gridline-color: #E5E7EB;
                    background-color: #FFFFFF;
                }
                QTableWidget::item {
                    padding: 8px;
                }
                QHeaderView::section {
                    background-color: #4472C4;
                    color: white;
                    padding: 10px;
                    border: none;
                    font-weight: bold;
                    font-size: 13px;
                }
            """)
            table.setAlternatingRowColors(True)
            table.setEditTriggers(QAbstractItemView.NoEditTriggers)

            table.setRowCount(len(data))
            for row_idx, row in enumerate(data):
                for col_idx, header in enumerate(headers):
                    value = row.get(header, '')
                    table.setItem(row_idx, col_idx, QTableWidgetItem(str(value) if value is not None else ''))

            layout.addWidget(table)

            btn_layout = QHBoxLayout()
            export_btn = QPushButton("导出Excel")
            export_btn.setStyleSheet(self.btn_secondary_style)
            export_btn.clicked.connect(lambda: self._export_excel_table(table, headers))
            close_btn = QPushButton("关闭")
            close_btn.setStyleSheet(self.btn_secondary_style)
            close_btn.clicked.connect(dialog.close)
            btn_layout.addWidget(export_btn)
            btn_layout.addStretch()
            btn_layout.addWidget(close_btn)
            layout.addLayout(btn_layout)

            dialog.setLayout(layout)
            dialog.exec_()

        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取失败: {str(e)}")

    def _export_excel_table(self, table, headers):
        filename, _ = QFileDialog.getSaveFileName(
            self, "导出Excel", "", "Excel files (*.xlsx)"
        )
        if filename:
            export_data = []
            for row in range(table.rowCount()):
                row_data = {}
                for col in range(table.columnCount()):
                    header = headers[col] if col < len(headers) else f"列{col}"
                    item = table.item(row, col)
                    row_data[header] = item.text() if item else ''
                export_data.append(row_data)
            self.excel_processor.export_to_excel(export_data, filename)
            QMessageBox.information(self, "完成", f"已导出到:\n{filename}")

    def open_level2_fullscreen(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("一级表格 - 全屏浏览")
        dialog.setGeometry(50, 50, 1400, 900)
        dialog.setStyleSheet("background-color: #F9FAFB;")

        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        title = QLabel("一级表格数据")
        title.setStyleSheet("font-size: 20px; font-weight: bold; color: #374151; padding-bottom: 15px;")
        layout.addWidget(title)

        table = QTableWidget()
        table.setColumnCount(8)
        table.setHorizontalHeaderLabels(['ID', '物料号', '物料描述', '供方', '总缺料', '分类', '序号', '订单号'])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #E5E7EB;
                border-radius: 8px;
                gridline-color: #E5E7EB;
                background-color: #FFFFFF;
            }
            QTableWidget::item {
                padding: 8px;
            }
            QHeaderView::section {
                background-color: #4472C4;
                color: white;
                padding: 10px;
                border: none;
                font-weight: bold;
                font-size: 13px;
            }
        """)
        table.setAlternatingRowColors(True)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)

        data = self.db.get_level1_all()
        table.setRowCount(len(data))
        for row_idx, row in enumerate(data):
            table.setItem(row_idx, 0, QTableWidgetItem(str(row.get('ID', ''))))
            table.setItem(row_idx, 1, QTableWidgetItem(str(row.get('物料号', ''))))
            table.setItem(row_idx, 2, QTableWidgetItem(str(row.get('物料描述', ''))))
            table.setItem(row_idx, 3, QTableWidgetItem(str(row.get('供方', ''))))
            table.setItem(row_idx, 4, QTableWidgetItem(str(row.get('总缺料', ''))))
            table.setItem(row_idx, 5, QTableWidgetItem(str(row.get('分类', ''))))
            table.setItem(row_idx, 6, QTableWidgetItem(str(row.get('序号', ''))))
            table.setItem(row_idx, 7, QTableWidgetItem(str(row.get('订单号', ''))))

        layout.addWidget(table)

        btn_layout = QHBoxLayout()
        refresh_btn = QPushButton("刷新")
        refresh_btn.setStyleSheet(self.btn_secondary_style)
        refresh_btn.clicked.connect(lambda: self._refresh_fullscreen_table(table, self.db.get_level1_all, [('ID', ''), ('物料号', ''), ('物料描述', ''), ('供方', ''), ('总缺料', ''), ('分类', ''), ('序号', ''), ('订单号', '')]))
        export_btn = QPushButton("导出Excel")
        export_btn.setStyleSheet(self.btn_secondary_style)
        export_btn.clicked.connect(lambda: self._export_fullscreen_table(table))
        close_btn = QPushButton("关闭")
        close_btn.setStyleSheet(self.btn_secondary_style)
        close_btn.clicked.connect(dialog.close)
        btn_layout.addWidget(refresh_btn)
        btn_layout.addWidget(export_btn)
        btn_layout.addStretch()
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)

        dialog.setLayout(layout)
        dialog.exec_()

    def open_level2_fullscreen(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("二级表格 - 全屏浏览")
        dialog.setGeometry(50, 50, 1400, 900)
        dialog.setStyleSheet("background-color: #F9FAFB;")

        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        title = QLabel("二级表格数据")
        title.setStyleSheet("font-size: 20px; font-weight: bold; color: #374151; padding-bottom: 15px;")
        layout.addWidget(title)

        table = QTableWidget()
        table.setColumnCount(8)
        table.setHorizontalHeaderLabels(['ID', '一级ID', '物料号', '物料描述', '序号', '订单号', '总缺料', '分类'])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #E5E7EB;
                border-radius: 8px;
                gridline-color: #E5E7EB;
                background-color: #FFFFFF;
            }
            QTableWidget::item {
                padding: 8px;
            }
            QHeaderView::section {
                background-color: #4472C4;
                color: white;
                padding: 10px;
                border: none;
                font-weight: bold;
                font-size: 13px;
            }
        """)
        table.setAlternatingRowColors(True)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)

        data = self.db.get_level2_all()
        table.setRowCount(len(data))
        for row_idx, row in enumerate(data):
            table.setItem(row_idx, 0, QTableWidgetItem(str(row.get('ID', ''))))
            table.setItem(row_idx, 1, QTableWidgetItem(str(row.get('一级ID', ''))))
            table.setItem(row_idx, 2, QTableWidgetItem(str(row.get('物料号', ''))))
            table.setItem(row_idx, 3, QTableWidgetItem(str(row.get('物料描述', ''))))
            table.setItem(row_idx, 4, QTableWidgetItem(str(row.get('序号', ''))))
            table.setItem(row_idx, 5, QTableWidgetItem(str(row.get('订单号', ''))))
            table.setItem(row_idx, 6, QTableWidgetItem(str(row.get('总缺料', ''))))
            table.setItem(row_idx, 7, QTableWidgetItem(str(row.get('分类', ''))))

        layout.addWidget(table)

        btn_layout = QHBoxLayout()
        refresh_btn = QPushButton("刷新")
        refresh_btn.setStyleSheet(self.btn_secondary_style)
        refresh_btn.clicked.connect(lambda: self._refresh_fullscreen_table(table, self.db.get_level2_all, [('ID', ''), ('一级ID', ''), ('物料号', ''), ('物料描述', ''), ('序号', ''), ('订单号', ''), ('总缺料', ''), ('分类', '')]))
        export_btn = QPushButton("导出Excel")
        export_btn.setStyleSheet(self.btn_secondary_style)
        export_btn.clicked.connect(lambda: self._export_fullscreen_table(table))
        import_btn = QPushButton("导入Excel")
        import_btn.setStyleSheet(self.btn_secondary_style)
        import_btn.clicked.connect(lambda: self._import_to_fullscreen(table, self.db.get_level2_all, self.import_level2))
        close_btn = QPushButton("关闭")
        close_btn.setStyleSheet(self.btn_secondary_style)
        close_btn.clicked.connect(dialog.close)
        btn_layout.addWidget(refresh_btn)
        btn_layout.addWidget(export_btn)
        btn_layout.addWidget(import_btn)
        btn_layout.addStretch()
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)

        dialog.setLayout(layout)
        dialog.exec_()

    def _refresh_fullscreen_table(self, table, data_func, columns):
        table.setRowCount(0)
        data = data_func()
        table.setRowCount(len(data))
        for row_idx, row in enumerate(data):
            for col_idx, (key, default) in enumerate(columns):
                table.setItem(row_idx, col_idx, QTableWidgetItem(str(row.get(key, default))))

    def _export_fullscreen_table(self, table):
        filename, _ = QFileDialog.getSaveFileName(
            self, "导出Excel", "", "Excel files (*.xlsx)"
        )
        if filename:
            export_data = []
            for row in range(table.rowCount()):
                row_data = {}
                for col in range(table.columnCount()):
                    header = table.horizontalHeaderItem(col).text()
                    item = table.item(row, col)
                    row_data[header] = item.text() if item else ''
                export_data.append(row_data)
            self.excel_processor.export_to_excel(export_data, filename)
            QMessageBox.information(self, "完成", f"已导出到:\n{filename}")

    def _export_fullscreen_fullscreen(self, table, title):
        self._export_fullscreen_table(table)

    def _import_to_fullscreen(self, table, data_func, import_func):
        import_func()
        table.setRowCount(0)
        data = data_func()
        columns = []
        if data:
            for key in data[0].keys():
                columns.append((key, ''))
        if columns:
            self._refresh_fullscreen_table(table, data_func, columns)

    def import_shipment(self):
        filepath = self.shipment_file_edit.text()
        if not filepath:
            QMessageBox.critical(self, "错误", "请选择出货表格文件")
            return

        try:
            headers, data = self.excel_processor.read_excel(filepath)

            self.shipment_table.setRowCount(0)

            shipment_data = []
            for row in data:
                material_number = row.get('物料号', '')
                quantity = row.get('总缺料', 0) or row.get('数量', 0)
                if material_number:
                    shipment_data.append({
                        '物料号': material_number,
                        '数量': quantity,
                        '订单号': row.get('销售订单', ''),
                        '备注': '待确认删除'
                    })
                    row_idx = self.shipment_table.rowCount()
                    self.shipment_table.insertRow(row_idx)
                    self.shipment_table.setItem(row_idx, 0, QTableWidgetItem(str(material_number)))
                    self.shipment_table.setItem(row_idx, 1, QTableWidgetItem(str(quantity)))
                    self.shipment_table.setItem(row_idx, 2, QTableWidgetItem(str(row.get('销售订单', ''))))

            self.current_shipment_data = shipment_data

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = os.path.join(self.output_edit.text(), f"出货记录高亮_{timestamp}.xlsx")
            self.excel_processor.save_shipment_highlighted(shipment_data, output_file)

            QMessageBox.information(self, "完成", f"已导入 {len(shipment_data)} 条记录\n高亮文件已保存到:\n{output_file}")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"导入失败: {str(e)}")

    def delete_shipment(self):
        selected_rows = self.shipment_table.selectedItems()
        if not selected_rows:
            QMessageBox.warning(self, "警告", "请先选中要删除的记录")
            return

        if not hasattr(self, 'current_shipment_data'):
            QMessageBox.warning(self, "警告", "请先导入出货表格")
            return

        reply = QMessageBox.question(self, "确认", "确认删除选中的出货记录？",
                                      QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            rows_to_delete = set()
            for item in selected_rows:
                rows_to_delete.add(item.row())

            for row in sorted(rows_to_delete, reverse=True):
                material_number = self.shipment_table.item(row, 0).text()
                for shipment in self.current_shipment_data:
                    if shipment['物料号'] == material_number:
                        self.db.delete_shipment(material_number)
                        break
                self.shipment_table.removeRow(row)

            QMessageBox.information(self, "完成", "已删除选中的出货记录")

    def export_shipment_highlighted(self):
        if not hasattr(self, 'current_shipment_data'):
            QMessageBox.warning(self, "警告", "没有可导出的数据")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "保存高亮表格", "", "Excel files (*.xlsx)"
        )
        if filename:
            self.excel_processor.save_shipment_highlighted(self.current_shipment_data, filename)
            QMessageBox.information(self, "完成", f"高亮表格已导出到:\n{filename}")

    def compare_shipment(self):
        shipment_path = self.shipment_compare_edit.text()
        target_path = self.shipment_target_edit.text()

        if not shipment_path or not target_path:
            QMessageBox.critical(self, "错误", "请选择出货表格和对比表格")
            return

        try:
            _, shipment_data = self.excel_processor.read_excel(shipment_path)
            _, target_data = self.excel_processor.read_excel(target_path)

            shipment_dict = {}
            for row in shipment_data:
                material = row.get('物料号', '')
                quantity = row.get('总缺料', 0) or row.get('数量', 0) or 0
                if material:
                    shipment_dict[material] = shipment_dict.get(material, 0) + quantity

            target_dict = {}
            for row in target_data:
                material = row.get('物料号', '')
                quantity = row.get('总缺料', 0) or row.get('数量', 0) or 0
                if material:
                    target_dict[material] = target_dict.get(material, 0) + quantity

            all_materials = set(shipment_dict.keys()) | set(target_dict.keys())

            self.shipment_compare_table.setRowCount(0)

            for material in sorted(all_materials):
                shipment_qty = shipment_dict.get(material, 0)
                target_qty = target_dict.get(material, 0)
                diff = shipment_qty - target_qty

                row_idx = self.shipment_compare_table.rowCount()
                self.shipment_compare_table.insertRow(row_idx)
                self.shipment_compare_table.setItem(row_idx, 0, QTableWidgetItem(str(material)))
                self.shipment_compare_table.setItem(row_idx, 1, QTableWidgetItem(str(shipment_qty)))
                self.shipment_compare_table.setItem(row_idx, 2, QTableWidgetItem(str(target_qty)))
                self.shipment_compare_table.setItem(row_idx, 3, QTableWidgetItem(str(diff)))

            QMessageBox.information(self, "完成", f"比对完成！\n共 {len(all_materials)} 种物料")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"比对失败: {str(e)}")

    def import_sequence(self):
        filepath = self.sequence_file_edit.text()
        if not filepath:
            QMessageBox.critical(self, "错误", "请选择序号表格文件")
            return

        try:
            headers, data = self.excel_processor.read_excel(filepath)

            sequence_col = self.sequence_col_edit.text()
            order_col = self.order_col_edit.text()

            count = 0
            for row in data:
                material_number = row.get('物料号', '')
                sequence = row.get(sequence_col, '')
                order_number = row.get(order_col, '')

                if material_number:
                    self.db.update_sequence_order(material_number, sequence, order_number)
                    count += 1

            self.refresh_level1()
            self.refresh_level2()

            QMessageBox.information(self, "完成", f"已导入 {count} 条序号信息")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"导入失败: {str(e)}")

    def load_excel_to_query_table(self):
        """直接将Excel内容显示在主表格上"""
        filepath = self.query_file_edit.text()

        if not filepath:
            QMessageBox.critical(self, "错误", "请先选择 Excel 文件")
            return

        try:
            headers, data = self.excel_processor.read_excel(filepath)

            if not data:
                QMessageBox.warning(self, "警告", "文件中没有数据")
                return

            col_count = len(headers)
            row_count = len(data)

            self.query_table.setColumnCount(col_count)
            self.query_table.setRowCount(row_count)

            for col_idx, header_text in enumerate(headers):
                item = QTableWidgetItem(str(header_text) if header_text else '')
                item.setForeground(QColor("#FFFFFF"))
                item.setTextAlignment(Qt.AlignCenter)
                self.query_table.setHorizontalHeaderItem(col_idx, item)

            self.query_table.horizontalHeader().update()

            for row_idx, row_data in enumerate(data):
                for col_idx, header in enumerate(headers):
                    value = row_data.get(header, '')
                    self.query_table.setItem(row_idx, col_idx, QTableWidgetItem(str(value) if value else ''))

            for col_idx in range(col_count):
                self.query_table.resizeColumnToContents(col_idx)
                w = self.query_table.columnWidth(col_idx)
                if w < 80:
                    self.query_table.setColumnWidth(col_idx, 80)
                elif w > 300:
                    self.query_table.setColumnWidth(col_idx, 300)

            self.query_table.viewport().update()
            self.query_table.horizontalHeader().raise_()
            self.query_table.horizontalScrollBar().setValue(0)

            QMessageBox.information(self, "完成", f"已导入 {row_count} 行数据")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导入失败：{str(e)}")

    def query_and_export(self):
        filepath = self.query_file_edit.text()
        if not filepath:
            QMessageBox.critical(self, "错误", "请选择查询表格文件")
            return

        try:
            headers, data = self.excel_processor.read_excel(filepath)

            material_col = self.query_col_edit.text()

            self.query_table.setRowCount(0)

            query_results = []
            for row_idx, row in enumerate(data):
                material_number = row.get(material_col, '')
                excel_row_num = row_idx + 2  # Excel行号（1行表头+1-indexed）
                if material_number:
                    db_results = self.db.search_by_material_number(material_number)
                    if db_results:
                        for db_row in db_results:
                            query_results.append({
                                '查询状态': '已匹配',
                                '原始行号': excel_row_num,
                                'ID': db_row.get('ID', ''),
                                '物料号': material_number,
                                '物料描述': db_row.get('物料描述', ''),
                                '供方': db_row.get('供方', ''),
                                '总缺料': db_row.get('总缺料', ''),
                                '分类': db_row.get('分类', ''),
                                '序号': db_row.get('序号', ''),
                                '订单号': db_row.get('订单号', '')
                            })
                    else:
                        query_results.append({
                            '查询状态': '未匹配',
                            '原始行号': excel_row_num,
                            'ID': '',
                            '物料号': material_number,
                            '物料描述': '',
                            '供方': '',
                            '总缺料': '',
                            '分类': '',
                            '序号': '',
                            '订单号': ''
                        })
                else:
                    query_results.append({
                        '查询状态': '未匹配',
                        '原始行号': excel_row_num,
                        'ID': '',
                        '物料号': '',
                        '物料描述': '',
                        '供方': '',
                        '总缺料': '',
                        '分类': '',
                        '序号': '',
                        '订单号': ''
                    })

            if query_results:
                result_headers = ['查询状态', '原始行号', 'ID', '物料号', '物料描述', '供方', '总缺料', '分类', '序号', '订单号']
                col_count = len(result_headers)
                row_count = len(query_results)

                self.query_table.setColumnCount(col_count)
                self.query_table.setRowCount(row_count)

                for col_idx, header in enumerate(result_headers):
                    item = QTableWidgetItem(str(header))
                    item.setForeground(QColor("#FFFFFF"))
                    item.setTextAlignment(Qt.AlignCenter)
                    self.query_table.setHorizontalHeaderItem(col_idx, item)
                self.query_table.horizontalHeader().update()

                matched_count = 0
                for row_idx, row_data in enumerate(query_results):
                    status = row_data.get('查询状态', '')
                    is_matched = (status == '已匹配')
                    if is_matched:
                        matched_count += 1

                    values = [
                        status,
                        str(row_data.get('原始行号', '')),
                        str(row_data.get('ID', '')),
                        str(row_data.get('物料号', '')),
                        str(row_data.get('物料描述', ''))[:30] if row_data.get('物料描述') else '',
                        str(row_data.get('供方', '')),
                        str(row_data.get('总缺料', '')),
                        str(row_data.get('分类', '')),
                        str(row_data.get('序号', '')),
                        str(row_data.get('订单号', ''))
                    ]

                    for col_idx, value in enumerate(values):
                        cell_item = QTableWidgetItem(value)
                        cell_item.setTextAlignment(Qt.AlignCenter)
                        if not is_matched:
                            cell_item.setBackground(QColor(255, 237, 213))
                            if col_idx == 0:
                                cell_item.setForeground(QColor("#EF4444"))
                        else:
                            if col_idx == 0:
                                cell_item.setForeground(QColor("#10B981"))
                        self.query_table.setItem(row_idx, col_idx, cell_item)

                for col_idx in range(col_count):
                    self.query_table.resizeColumnToContents(col_idx)
                    if col_idx == 4:  # 物料描述列
                        self.query_table.setColumnWidth(col_idx, 300)

                # 更新统计标签
                unmatched_count = row_count - matched_count
                self.query_stats_label.setText(
                    f'共 {row_count} 行 | '
                    f'<span style="color: #10B981;">已匹配 {matched_count} 行</span> | '
                    f'<span style="color: #EF4444;">未匹配 {unmatched_count} 行</span>'
                )
                self.query_stats_label.setTextFormat(Qt.RichText)

                self.query_table.viewport().update()
                self.query_table.horizontalHeader().raise_()
                self.query_table.horizontalScrollBar().setValue(0)

                filename, _ = QFileDialog.getSaveFileName(
                    self, "保存查询结果", "", "Excel files (*.xlsx)"
                )
                if filename:
                    self.excel_processor.export_to_excel(query_results, filename)
                    QMessageBox.information(self, "完成",
                        f"查询结果已导出到:\n{filename}\n\n"
                        f"共 {row_count} 行, 匹配 {matched_count}, 未匹配 {unmatched_count}")
            else:
                self.query_stats_label.setText("")
                QMessageBox.information(self, "提示", "没有查询到数据")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"查询失败: {str(e)}")

    def batch_import_excel(self):
        filenames, _ = QFileDialog.getOpenFileNames(
            self, "选择Excel文件", "", "Excel files (*.xlsx *.xls);;All Files (*)"
        )
        if not filenames:
            return

        # 判断是否有现有数据
        has_existing_data = (
            hasattr(self, 'batch_import_data') and
            len(self.batch_import_data) > 0
        )

        # 导入模式选择
        if has_existing_data:
            reply = QMessageBox.question(
                self, "导入模式选择",
                f"已存在 {len(self.batch_import_data)} 条数据\n\n"
                f"选择导入模式：\n"
                f"是(Yes)：追加 - 保留现有数据，添加新文件数据\n"
                f"否(No)：替换 - 清空现有数据，重新导入",
                QMessageBox.Yes | QMessageBox.No
            )

            if reply == QMessageBox.Yes:
                # 追加模式：保留现有数据
                import_mode = "追加"
            else:
                # 替换模式：清空现有数据
                import_mode = "替换"
                self.batch_import_data = []
                self.batch_import_headers = []
        else:
            # 首次导入，初始化数据
            import_mode = "首次导入"
            self.batch_import_data = []
            self.batch_import_headers = []

        new_data_count = 0
        new_headers = None

        for filepath in filenames:
            try:
                headers, data = self.excel_processor.read_excel(filepath)

                # 获取文件名作为订单名称（去除扩展名）
                order_name = os.path.splitext(os.path.basename(filepath))[0]

                # 从第一个文件获取列头（后续文件保持一致）
                if not self.batch_import_headers and headers:
                    self.batch_import_headers = list(headers)
                    # 在下单时间列后添加订单名称列
                    if '下单时间' in self.batch_import_headers:
                        time_index = self.batch_import_headers.index('下单时间')
                        self.batch_import_headers.insert(time_index + 1, '订单名称')
                    elif '下单时间' not in self.batch_import_headers:
                        # 如果没有下单时间列，直接在最后添加
                        self.batch_import_headers.append('订单名称')
                    new_headers = self.batch_import_headers
                    print(f"[DEBUG] 从文件 '{os.path.basename(filepath)}' 读取到列头: {self.batch_import_headers}")

                for row in data:
                    material_number = row.get('物料号', '') or row.get('物料编码', '')
                    if not material_number:
                        continue
                    # 保留完整的原始行数据，并添加勾选字段和订单名称
                    row_data = {'勾选': False, '订单名称': order_name}
                    row_data.update(row)
                    self.batch_import_data.append(row_data)
                    new_data_count += 1
            except Exception as e:
                QMessageBox.warning(self, "警告", f"导入文件 {filepath} 时出错:\n{str(e)}")
                continue

        self.batch_refresh_table()

        mode_text = f"{import_mode}模式"
        total_count = len(self.batch_import_data)

        if import_mode == "追加":
            self.batch_info_label.setText(f"{mode_text} - 新增 {new_data_count} 条，现有共 {total_count} 条")
            QMessageBox.information(
                self, "导入完成",
                f"已成功{import_mode}！\n"
                f"新增数据：{new_data_count} 条\n"
                f"总数据量：{total_count} 条"
            )
        else:
            self.batch_info_label.setText(f"已导入 {total_count} 条")
            QMessageBox.information(
                self, "导入完成",
                f"已成功导入 {total_count} 条数据！"
            )

        # 自动保存导入的数据
        self.save_batch_import_data()

    def batch_refresh_table(self):
        print(f"[DEBUG] batch_refresh_table() 开始执行")
        print(f"[DEBUG] 当前父容器: {self.batch_import_table.parent().__class__.__name__}")
        print(f"[DEBUG] 动态列头数量: {len(self.batch_import_headers)}")
        print(f"[DEBUG] 动态列头内容: {self.batch_import_headers}")

        # 根据Excel的原始列头动态设置表格
        if self.batch_import_headers:
            # 列数 = 1（勾选列）+ 原始Excel的列数
            total_cols = 1 + len(self.batch_import_headers)
            self.batch_import_table.setColumnCount(total_cols)

            # 设置表头：第一列为"勾选"，其余为Excel原始列头
            header_labels = ['勾选'] + self.batch_import_headers
            self.batch_import_table.setHorizontalHeaderLabels(header_labels)
            print(f"[DEBUG] 设置表格列数为: {total_cols}")
            print(f"[DEBUG] 设置表头: {header_labels}")
        else:
            # 如果没有导入数据，保持默认的单列（勾选列）
            self.batch_import_table.setColumnCount(1)
            self.batch_import_table.setHorizontalHeaderLabels(['勾选'])

        self.batch_import_table.setRowCount(0)
        for item in self.batch_import_data:
            row_idx = self.batch_import_table.rowCount()
            self.batch_import_table.insertRow(row_idx)

            # 第一列：勾选列
            checkbox = QCheckBox()
            checkbox.setChecked(item.get('勾选', False))
            checkbox.stateChanged.connect(lambda state, row=row_idx: self.batch_toggle_row(row, state))
            container = QWidget()
            layout = QHBoxLayout(container)
            layout.addWidget(checkbox)
            layout.setAlignment(Qt.AlignCenter)
            layout.setContentsMargins(0, 0, 0, 0)
            self.batch_import_table.setCellWidget(row_idx, 0, container)

            # 后续列：根据Excel原始列头动态填充数据
            for col_idx, header_name in enumerate(self.batch_import_headers):
                value = item.get(header_name, '')
                if value is None:
                    value = ''
                # 截断过长的文本
                display_value = str(value)[:50] if len(str(value)) > 50 else str(value)
                self.batch_import_table.setItem(row_idx, col_idx + 1, QTableWidgetItem(display_value))

        # 智能调整列宽
        total_cols = self.batch_import_table.columnCount()
        for col_idx in range(total_cols):
            if col_idx == 0:
                # 勾选列固定宽度
                self.batch_import_table.setColumnWidth(col_idx, 50)
            else:
                # 其他列自适应内容
                self.batch_import_table.resizeColumnToContents(col_idx)
                w = self.batch_import_table.columnWidth(col_idx)
                if w < 80:
                    self.batch_import_table.setColumnWidth(col_idx, 80)  # 最小宽度
                elif w > 300:
                    self.batch_import_table.setColumnWidth(col_idx, 300)  # 最大宽度

        print(f"[DEBUG] 表格刷新完成: {self.batch_import_table.rowCount()} 行 x {self.batch_import_table.columnCount()} 列")

    def batch_toggle_row(self, row, state):
        if 0 <= row < len(self.batch_import_data):
            self.batch_import_data[row]['勾选'] = (state == Qt.Checked)

    def batch_select_all_rows(self):
        for item in self.batch_import_data:
            item['勾选'] = True
        self.batch_refresh_table()

    def batch_deselect_all_rows(self):
        for item in self.batch_import_data:
            item['勾选'] = False
        self.batch_refresh_table()

    def batch_export_selected(self):
        selected = [item for item in self.batch_import_data if item['勾选']]
        if not selected:
            QMessageBox.warning(self, "提示", "请先选择要导出的数据")
            return

        # 按照列头顺序导出，确保下单时间在左，订单名称在右
        export_data = []
        for item in selected:
            row_dict = {}
            # 按照列头顺序导出所有列（排除'勾选'字段）
            for header in self.batch_import_headers:
                if header in item:
                    row_dict[header] = item[header]
            export_data.append(row_dict)

        # 生成带北京时间戳的文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"批量导出_{timestamp}.xlsx"
        default_path = os.path.join(self.output_dir if hasattr(self, 'output_dir') else get_app_dir(), default_filename)

        filename, _ = QFileDialog.getSaveFileName(
            self, "保存选中数据", default_path, "Excel files (*.xlsx)"
        )
        if filename:
            self.excel_processor.export_to_excel(export_data, filename)
            QMessageBox.information(self, "完成", f"已导出 {len(export_data)} 条记录到：\n{filename}")

    def export_batch_data_for_query(self):
        """数据查询页面：导出批量导入的数据为出货Excel格式（与数据库管理页面一致）"""
        if not hasattr(self, 'batch_import_data') or not self.batch_import_data:
            QMessageBox.warning(self, "提示", "没有数据可导出")
            return

        # 检查是否有勾选的行
        selected = [item for item in self.batch_import_data if item['勾选']]

        # 如果有勾选的行，只导出勾选的；否则导出所有数据
        if selected:
            export_data = []
            for item in selected:
                row_dict = {}
                # 按照列头顺序导出所有列（排除'勾选'字段）
                for header in self.batch_import_headers:
                    if header in item:
                        row_dict[header] = item[header]
                export_data.append(row_dict)
        else:
            # 导出所有数据
            export_data = []
            for item in self.batch_import_data:
                row_dict = {}
                for header in self.batch_import_headers:
                    if header in item:
                        row_dict[header] = item[header]
                export_data.append(row_dict)

        # 生成文件名：导出出货Excel_北京时间（与数据库管理页面一致）
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"导出出货Excel_{timestamp}.xlsx"
        default_path = os.path.join(
            self.output_dir if hasattr(self, 'output_dir') else get_app_dir(),
            default_filename
        )

        filename, _ = QFileDialog.getSaveFileName(
            self, "导出出货数据", default_path, "Excel files (*.xlsx)"
        )
        if filename:
            self.excel_processor.export_to_excel(export_data, filename)
            QMessageBox.information(self, "完成", f"已导出 {len(export_data)} 条数据到:\n{filename}")

    def delete_query_matched(self):
        """数据查询页面：删除已显示的匹配记录或选中的行"""
        if not hasattr(self, 'batch_import_data') or not self.batch_import_data:
            QMessageBox.warning(self, "提示", "没有可删除的数据")
            return

        checked_indices = [idx for idx, item in enumerate(self.batch_import_data) if item.get('勾选', False)]

        if checked_indices:
            checked_count = len(checked_indices)

            reply = QMessageBox.question(self, "确认删除",
                f"确定要删除这 {checked_count} 条选中的记录吗？\n\n此操作不可撤销！",
                QMessageBox.Yes | QMessageBox.No)

            if reply == QMessageBox.Yes:
                self.batch_import_data = [row for idx, row in enumerate(self.batch_import_data)
                                          if idx not in checked_indices]
                self.batch_refresh_table()
                self.save_batch_import_data()
                QMessageBox.information(self, "完成", f"已删除 {checked_count} 条记录")
                return

        if not hasattr(self, 'query_shipment_matched_indices') or not self.query_shipment_matched_indices:
            QMessageBox.warning(self, "提示", "请先勾选要删除的行\n或者导入出货Excel进行匹配对比")
            return

        matched_count = len(self.query_shipment_matched_indices)

        reply = QMessageBox.question(self, "确认删除",
            f"确定要删除这 {matched_count} 条匹配记录吗？\n\n此操作不可撤销！",
            QMessageBox.Yes | QMessageBox.No)

        if reply == QMessageBox.Yes:
            self.batch_import_data = [row for idx, row in enumerate(self.batch_import_data)
                                      if idx not in self.query_shipment_matched_indices]

            self.batch_refresh_table()
            self.save_batch_import_data()

            self.query_shipment_matched_indices = []

            if hasattr(self, 'batch_info_label'):
                self.batch_info_label.setText(f"已导入 {len(self.batch_import_data)} 条数据")

            QMessageBox.information(self, "完成", f"已删除 {matched_count} 条匹配记录")

    def import_shipment_to_compare_for_query(self):
        """数据查询页面专用：导入出货Excel表格，与批量导入数据对比并显示高亮"""
        if not hasattr(self, 'batch_import_data') or not self.batch_import_data:
            QMessageBox.warning(self, "提示", "请先导入Excel数据")
            return

        filename, _ = QFileDialog.getOpenFileName(
            self, "选择出货Excel文件", "", "Excel files (*.xlsx *.xls)"
        )
        if not filename:
            return

        try:
            shipment_headers, shipment_data = self.excel_processor.read_excel(filename)

            if not shipment_data:
                QMessageBox.warning(self, "提示", "出货Excel文件中没有数据")
                return

            # 查找出货表格中的关键列
            col_names = {
                'material': None,
                'sales_order': None,
                'sales_line': None,
                'internal_order': None
            }
            for h in shipment_headers:
                if h:
                    h_str = str(h)
                    if '物料号' in h_str or '物料编码' in h_str:
                        col_names['material'] = h
                    elif '销售订单' in h_str and '行号' not in h_str:
                        col_names['sales_order'] = h
                    elif '销售订单行号' in h_str:
                        col_names['sales_line'] = h
                    elif '内需单号' in h_str:
                        col_names['internal_order'] = h

            # 构建出货数据的匹配键列表
            shipment_keys = []
            for row in shipment_data:
                material_val = row.get(col_names['material'])
                material = str(material_val if material_val is not None else '').strip()
                if material:
                    sales_order_val = row.get(col_names['sales_order'])
                    sales_order = str(sales_order_val if sales_order_val is not None else '').strip()
                    sales_line_val = row.get(col_names['sales_line'])
                    sales_line = str(sales_line_val if sales_line_val is not None else '').strip()
                    internal_order_val = row.get(col_names['internal_order'])
                    internal_order = str(internal_order_val if internal_order_val is not None else '').strip()
                    # 构建完整匹配键
                    shipment_keys.append({
                        'material': material,
                        'sales_order': sales_order,
                        'sales_line': sales_line,
                        'internal_order': internal_order
                    })

            if not shipment_keys:
                QMessageBox.warning(self, "提示", "出货Excel文件中没有找到有效的物料号")
                return

            # 获取批量导入数据的关键列名
            headers = self.batch_import_headers
            batch_col_names = {
                'material': None,
                'sales_order': None,
                'sales_line': None,
                'internal_order': None
            }
            for h in headers:
                if h:
                    h_str = str(h)
                    if '物料号' in h_str or '物料编码' in h_str:
                        batch_col_names['material'] = h
                    elif '销售订单' in h_str and '行号' not in h_str:
                        batch_col_names['sales_order'] = h
                    elif '销售订单行号' in h_str:
                        batch_col_names['sales_line'] = h
                    elif '内需单号' in h_str:
                        batch_col_names['internal_order'] = h

            # 在批量导入数据中查找匹配的记录索引
            matched_indices = []
            for idx, row in enumerate(self.batch_import_data):
                batch_material = str(row.get(batch_col_names['material'], '') or '').strip()
                batch_sales_order = str(row.get(batch_col_names['sales_order'], '') or '').strip()
                batch_sales_line = str(row.get(batch_col_names['sales_line'], '') or '').strip()
                batch_internal_order = str(row.get(batch_col_names['internal_order'], '') or '').strip()

                # 检查是否匹配任一出货键
                for ship_key in shipment_keys:
                    # 必须物料号匹配
                    if batch_material.lower() != ship_key['material'].lower():
                        continue
                    # 检查销售订单（如果出货数据中有值）
                    if ship_key['sales_order'] and batch_sales_order.lower() != ship_key['sales_order'].lower():
                        continue
                    # 检查销售订单行号（如果出货数据中有值）
                    if ship_key['sales_line'] and batch_sales_line.lower() != ship_key['sales_line'].lower():
                        continue
                    # 检查内需单号（如果出货数据中有值）
                    if ship_key['internal_order'] and batch_internal_order.lower() != ship_key['internal_order'].lower():
                        continue
                    # 所有条件满足
                    matched_indices.append(idx)
                    break

            if not matched_indices:
                QMessageBox.information(self, "提示", f"出货表格中的 {len(shipment_keys)} 个物料号在当前数据中未找到匹配项")
                return

            # 保存匹配的索引用于后续操作
            self.query_shipment_matched_indices = matched_indices

            # 在批量导入表格中高亮显示匹配结果（红色高亮）
            self.display_batch_shipment_matched(matched_indices)

            # 更新标签
            self.batch_info_label.setText(f"出货对比模式 - 共 {len(matched_indices)} 条匹配记录（红色高亮显示）")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"处理失败: {str(e)}")

    def display_batch_shipment_matched(self, indices):
        """在批量导入表格中高亮显示出货匹配的记录，匹配项排在顶部，红色高亮"""
        red_fill = QColor(255, 100, 100)  # 红色高亮
        headers = self.batch_import_headers
        data = self.batch_import_data
        matched_set = set(indices)

        col_count = len(headers)

        # 将数据分为匹配和非匹配两组，匹配项放前面
        matched_data = [row for idx, row in enumerate(data) if idx in matched_set]
        unmatched_data = [row for idx, row in enumerate(data) if idx not in matched_set]
        display_data = matched_data + unmatched_data

        row_count = len(display_data)
        self.batch_import_table.setRowCount(row_count)

        for row_idx in range(row_count):
            row_data = display_data[row_idx]
            is_matched = row_idx < len(matched_data)  # 前len(matched_data)行是匹配的
            # 第一列：复选框使用QCheckBox居中显示
            checkbox_widget = QWidget()
            checkbox_layout = QHBoxLayout(checkbox_widget)
            checkbox_layout.setContentsMargins(0, 0, 0, 0)
            checkbox_layout.setAlignment(Qt.AlignCenter)
            checkbox = QCheckBox()
            checkbox.setChecked(is_matched)
            if is_matched:
                checkbox.setStyleSheet("QCheckBox { margin-left: 15px; background-color: rgb(255, 100, 100); }")
            else:
                checkbox.setStyleSheet("QCheckBox { margin-left: 15px; }")
            checkbox_layout.addWidget(checkbox)
            self.batch_import_table.setCellWidget(row_idx, 0, checkbox_widget)

            # 数据列
            for col_idx in range(col_count):
                key = headers[col_idx]
                value = row_data.get(key, '') if isinstance(row_data, dict) else ''
                if value is None:
                    value = ''
                item = QTableWidgetItem(str(value))
                # 匹配的记录红色高亮显示
                if is_matched:
                    item.setBackground(red_fill)
                self.batch_import_table.setItem(row_idx, col_idx + 1, item)

        # 保存红色高亮行的原始索引
        self.batch_red_highlight_indices = set(indices)

        # 自动调整列宽
        for col_idx in range(col_count + 1):
            if col_idx == 0:
                self.batch_import_table.setColumnWidth(col_idx, 50)
            else:
                self.batch_import_table.resizeColumnToContents(col_idx)
                if col_idx > 0 and headers[col_idx - 1] == '物料描述':
                    self.batch_import_table.setColumnWidth(col_idx, 300)

    def filter_batch_data(self, text):
        """搜索过滤批量导入数据"""
        if not hasattr(self, 'batch_import_data') or not self.batch_import_data:
            return

        keywords = [k.strip() for k in text.split(',') if k.strip()]

        if not keywords:
            # 没有搜索关键词，显示所有数据
            data = self.batch_import_data
            self.batch_search_keywords = []
            self.batch_search_data = data
            self.batch_refresh_table_with_data(data, [])
            self.batch_info_label.setText(f"显示 {len(data)} / {len(self.batch_import_data)} 条数据")
            return

        self.batch_search_keywords = keywords

        # 构建物料号到行数据的查找表（精确匹配）
        material_lookup = {}
        for row in self.batch_import_data:
            material_number = str(row.get('物料号', '') or row.get('物料编码', '')).strip()
            if material_number:
                if material_number not in material_lookup:
                    material_lookup[material_number] = []
                material_lookup[material_number].append(row)

        # 遍历所有搜索条件，构建结果列表
        headers = self.batch_import_headers
        data = []
        for kw in keywords:
            if kw in material_lookup:
                for row in material_lookup[kw]:
                    row_copy = dict(row)
                    row_copy['查询状态'] = '已匹配'
                    data.append(row_copy)
            else:
                # 未匹配：优先使用搜索条件Excel的原始行数据
                if hasattr(self, 'batch_search_source_rows') and kw in self.batch_search_source_rows:
                    src_row = self.batch_search_source_rows[kw]
                    unmatched_row = {h: '' for h in headers}
                    # 将搜索条件Excel的列数据映射过来
                    for src_h, src_v in src_row.items():
                        if src_h in unmatched_row:
                            unmatched_row[src_h] = src_v
                    unmatched_row['查询状态'] = '未匹配'
                    data.append(unmatched_row)
                else:
                    empty_row = {h: '' for h in headers}
                    empty_row['物料号'] = kw
                    empty_row['查询状态'] = '未匹配'
                    data.append(empty_row)

        self.batch_search_data = data
        self.batch_refresh_table_with_data(data, keywords)

        matched_count = sum(1 for r in data if r.get('查询状态') == '已匹配')
        self.batch_info_label.setText(
            f"共 {len(keywords)} 个搜索条件 | 匹配 {matched_count} | 未匹配 {len(keywords) - matched_count}")

    def batch_refresh_table_with_data(self, data, keywords=None):
        """使用指定数据刷新表格，可选的高亮显示"""
        if not data:
            self.batch_import_table.setRowCount(0)
            return

        headers = self.batch_import_headers
        col_count = len(headers)

        # 判断是否处于搜索状态
        has_keywords = keywords is not None and len(keywords) > 0
        # 判断是否有查询状态列（从filter_batch_data传入）
        has_status_col = has_keywords and any('查询状态' in row for row in data)

        # 计算实际列数：checkbox + [查询状态] + 数据列
        if has_status_col:
            total_cols = 1 + 1 + col_count  # checkbox + 查询状态 + 数据列
        else:
            total_cols = 1 + col_count  # checkbox + 数据列
        self.batch_import_table.setColumnCount(total_cols)

        # 设置表头
        if has_status_col:
            status_header = QTableWidgetItem("查询状态")
            status_header.setForeground(QColor("#FFFFFF"))
            status_header.setTextAlignment(Qt.AlignCenter)
            self.batch_import_table.setHorizontalHeaderItem(1, status_header)
            for i, h in enumerate(headers):
                header_item = QTableWidgetItem(str(h))
                header_item.setForeground(QColor("#FFFFFF"))
                header_item.setTextAlignment(Qt.AlignCenter)
                self.batch_import_table.setHorizontalHeaderItem(i + 2, header_item)
        else:
            for i, h in enumerate(headers):
                header_item = QTableWidgetItem(str(h))
                header_item.setForeground(QColor("#FFFFFF"))
                header_item.setTextAlignment(Qt.AlignCenter)
                self.batch_import_table.setHorizontalHeaderItem(i + 1, header_item)

        self.batch_import_table.setRowCount(len(data))

        for row_idx, row_data in enumerate(data):
            # 第一列：勾选列
            checkbox = QCheckBox()
            checkbox.setChecked(row_data.get('勾选', False))
            checkbox.stateChanged.connect(lambda state, row=row_idx: self.batch_toggle_row(row, state))
            container = QWidget()
            layout = QHBoxLayout(container)
            layout.addWidget(checkbox)
            layout.setAlignment(Qt.AlignCenter)
            layout.setContentsMargins(0, 0, 0, 0)
            self.batch_import_table.setCellWidget(row_idx, 0, container)

            if has_status_col:
                status = row_data.get('查询状态', '')
                is_matched = (status == '已匹配')

                # 第二列：查询状态
                status_item = QTableWidgetItem(status)
                status_item.setTextAlignment(Qt.AlignCenter)
                if is_matched:
                    status_item.setForeground(QColor("#10B981"))
                else:
                    status_item.setForeground(QColor("#EF4444"))
                    status_item.setBackground(QColor(255, 237, 213))
                self.batch_import_table.setItem(row_idx, 1, status_item)

                # 数据列（从第3列开始）
                for col_idx in range(col_count):
                    key = headers[col_idx]
                    value = row_data.get(key, '')
                    if value is None:
                        value = ''
                    display_value = str(value)[:50] if len(str(value)) > 50 else str(value)
                    item = QTableWidgetItem(display_value)
                    item.setTextAlignment(Qt.AlignCenter)
                    if not is_matched:
                        item.setBackground(QColor(255, 237, 213))
                    self.batch_import_table.setItem(row_idx, col_idx + 2, item)
            else:
                # 无状态列：普通模式
                highlight_color = QColor(173, 216, 230)
                needs_highlight = has_keywords
                for col_idx in range(col_count):
                    key = headers[col_idx]
                    value = row_data.get(key, '')
                    if value is None:
                        value = ''
                    display_value = str(value)[:50] if len(str(value)) > 50 else str(value)
                    item = QTableWidgetItem(display_value)
                    if needs_highlight:
                        item.setBackground(highlight_color)
                    self.batch_import_table.setItem(row_idx, col_idx + 1, item)

    def import_batch_search_conditions(self):
        """从Excel文件导入搜索条件"""
        filename, _ = QFileDialog.getOpenFileName(
            self, "选择搜索条件文件", "", "Excel files (*.xlsx *.xls)"
        )
        if not filename:
            return

        try:
            headers, data = self.excel_processor.read_excel(filename)

            # 查找物料号列
            material_col = None
            for h in headers:
                if '物料号' in str(h) or '物料编码' in str(h):
                    material_col = h
                    break

            if not material_col:
                QMessageBox.warning(self, "提示", "文件中没有找到物料号列")
                return

            # 提取所有物料号作为搜索条件，并保存原始行数据
            keywords = []
            source_rows = {}  # key=物料号, value=原始行dict
            for row in data:
                material = row.get(material_col, '')
                if material:
                    material_str = str(material).strip()
                    keywords.append(material_str)
                    source_rows[material_str] = row
            self.batch_search_source_rows = source_rows
            self.batch_search_source_headers = headers

            if keywords:
                self.batch_search_keywords = keywords
                self.batch_search_edit.setText(','.join(keywords))
                self.filter_batch_data(','.join(keywords))
                matched = len(self.batch_search_data) if hasattr(self, 'batch_search_data') else 0
                QMessageBox.information(self, "完成",
                    f"已导入 {len(keywords)} 个搜索条件\n"
                    f"匹配 {matched} 项，未匹配 {len(keywords) - matched} 项")
            else:
                QMessageBox.warning(self, "提示", "文件中没有找到有效的物料号")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"导入失败: {str(e)}")

    def export_batch_search_result(self):
        """导出搜索结果"""
        # 如果有搜索结果，导出搜索结果；否则导出所有数据
        if hasattr(self, 'batch_search_data') and self.batch_search_data:
            export_data = []
            has_status = any('查询状态' in row for row in self.batch_search_data)
            for row in self.batch_search_data:
                row_dict = {}
                if has_status:
                    row_dict['查询状态'] = row.get('查询状态', '')
                # 按照列头顺序导出所有列（排除'勾选'字段）
                for header in self.batch_import_headers:
                    if header in row:
                        row_dict[header] = row[header]
                export_data.append(row_dict)

            if not export_data:
                QMessageBox.warning(self, "提示", "没有可导出的数据")
                return

            # 生成带北京时间戳的文件名
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            default_filename = f"批量导出搜索结果_{timestamp}.xlsx"
            default_path = os.path.join(
                self.output_dir if hasattr(self, 'output_dir') else get_app_dir(),
                default_filename
            )

            filename, _ = QFileDialog.getSaveFileName(
                self, "保存搜索结果", default_path, "Excel files (*.xlsx)"
            )
            if filename:
                self.excel_processor.export_to_excel(export_data, filename)
                QMessageBox.information(self, "完成", f"已导出 {len(export_data)} 条搜索结果到：\n{filename}")
        else:
            QMessageBox.warning(self, "提示", "没有搜索结果可导出")

    def clear_batch_search(self):
        """清空搜索"""
        self.batch_search_edit.setText('')
        self.batch_search_source_rows = {}
        self.batch_search_source_headers = []
        self.filter_batch_data('')

    def batch_delete_selected(self):
        selected_count = sum(1 for item in self.batch_import_data if item['勾选'])
        if selected_count == 0:
            QMessageBox.warning(self, "提示", "请先选择要删除的数据")
            return

        reply = QMessageBox.question(self, "确认删除", f"确定要删除选中的 {selected_count} 条记录吗？",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.batch_import_data = [item for item in self.batch_import_data if not item['勾选']]
            self.batch_refresh_table()
            self.batch_info_label.setText(f"已导入 {len(self.batch_import_data)} 条")

            # 自动保存数据
            self.save_batch_import_data()

    def clear_batch_table(self):
        """清空数据查询页面的所有数据，与数据库页面clear_db_table逻辑一致"""
        reply = QMessageBox.question(self, "确认清空",
            "确定要清空所有数据吗？\n\n此操作不可撤销！",
            QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        # 安全检查：确保query_table对象仍然有效
        if self.is_widget_valid(getattr(self, 'query_table', None)):
            try:
                self.query_table.setRowCount(0)
                self.query_table.setColumnCount(0)
                self.query_table.setHorizontalHeaderLabels([])
            except Exception as e:
                print(f"[WARNING] 清空query_table时遇到问题: {e}")

        # 清空数据变量
        if hasattr(self, 'batch_import_data'):
            self.batch_import_data = []
        if hasattr(self, 'batch_import_headers'):
            self.batch_import_headers = []
        if hasattr(self, 'batch_search_data'):
            self.batch_search_data = []

        # 清空出货对比状态
        if hasattr(self, 'query_shipment_matched_indices'):
            self.query_shipment_matched_indices = []

        # 清空搜索框
        if self.is_widget_valid(getattr(self, 'batch_search_edit', None)):
            try:
                self.batch_search_edit.setText('')
            except Exception as e:
                print(f"[WARNING] 清空搜索框时遇到问题: {e}")

        # 更新状态标签
        if self.is_widget_valid(getattr(self, 'batch_info_label', None)):
            try:
                self.batch_info_label.setText("请导入Excel数据")
            except Exception as e:
                print(f"[WARNING] 更新标签时遇到问题: {e}")

        # 保存空的批量导入数据
        self.save_batch_import_data()

        QMessageBox.information(self, "完成", "已清空所有数据")

    def batch_import_single_excel(self):
        filename, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", "", "Excel files (*.xlsx *.xls);;All Files (*)"
        )
        if not filename:
            return

        try:
            headers, data = self.excel_processor.read_excel(filename)

            # 获取文件名作为订单名称（去除扩展名）
            order_name = os.path.splitext(os.path.basename(filename))[0]

            # 保存原始列头，并在下单时间列后添加订单名称列
            self.batch_import_headers = list(headers)
            if '下单时间' in self.batch_import_headers:
                time_index = self.batch_import_headers.index('下单时间')
                self.batch_import_headers.insert(time_index + 1, '订单名称')
            elif '下单时间' not in self.batch_import_headers:
                self.batch_import_headers.append('订单名称')
            print(f"[DEBUG] 从文件 '{os.path.basename(filename)}' 读取到列头: {self.batch_import_headers}")

            self.batch_import_data = []
            for row in data:
                material_number = row.get('物料号', '') or row.get('物料编码', '')
                if not material_number:
                    continue
                # 保留完整的原始行数据，并添加勾选字段和订单名称
                row_data = {'勾选': False, '订单名称': order_name}
                row_data.update(row)  # 将所有原始数据合并进来
                self.batch_import_data.append(row_data)

            self.batch_refresh_table()
            self.batch_info_label.setText(f"已导入 {len(self.batch_import_data)} 条")
            QMessageBox.information(self, "完成", f"已导入 {len(self.batch_import_data)} 条记录")

            # 自动保存导入的数据
            self.save_batch_import_data()

        except Exception as e:
            QMessageBox.critical(self, "错误", f"导入失败: {str(e)}")

    def export_by_category(self):
        output_dir = self.category_output_edit.text()
        if not output_dir:
            QMessageBox.critical(self, "错误", "请选择输出目录")
            return

        try:
            data = self.db.get_level1_all()

            if not data:
                QMessageBox.warning(self, "提示", "数据库中没有数据")
                return

            output_files = self.excel_processor.export_category_files(data, output_dir)

            result_msg = f"已导出 {len(output_files)} 个分类文件:\n\n"
            for category, filepath in output_files:
                result_msg += f"{category}: {filepath}\n"

            QMessageBox.information(self, "完成", result_msg)

        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")

    def calculate_pending_delivery(self):
        supply_path = self.pending_supply_edit.text()
        delivery_path = self.pending_delivery_edit.text()

        if not supply_path or not delivery_path:
            QMessageBox.critical(self, "错误", "请选择供货计划和送货单明细文件")
            return

        try:
            import openpyxl

            # 读取供货计划
            wb_supply = openpyxl.load_workbook(supply_path, read_only=True, data_only=True)
            ws_supply = wb_supply.active
            supply_headers = [cell.value for cell in ws_supply[1]]
            supply_data = []
            for row in ws_supply.iter_rows(min_row=2, values_only=True):
                if any(v is not None for v in row):
                    supply_data.append(dict(zip(supply_headers, row)))
            wb_supply.close()

            # 读取送货单
            wb_delivery = openpyxl.load_workbook(delivery_path, read_only=True, data_only=True)
            ws_delivery = wb_delivery.active
            delivery_headers = [cell.value for cell in ws_delivery[1]]
            delivery_data = []
            for row in ws_delivery.iter_rows(min_row=2, values_only=True):
                if any(v is not None for v in row):
                    delivery_data.append(dict(zip(delivery_headers, row)))
            wb_delivery.close()

            # 按物料编码汇总已送货数量
            delivered_dict = {}
            for row in delivery_data:
                material = row.get('物料编码', '') or ''
                if material:
                    qty = row.get('收货数量', 0) or 0
                    delivered_dict[material] = delivered_dict.get(material, 0) + qty

            # 合并数据：供货计划 + 已送货，计算还需交货
            pending_list = []
            for row in supply_data:
                material = row.get('物料号', '') or ''
                if not material:
                    continue
                total = row.get('总缺料', 0) or 0
                delivered = delivered_dict.get(material, 0)
                remaining = total - delivered
                if remaining > 0:
                    description = row.get('物料描述', '') or ''
                    in_transit = row.get('供应商在途量', 0) or 0
                    committed = row.get('已承诺量', 0) or 0
                    pending_list.append({
                        '物料号': material,
                        '物料描述': description,
                        '总缺料': total,
                        '已送货': delivered,
                        '还需交货': remaining,
                        '供应商在途量': in_transit,
                        '已承诺量': committed,
                    })

            # 按还需交货量降序排序
            pending_list.sort(key=lambda x: x['还需交货'], reverse=True)

            # 渲染到表格
            self.pending_table.setRowCount(0)
            self.pending_data = []
            for item in pending_list:
                row_idx = self.pending_table.rowCount()
                self.pending_table.insertRow(row_idx)
                self.pending_table.setItem(row_idx, 0, QTableWidgetItem(str(item['物料号'])))
                self.pending_table.setItem(row_idx, 1, QTableWidgetItem(str(item['物料描述'])[:50]))
                self.pending_table.setItem(row_idx, 2, QTableWidgetItem(str(item['总缺料'])))
                self.pending_table.setItem(row_idx, 3, QTableWidgetItem(str(item['已送货'])))
                self.pending_table.setItem(row_idx, 4, QTableWidgetItem(str(item['还需交货'])))
                self.pending_table.setItem(row_idx, 5, QTableWidgetItem(str(item['供应商在途量'])))
                self.pending_table.setItem(row_idx, 6, QTableWidgetItem(str(item['已承诺量'])))
                self.pending_data.append(item)

            QMessageBox.information(self, "完成", f"还需交货: {len(self.pending_data)} 条记录")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"计算失败: {str(e)}")

    def export_pending_delivery(self):
        if not self.pending_data:
            QMessageBox.warning(self, "提示", "没有可导出的数据，请先点击计算")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "保存还需交货数据", "", "Excel files (*.xlsx)"
        )
        if filename:
            self.excel_processor.export_to_excel(self.pending_data, filename)
            QMessageBox.information(self, "完成", f"还需交货数据已导出到:\n{filename}")

    def select_all_rows(self):
        """全选所有行"""
        for row_idx in range(self.db_table.rowCount()):
            widget = self.db_table.cellWidget(row_idx, 0)
            if widget:
                checkbox = widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(True)
            else:
                item = self.db_table.item(row_idx, 0)
                if item:
                    item.setCheckState(Qt.Checked)

    def deselect_all_rows(self):
        """取消全选所有行"""
        for row_idx in range(self.db_table.rowCount()):
            widget = self.db_table.cellWidget(row_idx, 0)
            if widget:
                checkbox = widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(False)
            else:
                item = self.db_table.item(row_idx, 0)
                if item:
                    item.setCheckState(Qt.Unchecked)

    def get_selected_rows_indices(self):
        """获取所有被勾选的行的索引"""
        selected_indices = []
        for row_idx in range(self.db_table.rowCount()):
            widget = self.db_table.cellWidget(row_idx, 0)
            if widget:
                checkbox = widget.findChild(QCheckBox)
                if checkbox and checkbox.isChecked():
                    selected_indices.append(row_idx)
            else:
                item = self.db_table.item(row_idx, 0)
                if item and item.checkState() == Qt.Checked:
                    selected_indices.append(row_idx)
        return selected_indices

    def delete_selected_rows(self):
        """删除选中的行（通过复选框选择的）"""
        checked_indices = self.get_selected_rows_indices()

        if not checked_indices:
            QMessageBox.warning(self, "提示", "请先勾选要删除的行")
            return

        checked_count = len(checked_indices)
        reply = QMessageBox.question(self, "确认删除",
            f"确定要删除这 {checked_count} 条选中的记录吗？\n\n此操作不可撤销！",
            QMessageBox.Yes | QMessageBox.No)

        if reply == QMessageBox.Yes:
            self.db_all_data = [row for idx, row in enumerate(self.db_all_data)
                                if idx not in checked_indices]
            self.refresh_db_table()
            self.save_db_data()
            QMessageBox.information(self, "完成", f"已删除 {checked_count} 条记录")

    def closeEvent(self, event):
        """关闭窗口事件 - 最小化到托盘而不是退出"""
        # 隐藏窗口
        self.hide()

        # 确保托盘图标可见
        if hasattr(self, 'tray_icon') and self.tray_icon:
            if not self.tray_icon.isVisible():
                self.tray_icon.show()

            # 显示托盘通知
            self.tray_icon.showMessage(
                f"TCL表格比对系统 v{__version__}",
                "程序已最小化到托盘，右键点击托盘图标可以退出程序",
                QSystemTrayIcon.Information,
                2000
            )

        # 忽略关闭事件，只隐藏窗口
        event.ignore()

    def auto_start_server(self):
        """自动启动服务器（开机自启模式）"""
        try:
            from server_db import ServerDatabase
            db_path = os.path.join(get_app_dir(), 'tcl_server_data.db')
            db = ServerDatabase(db_path)
            port = int(db.load_config('server_port', 5000))

            self._server_thread = ServerThread(host='0.0.0.0', port=port)
            self._server_thread.started.connect(lambda: self._on_auto_server_started(port))
            self._server_thread.stopped.connect(self._on_auto_server_stopped)
            self._server_thread.error.connect(self._on_auto_server_error)
            self._server_thread.start()
        except Exception as e:
            print(f"[ERROR] 自动启动服务器失败: {e}")

    def _on_auto_server_started(self, port):
        """自动启动服务器成功"""
        self._is_server_running = True
        self.update_network_status()
        print(f"[OK] 服务器已自动启动，端口: {port}")

    def _on_auto_server_stopped(self):
        """自动启动的服务器已停止"""
        self._is_server_running = False
        self._server_thread = None
        self.update_network_status()

    def _on_auto_server_error(self, error_msg):
        """自动启动服务器失败"""
        print(f"[ERROR] 服务器自动启动失败: {error_msg}")

    def auto_connect_client(self):
        """自动连接到服务器（开机自启客户端模式）"""
        try:
            if not self.network.is_server_mode or not self.network.server_url:
                print("[WARN] 客户端自动连接: 未配置服务器地址，跳过")
                return

            print(f"[INFO] 正在自动连接到服务器: {self.network.server_url}")
            success, msg = self.network.test_connection()
            if success:
                print(f"[OK] 自动连接成功: {msg}")
                self.update_network_status()
            else:
                print(f"[WARN] 自动连接失败: {msg}")
                self.network_mode_label.setText("● 连接失败（自动重试中...）")
                self.network_mode_label.setStyleSheet("""
                    QLabel {
                        color: #EF4444;
                        font-size: 11px;
                        font-weight: bold;
                        padding: 3px;
                    }
                """)
                # 30秒后重试，最多3次
                self._auto_connect_retry_count = getattr(self, '_auto_connect_retry_count', 0) + 1
                if self._auto_connect_retry_count <= 3:
                    QTimer.singleShot(30000, self.auto_connect_client)
                else:
                    print("[WARN] 自动连接重试次数已用完，请手动重连")
                    self.network_mode_label.setText("● 自动连接失败")
                    self._auto_connect_retry_count = 0
        except Exception as e:
            print(f"[ERROR] 自动连接失败: {e}")

    def open_server_settings(self):
        """打开服务器设置对话框"""
        dialog = ServerSettingsDialog(self)
        dialog.exec_()
        # 关闭对话框后获取服务器状态并更新显示
        self._server_thread = dialog.server_thread
        self._is_server_running = dialog.is_server_running
        self.update_network_status()

    def update_network_status(self):
        """更新侧边栏网络状态显示"""
        if self._is_server_running and self._server_thread and self._server_thread.isRunning():
            # 本机服务器运行中
            self.network_mode_label.setText("● 服务器运行中")
            self.network_mode_label.setStyleSheet("""
                QLabel {
                    color: #10B981;
                    font-size: 11px;
                    font-weight: bold;
                    padding: 3px;
                }
            """)
            # 显示监听地址
            try:
                import socket
                s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
                s.connect(("8.8.8.8", 80))
                local_ip = s.getsockname()[0]
                s.close()
                port = self._server_thread.port if hasattr(self._server_thread, 'port') else 5000
                self.server_address_label.setText(f"http://{local_ip}:{port}")
            except:
                self.server_address_label.setText("http://127.0.0.1:5000")
        elif self.network.is_server_mode:
            # 连接到远程服务器
            self.network_mode_label.setText("● 服务器模式")
            self.network_mode_label.setStyleSheet("""
                QLabel {
                    color: #3B82F6;
                    font-size: 11px;
                    font-weight: bold;
                    padding: 3px;
                }
            """)
            self.server_address_label.setText(self.network.server_url)
        else:
            # 本地模式
            self.network_mode_label.setText("● 本地模式")
            self.network_mode_label.setStyleSheet("""
                QLabel {
                    color: #10B981;
                    font-size: 11px;
                    font-weight: bold;
                    padding: 3px;
                }
            """)
            self.server_address_label.setText("")

    # ==================== 数据保存/加载 - 支持双模式 ====================

    def _save_with_mode(self, headers, data, server_save, local_save, fallback_message):
        if self.network.is_server_mode:
            success, msg = server_save(headers, data)
            if not success:
                QMessageBox.warning(self, "服务器错误", fallback_message.format(msg=msg))
                local_save(headers, data)
        else:
            local_save(headers, data)

    def _load_with_mode(self, server_load, local_load, fallback_message):
        if self.network.is_server_mode:
            headers, data, error = server_load()
            if error:
                QMessageBox.warning(self, "服务器错误", fallback_message.format(error=error))
                return local_load()
            return headers, data
        return local_load()

    def save_db_data_with_mode(self, headers, data):
        """根据当前模式保存数据库数据（本地或服务器）"""
        self._save_with_mode(
            headers,
            data,
            self.network.save_db_data,
            self._save_db_data_local,
            "保存到服务器失败:\n{msg}\n\n数据将尝试保存到本地"
        )

    def load_db_data_with_mode(self):
        """根据当前模式加载数据库数据（本地或服务器）"""
        return self._load_with_mode(
            self.network.load_db_data,
            self._load_db_data_local,
            "从服务器加载数据失败:\n{error}\n\n将加载本地数据"
        )

    def save_batch_import_data_with_mode(self, headers, data):
        """根据当前模式保存批量导入数据（本地或服务器）"""
        self._save_with_mode(
            headers,
            data,
            self.network.save_batch_import_data,
            self._save_batch_import_data_local,
            "保存到服务器失败:\n{msg}"
        )

    def load_batch_import_data_with_mode(self):
        """根据当前模式加载批量导入数据（本地或服务器）"""
        return self._load_with_mode(
            self.network.load_batch_import_data,
            self._load_batch_import_data_local,
            "从服务器加载数据失败:\n{error}"
        )


if __name__ == "__main__":
    import sys
    from PyQt5.QtWidgets import QApplication

    app = QApplication(sys.argv)
    window = TCLApplication()
    window.show()
    sys.exit(app.exec_())